import os
import io
import re
import base64
import logging
import ipaddress
from flask import Flask, jsonify, request, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine, func, or_
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
try:
    from api.invoice import generate_invoice_pdf, _collect_line_items
except ImportError:
    from invoice import generate_invoice_pdf, _collect_line_items
try:
    from api.chat import chat_with_ollama
except ImportError:
    from chat import chat_with_ollama

logging.basicConfig(level=logging.DEBUG)

# Serve React build in production (static_folder points to React build output)
static_folder = os.environ.get('STATIC_FOLDER',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'webapp', 'customer-app', 'build'))
if os.path.isdir(static_folder):
    app = Flask(__name__, static_folder=static_folder, static_url_path='')
else:
    app = Flask(__name__)

# ===========================================================================
# NETWORK ACCESS CONTROL — Local network only
# ===========================================================================
# Private/local IP ranges (RFC 1918 + loopback + link-local)
LOCAL_NETWORKS = [
    ipaddress.ip_network('127.0.0.0/8'),       # Loopback
    ipaddress.ip_network('10.0.0.0/8'),         # Class A private
    ipaddress.ip_network('172.16.0.0/12'),      # Class B private
    ipaddress.ip_network('192.168.0.0/16'),     # Class C private
    ipaddress.ip_network('169.254.0.0/16'),     # Link-local
    ipaddress.ip_network('::1/128'),            # IPv6 loopback
    ipaddress.ip_network('fe80::/10'),          # IPv6 link-local
    ipaddress.ip_network('fc00::/7'),           # IPv6 unique local
]

# Additional allowed IPs/CIDRs from env (comma-separated), e.g. "203.0.113.5,198.51.100.0/24"
extra_allowed = os.environ.get('ALLOWED_IPS', '')
if extra_allowed:
    for cidr in extra_allowed.split(','):
        cidr = cidr.strip()
        if cidr:
            try:
                LOCAL_NETWORKS.append(ipaddress.ip_network(cidr, strict=False))
            except ValueError:
                logging.warning(f"Invalid ALLOWED_IPS entry: {cidr}")

# LAN_ONLY can be set to "false" to disable the restriction (e.g. during development)
LAN_ONLY = os.environ.get('LAN_ONLY', 'true').lower() != 'false'

# SMTP Configuration
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.office365.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USERNAME = os.environ.get('SMTP_USERNAME', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM = os.environ.get('SMTP_FROM', 'clientsupport@njgroups.com')
SMTP_USE_TLS = os.environ.get('SMTP_USE_TLS', 'true').lower() != 'false'


def is_local_ip(ip_str):
    """Check if an IP address belongs to a local/private network."""
    try:
        addr = ipaddress.ip_address(ip_str)
        return any(addr in network for network in LOCAL_NETWORKS)
    except ValueError:
        return False


@app.before_request
def restrict_to_local_network():
    """Block requests from non-local IP addresses."""
    if not LAN_ONLY:
        return None

    client_ip = request.remote_addr
    if not is_local_ip(client_ip):
        logging.warning(f"Blocked request from non-local IP: {client_ip}")
        abort(403)


# ===========================================================================
# CORS CONFIGURATION
# ===========================================================================
def is_allowed_origin(origin):
    """Check if origin is allowed."""
    allowed_origins = [
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:5001",
        "http://127.0.0.1:5001"
    ]
    # Add origins from ALLOWED_ORIGINS env var (comma-separated)
    extra = os.environ.get('ALLOWED_ORIGINS', '')
    if extra:
        allowed_origins.extend([o.strip() for o in extra.split(',') if o.strip()])
    if origin in allowed_origins:
        return True
    # Allow private network origins (http://192.168.x.x:port, http://10.x.x.x:port, etc.)
    if origin:
        try:
            # Extract hostname from origin (e.g. "http://192.168.1.50:5000" -> "192.168.1.50")
            host = origin.split('://')[1].split(':')[0] if '://' in origin else origin.split(':')[0]
            if is_local_ip(host):
                return True
        except (IndexError, ValueError):
            pass
    return False

@app.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    if origin and is_allowed_origin(origin):
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS'
        response.headers['Access-Control-Allow-Credentials'] = 'true'
    return response

# ===========================================================================
# DATABASE CONFIGURATION
# ===========================================================================
db_uri = os.environ.get('DATABASE_URI', 'postgresql://localhost/client_portal')

app.config['SQLALCHEMY_DATABASE_URI'] = db_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'])
db = SQLAlchemy(app)
Session = sessionmaker(bind=engine)
logging.info(f"Database URI: {db_uri}")

# ===========================================================================
# DATABASE MODELS
# ===========================================================================

class Client(db.Model):
    __tablename__ = 'clients'

    id = db.Column(db.Integer, primary_key=True)
    tax_id = db.Column(db.String(50), unique=True, nullable=False)
    client_name = db.Column(db.String(200))
    dba = db.Column(db.String(200))
    contact_person = db.Column(db.String(200))
    email = db.Column(db.String(200))
    phone_number = db.Column(db.String(50))
    address_line_1 = db.Column(db.String(200))
    address_line_2 = db.Column(db.String(200))
    city = db.Column(db.String(100))
    state = db.Column(db.String(50))
    zip_code = db.Column(db.String(20))
    status = db.Column(db.String(50), default='Active')
    gross_revenue = db.Column(db.Numeric(15, 2))
    total_ees = db.Column(db.Integer)
    industry = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    employee_benefits = db.relationship('EmployeeBenefit', back_populates='client', cascade='all, delete-orphan')
    commercial_insurance = db.relationship('CommercialInsurance', back_populates='client', cascade='all, delete-orphan')
    contacts = db.relationship('ClientContact', back_populates='client', cascade='all, delete-orphan', order_by='ClientContact.sort_order')

    def to_dict(self):
        contacts_list = [c.to_dict() for c in self.contacts] if self.contacts else []
        # For backward compatibility, populate flat contact fields from first contact
        first_contact = contacts_list[0] if contacts_list else {}
        return {
            'id': self.id,
            'tax_id': self.tax_id,
            'client_name': self.client_name,
            'dba': self.dba,
            'industry': self.industry,
            'contact_person': first_contact.get('contact_person', self.contact_person or ''),
            'email': first_contact.get('email', self.email or ''),
            'phone_number': first_contact.get('phone_number', self.phone_number or ''),
            'phone_extension': first_contact.get('phone_extension', ''),
            'address_line_1': first_contact.get('address_line_1', self.address_line_1 or ''),
            'address_line_2': first_contact.get('address_line_2', self.address_line_2 or ''),
            'city': first_contact.get('city', self.city or ''),
            'state': first_contact.get('state', self.state or ''),
            'zip_code': first_contact.get('zip_code', self.zip_code or ''),
            'status': self.status,
            'gross_revenue': float(self.gross_revenue) if self.gross_revenue else None,
            'total_ees': self.total_ees,
            'has_employee_benefits': len(self.employee_benefits) > 0,
            'has_commercial_insurance': len(self.commercial_insurance) > 0,
            'contacts': contacts_list
        }


class ClientContact(db.Model):
    __tablename__ = 'client_contacts'

    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('clients.id'), nullable=False)
    contact_person = db.Column(db.String(200))
    email = db.Column(db.String(200))
    phone_number = db.Column(db.String(50))
    phone_extension = db.Column(db.String(20))
    address_line_1 = db.Column(db.String(200))
    address_line_2 = db.Column(db.String(200))
    city = db.Column(db.String(100))
    state = db.Column(db.String(50))
    zip_code = db.Column(db.String(20))
    sort_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    client = db.relationship('Client', back_populates='contacts')

    def to_dict(self):
        return {
            'id': self.id,
            'contact_person': self.contact_person or '',
            'email': self.email or '',
            'phone_number': self.phone_number or '',
            'phone_extension': self.phone_extension or '',
            'address_line_1': self.address_line_1 or '',
            'address_line_2': self.address_line_2 or '',
            'city': self.city or '',
            'state': self.state or '',
            'zip_code': self.zip_code or '',
            'sort_order': self.sort_order
        }


class Individual(db.Model):
    __tablename__ = 'individuals'

    id = db.Column(db.Integer, primary_key=True)
    individual_id = db.Column(db.String(50), unique=True, nullable=False)
    first_name = db.Column(db.String(200))
    last_name = db.Column(db.String(200))
    email = db.Column(db.String(200))
    phone_number = db.Column(db.String(50))
    address_line_1 = db.Column(db.String(200))
    address_line_2 = db.Column(db.String(200))
    city = db.Column(db.String(100))
    state = db.Column(db.String(50))
    zip_code = db.Column(db.String(20))
    status = db.Column(db.String(50), default='Active')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    personal_insurance = db.relationship('PersonalInsurance', back_populates='individual', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'individual_id': self.individual_id,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'full_name': f"{self.first_name or ''} {self.last_name or ''}".strip(),
            'email': self.email,
            'phone_number': self.phone_number,
            'address_line_1': self.address_line_1,
            'address_line_2': self.address_line_2,
            'city': self.city,
            'state': self.state,
            'zip_code': self.zip_code,
            'status': self.status
        }


class EmployeeBenefit(db.Model):
    __tablename__ = 'employee_benefits'

    id = db.Column(db.Integer, primary_key=True)
    tax_id = db.Column(db.String(50), db.ForeignKey('clients.tax_id'), nullable=False)
    parent_client = db.Column(db.String(200))

    # Core fields
    status = db.Column(db.String(50))
    outstanding_item = db.Column(db.String(50))
    remarks = db.Column(db.Text)
    form_fire_code = db.Column(db.String(100))
    enrollment_poc = db.Column(db.String(200))
    renewal_date = db.Column(db.Date)
    funding = db.Column(db.String(100))
    current_carrier = db.Column(db.String(200))
    num_employees_at_renewal = db.Column(db.Integer)
    enrolled_ees = db.Column(db.Integer)
    waiting_period = db.Column(db.String(100))
    deductible_accumulation = db.Column(db.String(100))
    previous_carrier = db.Column(db.String(200))
    cobra_carrier = db.Column(db.String(200))
    employer_contribution = db.Column(db.String(50))
    employee_contribution = db.Column(db.String(50))

    # Dental
    dental_renewal_date = db.Column(db.Date)
    dental_carrier = db.Column(db.String(200))

    # Vision
    vision_renewal_date = db.Column(db.Date)
    vision_carrier = db.Column(db.String(200))

    # Life & AD&D
    life_adnd_renewal_date = db.Column(db.Date)
    life_adnd_carrier = db.Column(db.String(200))

    # LTD
    ltd_renewal_date = db.Column(db.Date)
    ltd_carrier = db.Column(db.String(200))

    # STD
    std_renewal_date = db.Column(db.Date)
    std_carrier = db.Column(db.String(200))

    # 401K
    k401_renewal_date = db.Column(db.Date)
    k401_carrier = db.Column(db.String(200))

    # Critical Illness
    critical_illness_renewal_date = db.Column(db.Date)
    critical_illness_carrier = db.Column(db.String(200))

    # Accident
    accident_renewal_date = db.Column(db.Date)
    accident_carrier = db.Column(db.String(200))

    # Hospital
    hospital_renewal_date = db.Column(db.Date)
    hospital_carrier = db.Column(db.String(200))

    # Voluntary Life
    voluntary_life_renewal_date = db.Column(db.Date)
    voluntary_life_carrier = db.Column(db.String(200))

    # Flag columns for single-plan types (deprecated, kept for backward compat)
    ltd_flag = db.Column(db.Boolean, default=False)
    std_flag = db.Column(db.Boolean, default=False)
    k401_flag = db.Column(db.Boolean, default=False)
    critical_illness_flag = db.Column(db.Boolean, default=False)
    accident_flag = db.Column(db.Boolean, default=False)
    hospital_flag = db.Column(db.Boolean, default=False)
    voluntary_life_flag = db.Column(db.Boolean, default=False)

    # Remarks columns for single-plan types
    ltd_remarks = db.Column(db.Text)
    std_remarks = db.Column(db.Text)
    k401_remarks = db.Column(db.Text)
    critical_illness_remarks = db.Column(db.Text)
    accident_remarks = db.Column(db.Text)
    hospital_remarks = db.Column(db.Text)
    voluntary_life_remarks = db.Column(db.Text)

    # Outstanding item columns for single-plan types
    outstanding_item_due_date = db.Column(db.Date)
    ltd_outstanding_item = db.Column(db.Text)
    ltd_outstanding_item_due_date = db.Column(db.Date)
    std_outstanding_item = db.Column(db.Text)
    std_outstanding_item_due_date = db.Column(db.Date)
    k401_outstanding_item = db.Column(db.Text)
    k401_outstanding_item_due_date = db.Column(db.Date)
    critical_illness_outstanding_item = db.Column(db.Text)
    critical_illness_outstanding_item_due_date = db.Column(db.Date)
    accident_outstanding_item = db.Column(db.Text)
    accident_outstanding_item_due_date = db.Column(db.Date)
    hospital_outstanding_item = db.Column(db.Text)
    hospital_outstanding_item_due_date = db.Column(db.Date)
    voluntary_life_outstanding_item = db.Column(db.Text)
    voluntary_life_outstanding_item_due_date = db.Column(db.Date)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    client = db.relationship('Client', back_populates='employee_benefits')
    plans = db.relationship('BenefitPlan', back_populates='employee_benefit', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'tax_id': self.tax_id,
            'client_name': self.client.client_name if self.client else None,
            'client_status': self.client.status if self.client else None,
            'parent_client': self.parent_client,
            'form_fire_code': self.form_fire_code,
            'enrollment_poc': self.enrollment_poc,
            'renewal_date': self.renewal_date.isoformat() if self.renewal_date else None,
            'outstanding_item_due_date': self.outstanding_item_due_date.isoformat() if self.outstanding_item_due_date else None,
            'funding': self.funding,
            'current_carrier': self.current_carrier,
            'num_employees_at_renewal': self.num_employees_at_renewal,
            'enrolled_ees': self.enrolled_ees,
            'waiting_period': self.waiting_period,
            'deductible_accumulation': self.deductible_accumulation,
            'previous_carrier': self.previous_carrier,
            'cobra_carrier': self.cobra_carrier,
            'employee_contribution': self.employee_contribution,
            'dental_renewal_date': self.dental_renewal_date.isoformat() if self.dental_renewal_date else None,
            'dental_carrier': self.dental_carrier,
            'vision_renewal_date': self.vision_renewal_date.isoformat() if self.vision_renewal_date else None,
            'vision_carrier': self.vision_carrier,
            'life_adnd_renewal_date': self.life_adnd_renewal_date.isoformat() if self.life_adnd_renewal_date else None,
            'life_adnd_carrier': self.life_adnd_carrier,
            'ltd_renewal_date': self.ltd_renewal_date.isoformat() if self.ltd_renewal_date else None,
            'ltd_carrier': self.ltd_carrier,
            'ltd_remarks': self.ltd_remarks,
            'ltd_outstanding_item': self.ltd_outstanding_item,
            'ltd_outstanding_item_due_date': self.ltd_outstanding_item_due_date.isoformat() if self.ltd_outstanding_item_due_date else None,
            'std_renewal_date': self.std_renewal_date.isoformat() if self.std_renewal_date else None,
            'std_carrier': self.std_carrier,
            'std_remarks': self.std_remarks,
            'std_outstanding_item': self.std_outstanding_item,
            'std_outstanding_item_due_date': self.std_outstanding_item_due_date.isoformat() if self.std_outstanding_item_due_date else None,
            'k401_renewal_date': self.k401_renewal_date.isoformat() if self.k401_renewal_date else None,
            'k401_carrier': self.k401_carrier,
            'k401_remarks': self.k401_remarks,
            'k401_outstanding_item': self.k401_outstanding_item,
            'k401_outstanding_item_due_date': self.k401_outstanding_item_due_date.isoformat() if self.k401_outstanding_item_due_date else None,
            'critical_illness_renewal_date': self.critical_illness_renewal_date.isoformat() if self.critical_illness_renewal_date else None,
            'critical_illness_carrier': self.critical_illness_carrier,
            'critical_illness_remarks': self.critical_illness_remarks,
            'critical_illness_outstanding_item': self.critical_illness_outstanding_item,
            'critical_illness_outstanding_item_due_date': self.critical_illness_outstanding_item_due_date.isoformat() if self.critical_illness_outstanding_item_due_date else None,
            'accident_renewal_date': self.accident_renewal_date.isoformat() if self.accident_renewal_date else None,
            'accident_carrier': self.accident_carrier,
            'accident_remarks': self.accident_remarks,
            'accident_outstanding_item': self.accident_outstanding_item,
            'accident_outstanding_item_due_date': self.accident_outstanding_item_due_date.isoformat() if self.accident_outstanding_item_due_date else None,
            'hospital_renewal_date': self.hospital_renewal_date.isoformat() if self.hospital_renewal_date else None,
            'hospital_carrier': self.hospital_carrier,
            'hospital_remarks': self.hospital_remarks,
            'hospital_outstanding_item': self.hospital_outstanding_item,
            'hospital_outstanding_item_due_date': self.hospital_outstanding_item_due_date.isoformat() if self.hospital_outstanding_item_due_date else None,
            'voluntary_life_renewal_date': self.voluntary_life_renewal_date.isoformat() if self.voluntary_life_renewal_date else None,
            'voluntary_life_carrier': self.voluntary_life_carrier,
            'voluntary_life_remarks': self.voluntary_life_remarks,
            'voluntary_life_outstanding_item': self.voluntary_life_outstanding_item,
            'voluntary_life_outstanding_item_due_date': self.voluntary_life_outstanding_item_due_date.isoformat() if self.voluntary_life_outstanding_item_due_date else None,
            'plans': self._get_plans_dict()
        }

    def _get_plans_dict(self):
        """Group child BenefitPlan records by type."""
        plans_dict = {'medical': [], 'dental': [], 'vision': [], 'life_adnd': []}
        for plan in (self.plans or []):
            if plan.plan_type in plans_dict:
                plans_dict[plan.plan_type].append(plan.to_dict())
        for pt in plans_dict:
            plans_dict[pt].sort(key=lambda p: p['plan_number'])
        return plans_dict


MULTI_PLAN_TYPES = ['medical', 'dental', 'vision', 'life_adnd']


class BenefitPlan(db.Model):
    __tablename__ = 'benefit_plans'

    id = db.Column(db.Integer, primary_key=True)
    employee_benefit_id = db.Column(db.Integer, db.ForeignKey('employee_benefits.id'), nullable=False)
    plan_type = db.Column(db.String(50), nullable=False)
    plan_number = db.Column(db.Integer, nullable=False, default=1)
    carrier = db.Column(db.String(200))
    renewal_date = db.Column(db.Date)
    flag = db.Column(db.Boolean, default=False)
    waiting_period = db.Column(db.String(100))
    remarks = db.Column(db.Text)
    outstanding_item = db.Column(db.String(50))

    # Relationship
    employee_benefit = db.relationship('EmployeeBenefit', back_populates='plans')

    def to_dict(self):
        return {
            'id': self.id,
            'plan_type': self.plan_type,
            'plan_number': self.plan_number,
            'carrier': self.carrier,
            'renewal_date': self.renewal_date.isoformat() if self.renewal_date else None,
            'waiting_period': self.waiting_period,
            'remarks': self.remarks,
            'outstanding_item': self.outstanding_item
        }


class CommercialInsurance(db.Model):
    __tablename__ = 'commercial_insurance'

    id = db.Column(db.Integer, primary_key=True)
    tax_id = db.Column(db.String(50), db.ForeignKey('clients.tax_id'), nullable=False)

    parent_client = db.Column(db.String(200))
    assigned_to = db.Column(db.String(200))

    # Core fields
    remarks = db.Column(db.Text)
    status = db.Column(db.String(50))
    outstanding_item = db.Column(db.String(50))

    # 1. Commercial General Liability
    general_liability_carrier = db.Column(db.String(200))
    general_liability_agency = db.Column(db.String(200))
    general_liability_policy_number = db.Column(db.String(100))
    general_liability_occ_limit = db.Column(db.String(100))
    general_liability_agg_limit = db.Column(db.String(100))
    general_liability_premium = db.Column(db.Numeric(12, 2))
    general_liability_renewal_date = db.Column(db.Date)
    general_liability_endorsement_bop = db.Column(db.Boolean, default=False)
    general_liability_endorsement_marine = db.Column(db.Boolean, default=False)
    general_liability_endorsement_foreign = db.Column(db.Boolean, default=False)
    general_liability_endorsement_molestation = db.Column(db.Boolean, default=False)
    general_liability_endorsement_staffing = db.Column(db.Boolean, default=False)
    general_liability_endorsement_accidental_medical = db.Column(db.Boolean, default=False)
    general_liability_endorsement_liquor_liability = db.Column(db.Boolean, default=False)

    # 2. Commercial Property
    property_carrier = db.Column(db.String(200))
    property_agency = db.Column(db.String(200))
    property_policy_number = db.Column(db.String(100))
    property_occ_limit = db.Column(db.String(100))
    property_agg_limit = db.Column(db.String(100))
    property_premium = db.Column(db.Numeric(12, 2))
    property_renewal_date = db.Column(db.Date)

    # 3. Business Owners Policy (BOP)
    bop_carrier = db.Column(db.String(200))
    bop_agency = db.Column(db.String(200))
    bop_policy_number = db.Column(db.String(100))
    bop_occ_limit = db.Column(db.String(100))
    bop_agg_limit = db.Column(db.String(100))
    bop_premium = db.Column(db.Numeric(12, 2))
    bop_renewal_date = db.Column(db.Date)
    bop_building_limit = db.Column(db.Numeric(15, 2))
    bop_personal_property = db.Column(db.Numeric(15, 2))

    # 4. Umbrella Liability
    umbrella_carrier = db.Column(db.String(200))
    umbrella_agency = db.Column(db.String(200))
    umbrella_policy_number = db.Column(db.String(100))
    umbrella_occ_limit = db.Column(db.String(100))
    umbrella_agg_limit = db.Column(db.String(100))
    umbrella_premium = db.Column(db.Numeric(12, 2))
    umbrella_renewal_date = db.Column(db.Date)

    # 5. Workers Compensation
    workers_comp_carrier = db.Column(db.String(200))
    workers_comp_agency = db.Column(db.String(200))
    workers_comp_policy_number = db.Column(db.String(100))
    workers_comp_occ_limit = db.Column(db.String(100))
    workers_comp_agg_limit = db.Column(db.String(100))
    workers_comp_premium = db.Column(db.Numeric(12, 2))
    workers_comp_renewal_date = db.Column(db.Date)

    # 6. Professional or E&O
    professional_eo_carrier = db.Column(db.String(200))
    professional_eo_agency = db.Column(db.String(200))
    professional_eo_policy_number = db.Column(db.String(100))
    professional_eo_occ_limit = db.Column(db.String(100))
    professional_eo_agg_limit = db.Column(db.String(100))
    professional_eo_premium = db.Column(db.Numeric(12, 2))
    professional_eo_renewal_date = db.Column(db.Date)

    # 7. Cyber Liability
    cyber_carrier = db.Column(db.String(200))
    cyber_agency = db.Column(db.String(200))
    cyber_policy_number = db.Column(db.String(100))
    cyber_occ_limit = db.Column(db.String(100))
    cyber_agg_limit = db.Column(db.String(100))
    cyber_premium = db.Column(db.Numeric(12, 2))
    cyber_renewal_date = db.Column(db.Date)

    # 8. Commercial Auto
    auto_carrier = db.Column(db.String(200))
    auto_agency = db.Column(db.String(200))
    auto_policy_number = db.Column(db.String(100))
    auto_occ_limit = db.Column(db.String(100))
    auto_agg_limit = db.Column(db.String(100))
    auto_premium = db.Column(db.Numeric(12, 2))
    auto_renewal_date = db.Column(db.Date)

    # 9. EPLI
    epli_carrier = db.Column(db.String(200))
    epli_agency = db.Column(db.String(200))
    epli_policy_number = db.Column(db.String(100))
    epli_occ_limit = db.Column(db.String(100))
    epli_agg_limit = db.Column(db.String(100))
    epli_premium = db.Column(db.Numeric(12, 2))
    epli_renewal_date = db.Column(db.Date)

    # 10. NYDBL
    nydbl_carrier = db.Column(db.String(200))
    nydbl_agency = db.Column(db.String(200))
    nydbl_policy_number = db.Column(db.String(100))
    nydbl_occ_limit = db.Column(db.String(100))
    nydbl_agg_limit = db.Column(db.String(100))
    nydbl_premium = db.Column(db.Numeric(12, 2))
    nydbl_renewal_date = db.Column(db.Date)

    # 11. Surety Bond
    surety_carrier = db.Column(db.String(200))
    surety_agency = db.Column(db.String(200))
    surety_policy_number = db.Column(db.String(100))
    surety_occ_limit = db.Column(db.String(100))
    surety_agg_limit = db.Column(db.String(100))
    surety_premium = db.Column(db.Numeric(12, 2))
    surety_renewal_date = db.Column(db.Date)

    # 12. Product Liability
    product_liability_carrier = db.Column(db.String(200))
    product_liability_agency = db.Column(db.String(200))
    product_liability_policy_number = db.Column(db.String(100))
    product_liability_occ_limit = db.Column(db.String(100))
    product_liability_agg_limit = db.Column(db.String(100))
    product_liability_premium = db.Column(db.Numeric(12, 2))
    product_liability_renewal_date = db.Column(db.Date)

    # 13. Flood
    flood_carrier = db.Column(db.String(200))
    flood_agency = db.Column(db.String(200))
    flood_policy_number = db.Column(db.String(100))
    flood_occ_limit = db.Column(db.String(100))
    flood_agg_limit = db.Column(db.String(100))
    flood_premium = db.Column(db.Numeric(12, 2))
    flood_renewal_date = db.Column(db.Date)

    # 14. Crime or Fidelity Bond
    crime_carrier = db.Column(db.String(200))
    crime_agency = db.Column(db.String(200))
    crime_policy_number = db.Column(db.String(100))
    crime_occ_limit = db.Column(db.String(100))
    crime_agg_limit = db.Column(db.String(100))
    crime_premium = db.Column(db.Numeric(12, 2))
    crime_renewal_date = db.Column(db.Date)

    # 15. Directors & Officers
    directors_officers_carrier = db.Column(db.String(200))
    directors_officers_agency = db.Column(db.String(200))
    directors_officers_policy_number = db.Column(db.String(100))
    directors_officers_occ_limit = db.Column(db.String(100))
    directors_officers_agg_limit = db.Column(db.String(100))
    directors_officers_premium = db.Column(db.Numeric(12, 2))
    directors_officers_renewal_date = db.Column(db.Date)

    # 16. Fiduciary Bond
    fiduciary_carrier = db.Column(db.String(200))
    fiduciary_agency = db.Column(db.String(200))
    fiduciary_policy_number = db.Column(db.String(100))
    fiduciary_occ_limit = db.Column(db.String(100))
    fiduciary_agg_limit = db.Column(db.String(100))
    fiduciary_premium = db.Column(db.Numeric(12, 2))
    fiduciary_renewal_date = db.Column(db.Date)

    # 17. Inland Marine
    inland_marine_carrier = db.Column(db.String(200))
    inland_marine_agency = db.Column(db.String(200))
    inland_marine_policy_number = db.Column(db.String(100))
    inland_marine_occ_limit = db.Column(db.String(100))
    inland_marine_agg_limit = db.Column(db.String(100))
    inland_marine_premium = db.Column(db.Numeric(12, 2))
    inland_marine_renewal_date = db.Column(db.Date)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Flag columns for single-plan types (deprecated, kept for backward compat)
    general_liability_flag = db.Column(db.Boolean, default=False)
    property_flag = db.Column(db.Boolean, default=False)
    bop_flag = db.Column(db.Boolean, default=False)
    workers_comp_flag = db.Column(db.Boolean, default=False)
    auto_flag = db.Column(db.Boolean, default=False)
    epli_flag = db.Column(db.Boolean, default=False)
    nydbl_flag = db.Column(db.Boolean, default=False)
    surety_flag = db.Column(db.Boolean, default=False)
    product_liability_flag = db.Column(db.Boolean, default=False)
    flood_flag = db.Column(db.Boolean, default=False)
    directors_officers_flag = db.Column(db.Boolean, default=False)
    fiduciary_flag = db.Column(db.Boolean, default=False)
    inland_marine_flag = db.Column(db.Boolean, default=False)

    # Remarks columns for single-plan types
    general_liability_remarks = db.Column(db.Text)
    property_remarks = db.Column(db.Text)
    bop_remarks = db.Column(db.Text)
    workers_comp_remarks = db.Column(db.Text)
    auto_remarks = db.Column(db.Text)
    epli_remarks = db.Column(db.Text)
    nydbl_remarks = db.Column(db.Text)
    surety_remarks = db.Column(db.Text)
    product_liability_remarks = db.Column(db.Text)
    flood_remarks = db.Column(db.Text)
    directors_officers_remarks = db.Column(db.Text)
    fiduciary_remarks = db.Column(db.Text)
    inland_marine_remarks = db.Column(db.Text)

    # Outstanding item columns for single-plan types
    general_liability_outstanding_item = db.Column(db.Text)
    property_outstanding_item = db.Column(db.Text)
    bop_outstanding_item = db.Column(db.Text)
    workers_comp_outstanding_item = db.Column(db.Text)
    auto_outstanding_item = db.Column(db.Text)
    epli_outstanding_item = db.Column(db.Text)
    nydbl_outstanding_item = db.Column(db.Text)
    surety_outstanding_item = db.Column(db.Text)
    product_liability_outstanding_item = db.Column(db.Text)
    flood_outstanding_item = db.Column(db.Text)
    directors_officers_outstanding_item = db.Column(db.Text)
    fiduciary_outstanding_item = db.Column(db.Text)
    inland_marine_outstanding_item = db.Column(db.Text)

    # Outstanding item due date columns
    general_liability_outstanding_item_due_date = db.Column(db.Date)
    property_outstanding_item_due_date = db.Column(db.Date)
    bop_outstanding_item_due_date = db.Column(db.Date)
    workers_comp_outstanding_item_due_date = db.Column(db.Date)
    auto_outstanding_item_due_date = db.Column(db.Date)
    epli_outstanding_item_due_date = db.Column(db.Date)
    nydbl_outstanding_item_due_date = db.Column(db.Date)
    surety_outstanding_item_due_date = db.Column(db.Date)
    product_liability_outstanding_item_due_date = db.Column(db.Date)
    flood_outstanding_item_due_date = db.Column(db.Date)
    directors_officers_outstanding_item_due_date = db.Column(db.Date)
    fiduciary_outstanding_item_due_date = db.Column(db.Date)
    inland_marine_outstanding_item_due_date = db.Column(db.Date)

    # Relationships
    client = db.relationship('Client', back_populates='commercial_insurance')
    commercial_plans = db.relationship('CommercialPlan', back_populates='commercial_insurance', cascade='all, delete-orphan')

    def to_dict(self):
        def format_premium(val):
            return float(val) if val else None

        result = {
            'id': self.id,
            'tax_id': self.tax_id,
            'client_name': self.client.client_name if self.client else None,
            'client_status': self.client.status if self.client else None,
            'parent_client': self.parent_client,
            'assigned_to': self.assigned_to,
            'general_liability_carrier': self.general_liability_carrier,
            'general_liability_agency': self.general_liability_agency,
            'general_liability_policy_number': self.general_liability_policy_number,
            'general_liability_occ_limit': self.general_liability_occ_limit,
            'general_liability_agg_limit': self.general_liability_agg_limit,
            'general_liability_premium': format_premium(self.general_liability_premium),
            'general_liability_renewal_date': self.general_liability_renewal_date.isoformat() if self.general_liability_renewal_date else None,
            'general_liability_remarks': self.general_liability_remarks,
            'general_liability_outstanding_item': self.general_liability_outstanding_item,
            'general_liability_outstanding_item_due_date': self.general_liability_outstanding_item_due_date.isoformat() if self.general_liability_outstanding_item_due_date else None,
            'general_liability_endorsement_bop': self.general_liability_endorsement_bop or False,
            'general_liability_endorsement_marine': self.general_liability_endorsement_marine or False,
            'general_liability_endorsement_foreign': self.general_liability_endorsement_foreign or False,
            'general_liability_endorsement_molestation': self.general_liability_endorsement_molestation or False,
            'general_liability_endorsement_staffing': self.general_liability_endorsement_staffing or False,
            'general_liability_endorsement_accidental_medical': self.general_liability_endorsement_accidental_medical or False,
            'general_liability_endorsement_liquor_liability': self.general_liability_endorsement_liquor_liability or False,
            'property_carrier': self.property_carrier,
            'property_agency': self.property_agency,
            'property_policy_number': self.property_policy_number,
            'property_occ_limit': self.property_occ_limit,
            'property_agg_limit': self.property_agg_limit,
            'property_premium': format_premium(self.property_premium),
            'property_renewal_date': self.property_renewal_date.isoformat() if self.property_renewal_date else None,
            'property_remarks': self.property_remarks,
            'property_outstanding_item': self.property_outstanding_item,
            'property_outstanding_item_due_date': self.property_outstanding_item_due_date.isoformat() if self.property_outstanding_item_due_date else None,
            'bop_carrier': self.bop_carrier,
            'bop_agency': self.bop_agency,
            'bop_policy_number': self.bop_policy_number,
            'bop_occ_limit': self.bop_occ_limit,
            'bop_agg_limit': self.bop_agg_limit,
            'bop_premium': format_premium(self.bop_premium),
            'bop_renewal_date': self.bop_renewal_date.isoformat() if self.bop_renewal_date else None,
            'bop_building_limit': format_premium(self.bop_building_limit),
            'bop_personal_property': format_premium(self.bop_personal_property),
            'bop_remarks': self.bop_remarks,
            'bop_outstanding_item': self.bop_outstanding_item,
            'bop_outstanding_item_due_date': self.bop_outstanding_item_due_date.isoformat() if self.bop_outstanding_item_due_date else None,
            'umbrella_carrier': self.umbrella_carrier,
            'umbrella_agency': self.umbrella_agency,
            'umbrella_policy_number': self.umbrella_policy_number,
            'umbrella_occ_limit': self.umbrella_occ_limit,
            'umbrella_agg_limit': self.umbrella_agg_limit,
            'umbrella_premium': format_premium(self.umbrella_premium),
            'umbrella_renewal_date': self.umbrella_renewal_date.isoformat() if self.umbrella_renewal_date else None,
            'workers_comp_carrier': self.workers_comp_carrier,
            'workers_comp_agency': self.workers_comp_agency,
            'workers_comp_policy_number': self.workers_comp_policy_number,
            'workers_comp_occ_limit': self.workers_comp_occ_limit,
            'workers_comp_agg_limit': self.workers_comp_agg_limit,
            'workers_comp_premium': format_premium(self.workers_comp_premium),
            'workers_comp_renewal_date': self.workers_comp_renewal_date.isoformat() if self.workers_comp_renewal_date else None,
            'workers_comp_remarks': self.workers_comp_remarks,
            'workers_comp_outstanding_item': self.workers_comp_outstanding_item,
            'workers_comp_outstanding_item_due_date': self.workers_comp_outstanding_item_due_date.isoformat() if self.workers_comp_outstanding_item_due_date else None,
            'professional_eo_carrier': self.professional_eo_carrier,
            'professional_eo_agency': self.professional_eo_agency,
            'professional_eo_policy_number': self.professional_eo_policy_number,
            'professional_eo_occ_limit': self.professional_eo_occ_limit,
            'professional_eo_agg_limit': self.professional_eo_agg_limit,
            'professional_eo_premium': format_premium(self.professional_eo_premium),
            'professional_eo_renewal_date': self.professional_eo_renewal_date.isoformat() if self.professional_eo_renewal_date else None,
            'cyber_carrier': self.cyber_carrier,
            'cyber_agency': self.cyber_agency,
            'cyber_policy_number': self.cyber_policy_number,
            'cyber_occ_limit': self.cyber_occ_limit,
            'cyber_agg_limit': self.cyber_agg_limit,
            'cyber_premium': format_premium(self.cyber_premium),
            'cyber_renewal_date': self.cyber_renewal_date.isoformat() if self.cyber_renewal_date else None,
            'auto_carrier': self.auto_carrier,
            'auto_agency': self.auto_agency,
            'auto_policy_number': self.auto_policy_number,
            'auto_occ_limit': self.auto_occ_limit,
            'auto_agg_limit': self.auto_agg_limit,
            'auto_premium': format_premium(self.auto_premium),
            'auto_renewal_date': self.auto_renewal_date.isoformat() if self.auto_renewal_date else None,
            'auto_remarks': self.auto_remarks,
            'auto_outstanding_item': self.auto_outstanding_item,
            'auto_outstanding_item_due_date': self.auto_outstanding_item_due_date.isoformat() if self.auto_outstanding_item_due_date else None,
            'epli_carrier': self.epli_carrier,
            'epli_agency': self.epli_agency,
            'epli_policy_number': self.epli_policy_number,
            'epli_occ_limit': self.epli_occ_limit,
            'epli_agg_limit': self.epli_agg_limit,
            'epli_premium': format_premium(self.epli_premium),
            'epli_renewal_date': self.epli_renewal_date.isoformat() if self.epli_renewal_date else None,
            'epli_remarks': self.epli_remarks,
            'epli_outstanding_item': self.epli_outstanding_item,
            'epli_outstanding_item_due_date': self.epli_outstanding_item_due_date.isoformat() if self.epli_outstanding_item_due_date else None,
            'nydbl_carrier': self.nydbl_carrier,
            'nydbl_agency': self.nydbl_agency,
            'nydbl_policy_number': self.nydbl_policy_number,
            'nydbl_occ_limit': self.nydbl_occ_limit,
            'nydbl_agg_limit': self.nydbl_agg_limit,
            'nydbl_premium': format_premium(self.nydbl_premium),
            'nydbl_renewal_date': self.nydbl_renewal_date.isoformat() if self.nydbl_renewal_date else None,
            'nydbl_remarks': self.nydbl_remarks,
            'nydbl_outstanding_item': self.nydbl_outstanding_item,
            'nydbl_outstanding_item_due_date': self.nydbl_outstanding_item_due_date.isoformat() if self.nydbl_outstanding_item_due_date else None,
            'surety_carrier': self.surety_carrier,
            'surety_agency': self.surety_agency,
            'surety_policy_number': self.surety_policy_number,
            'surety_occ_limit': self.surety_occ_limit,
            'surety_agg_limit': self.surety_agg_limit,
            'surety_premium': format_premium(self.surety_premium),
            'surety_renewal_date': self.surety_renewal_date.isoformat() if self.surety_renewal_date else None,
            'surety_remarks': self.surety_remarks,
            'surety_outstanding_item': self.surety_outstanding_item,
            'surety_outstanding_item_due_date': self.surety_outstanding_item_due_date.isoformat() if self.surety_outstanding_item_due_date else None,
            'product_liability_carrier': self.product_liability_carrier,
            'product_liability_agency': self.product_liability_agency,
            'product_liability_policy_number': self.product_liability_policy_number,
            'product_liability_occ_limit': self.product_liability_occ_limit,
            'product_liability_agg_limit': self.product_liability_agg_limit,
            'product_liability_premium': format_premium(self.product_liability_premium),
            'product_liability_renewal_date': self.product_liability_renewal_date.isoformat() if self.product_liability_renewal_date else None,
            'product_liability_remarks': self.product_liability_remarks,
            'product_liability_outstanding_item': self.product_liability_outstanding_item,
            'product_liability_outstanding_item_due_date': self.product_liability_outstanding_item_due_date.isoformat() if self.product_liability_outstanding_item_due_date else None,
            'flood_carrier': self.flood_carrier,
            'flood_agency': self.flood_agency,
            'flood_policy_number': self.flood_policy_number,
            'flood_occ_limit': self.flood_occ_limit,
            'flood_agg_limit': self.flood_agg_limit,
            'flood_premium': format_premium(self.flood_premium),
            'flood_renewal_date': self.flood_renewal_date.isoformat() if self.flood_renewal_date else None,
            'flood_remarks': self.flood_remarks,
            'flood_outstanding_item': self.flood_outstanding_item,
            'flood_outstanding_item_due_date': self.flood_outstanding_item_due_date.isoformat() if self.flood_outstanding_item_due_date else None,
            'crime_carrier': self.crime_carrier,
            'crime_agency': self.crime_agency,
            'crime_policy_number': self.crime_policy_number,
            'crime_occ_limit': self.crime_occ_limit,
            'crime_agg_limit': self.crime_agg_limit,
            'crime_premium': format_premium(self.crime_premium),
            'crime_renewal_date': self.crime_renewal_date.isoformat() if self.crime_renewal_date else None,
            'directors_officers_carrier': self.directors_officers_carrier,
            'directors_officers_agency': self.directors_officers_agency,
            'directors_officers_policy_number': self.directors_officers_policy_number,
            'directors_officers_occ_limit': self.directors_officers_occ_limit,
            'directors_officers_agg_limit': self.directors_officers_agg_limit,
            'directors_officers_premium': format_premium(self.directors_officers_premium),
            'directors_officers_renewal_date': self.directors_officers_renewal_date.isoformat() if self.directors_officers_renewal_date else None,
            'directors_officers_remarks': self.directors_officers_remarks,
            'directors_officers_outstanding_item': self.directors_officers_outstanding_item,
            'directors_officers_outstanding_item_due_date': self.directors_officers_outstanding_item_due_date.isoformat() if self.directors_officers_outstanding_item_due_date else None,
            'fiduciary_carrier': self.fiduciary_carrier,
            'fiduciary_agency': self.fiduciary_agency,
            'fiduciary_policy_number': self.fiduciary_policy_number,
            'fiduciary_occ_limit': self.fiduciary_occ_limit,
            'fiduciary_agg_limit': self.fiduciary_agg_limit,
            'fiduciary_premium': format_premium(self.fiduciary_premium),
            'fiduciary_renewal_date': self.fiduciary_renewal_date.isoformat() if self.fiduciary_renewal_date else None,
            'fiduciary_remarks': self.fiduciary_remarks,
            'fiduciary_outstanding_item': self.fiduciary_outstanding_item,
            'fiduciary_outstanding_item_due_date': self.fiduciary_outstanding_item_due_date.isoformat() if self.fiduciary_outstanding_item_due_date else None,
            'inland_marine_carrier': self.inland_marine_carrier,
            'inland_marine_agency': self.inland_marine_agency,
            'inland_marine_policy_number': self.inland_marine_policy_number,
            'inland_marine_occ_limit': self.inland_marine_occ_limit,
            'inland_marine_agg_limit': self.inland_marine_agg_limit,
            'inland_marine_premium': format_premium(self.inland_marine_premium),
            'inland_marine_renewal_date': self.inland_marine_renewal_date.isoformat() if self.inland_marine_renewal_date else None,
            'inland_marine_remarks': self.inland_marine_remarks,
            'inland_marine_outstanding_item': self.inland_marine_outstanding_item,
            'inland_marine_outstanding_item_due_date': self.inland_marine_outstanding_item_due_date.isoformat() if self.inland_marine_outstanding_item_due_date else None,
            'plans': self._get_commercial_plans_dict()
        }
        return result

    def _get_commercial_plans_dict(self):
        """Group child CommercialPlan records by type."""
        plans_dict = {'umbrella': [], 'professional_eo': [], 'cyber': [], 'crime': []}
        for plan in (self.commercial_plans or []):
            if plan.plan_type in plans_dict:
                plans_dict[plan.plan_type].append(plan.to_dict())
        for pt in plans_dict:
            plans_dict[pt].sort(key=lambda p: p['plan_number'])
        return plans_dict

MULTI_PLAN_COMMERCIAL_TYPES = ['umbrella', 'professional_eo', 'cyber', 'crime']


class CommercialPlan(db.Model):
    __tablename__ = 'commercial_plans'

    id = db.Column(db.Integer, primary_key=True)
    commercial_insurance_id = db.Column(db.Integer, db.ForeignKey('commercial_insurance.id'), nullable=False)
    plan_type = db.Column(db.String(50), nullable=False)
    plan_number = db.Column(db.Integer, nullable=False, default=1)
    carrier = db.Column(db.String(200))
    agency = db.Column(db.String(200))
    policy_number = db.Column(db.String(100))
    coverage_occ_limit = db.Column(db.String(100))
    coverage_agg_limit = db.Column(db.String(100))
    premium = db.Column(db.Numeric(12, 2))
    renewal_date = db.Column(db.Date)
    flag = db.Column(db.Boolean, default=False)
    remarks = db.Column(db.Text)
    outstanding_item = db.Column(db.Text)
    outstanding_item_due_date = db.Column(db.Date)

    # Endorsements (used by professional_eo plans)
    endorsement_tech_eo = db.Column(db.Boolean, default=False)
    endorsement_allied_healthcare = db.Column(db.Boolean, default=False)
    endorsement_staffing = db.Column(db.Boolean, default=False)
    endorsement_medical_malpractice = db.Column(db.Boolean, default=False)

    # Relationship
    commercial_insurance = db.relationship('CommercialInsurance', back_populates='commercial_plans')

    def to_dict(self):
        return {
            'id': self.id,
            'plan_type': self.plan_type,
            'plan_number': self.plan_number,
            'carrier': self.carrier,
            'agency': self.agency,
            'policy_number': self.policy_number,
            'occ_limit': self.coverage_occ_limit,
            'agg_limit': self.coverage_agg_limit,
            'premium': float(self.premium) if self.premium else None,
            'renewal_date': self.renewal_date.isoformat() if self.renewal_date else None,
            'remarks': self.remarks,
            'outstanding_item': self.outstanding_item,
            'outstanding_item_due_date': self.outstanding_item_due_date.isoformat() if self.outstanding_item_due_date else None,
            'endorsement_tech_eo': self.endorsement_tech_eo or False,
            'endorsement_allied_healthcare': self.endorsement_allied_healthcare or False,
            'endorsement_staffing': self.endorsement_staffing or False,
            'endorsement_medical_malpractice': self.endorsement_medical_malpractice or False
        }


class HomeownersPolicy(db.Model):
    __tablename__ = 'homeowners_policies'

    id = db.Column(db.Integer, primary_key=True)
    personal_insurance_id = db.Column(db.Integer, db.ForeignKey('personal_insurance.id'), nullable=False)
    policy_number = db.Column(db.Integer, nullable=False, default=1)
    carrier = db.Column(db.String(200))
    dwelling_limit = db.Column(db.String(100))
    liability_limit = db.Column(db.String(100))
    premium = db.Column(db.Numeric(12, 2))
    renewal_date = db.Column(db.Date)
    remarks = db.Column(db.Text)
    outstanding_item = db.Column(db.Text)
    outstanding_item_due_date = db.Column(db.Date)
    property_address_line_1 = db.Column(db.String(200))
    property_address_line_2 = db.Column(db.String(200))
    property_city = db.Column(db.String(100))
    property_state = db.Column(db.String(50))
    property_zip = db.Column(db.String(20))
    is_primary_residence = db.Column(db.Boolean, default=False)

    personal_insurance = db.relationship('PersonalInsurance', back_populates='homeowners_policies')

    def to_dict(self):
        return {
            'id': self.id,
            'policy_number': self.policy_number,
            'carrier': self.carrier,
            'dwelling_limit': self.dwelling_limit,
            'liability_limit': self.liability_limit,
            'premium': float(self.premium) if self.premium else None,
            'renewal_date': self.renewal_date.isoformat() if self.renewal_date else None,
            'remarks': self.remarks,
            'outstanding_item': self.outstanding_item,
            'outstanding_item_due_date': self.outstanding_item_due_date.isoformat() if self.outstanding_item_due_date else None,
            'property_address_line_1': self.property_address_line_1,
            'property_address_line_2': self.property_address_line_2,
            'property_city': self.property_city,
            'property_state': self.property_state,
            'property_zip': self.property_zip,
            'is_primary_residence': self.is_primary_residence or False
        }


class PersonalInsurance(db.Model):
    __tablename__ = 'personal_insurance'

    id = db.Column(db.Integer, primary_key=True)
    individual_id = db.Column(db.String(50), db.ForeignKey('individuals.individual_id'), nullable=False)

    # 1. Personal Auto
    personal_auto_carrier = db.Column(db.String(200))
    personal_auto_bi_occ_limit = db.Column(db.String(100))
    personal_auto_bi_agg_limit = db.Column(db.String(100))
    personal_auto_pd_limit = db.Column(db.String(100))
    personal_auto_renewal_date = db.Column(db.Date)
    personal_auto_premium = db.Column(db.Numeric(12, 2))
    personal_auto_outstanding_item = db.Column(db.String(50))
    personal_auto_remarks = db.Column(db.Text)

    # 2. Homeowners
    homeowners_carrier = db.Column(db.String(200))
    homeowners_dwelling_limit = db.Column(db.String(100))
    homeowners_liability_limit = db.Column(db.String(100))
    homeowners_renewal_date = db.Column(db.Date)
    homeowners_premium = db.Column(db.Numeric(12, 2))
    homeowners_outstanding_item = db.Column(db.String(50))
    homeowners_remarks = db.Column(db.Text)

    # 3. Personal Umbrella
    personal_umbrella_carrier = db.Column(db.String(200))
    personal_umbrella_liability_limit = db.Column(db.String(100))
    personal_umbrella_deductible = db.Column(db.Numeric(12, 2))
    personal_umbrella_renewal_date = db.Column(db.Date)
    personal_umbrella_premium = db.Column(db.Numeric(12, 2))
    personal_umbrella_outstanding_item = db.Column(db.String(50))
    personal_umbrella_remarks = db.Column(db.Text)

    # 4. Event Insurance
    event_carrier = db.Column(db.String(200))
    event_type = db.Column(db.String(200))
    event_location = db.Column(db.String(500))
    event_start_date = db.Column(db.Date)
    event_end_date = db.Column(db.Date)
    event_entry_fee = db.Column(db.Numeric(12, 2))
    event_audience_count = db.Column(db.Integer)
    event_premium = db.Column(db.Numeric(12, 2))
    event_outstanding_item = db.Column(db.String(50))
    event_remarks = db.Column(db.Text)

    # 5. Visitors Medical
    visitors_medical_carrier = db.Column(db.String(200))
    visitors_medical_start_date = db.Column(db.Date)
    visitors_medical_end_date = db.Column(db.Date)
    visitors_medical_destination_country = db.Column(db.String(200))
    visitors_medical_premium = db.Column(db.Numeric(12, 2))
    visitors_medical_outstanding_item = db.Column(db.String(50))
    visitors_medical_remarks = db.Column(db.Text)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    individual = db.relationship('Individual', back_populates='personal_insurance')
    homeowners_policies = db.relationship('HomeownersPolicy', back_populates='personal_insurance', cascade='all, delete-orphan')

    def to_dict(self):
        def format_premium(val):
            return float(val) if val else None

        return {
            'id': self.id,
            'individual_id': self.individual_id,
            'individual_name': f"{self.individual.first_name or ''} {self.individual.last_name or ''}".strip() if self.individual else None,
            'individual_status': self.individual.status if self.individual else None,
            # Personal Auto
            'personal_auto_carrier': self.personal_auto_carrier,
            'personal_auto_bi_occ_limit': self.personal_auto_bi_occ_limit,
            'personal_auto_bi_agg_limit': self.personal_auto_bi_agg_limit,
            'personal_auto_pd_limit': self.personal_auto_pd_limit,
            'personal_auto_renewal_date': self.personal_auto_renewal_date.isoformat() if self.personal_auto_renewal_date else None,
            'personal_auto_premium': format_premium(self.personal_auto_premium),
            'personal_auto_outstanding_item': self.personal_auto_outstanding_item,
            'personal_auto_remarks': self.personal_auto_remarks,
            # Homeowners
            'homeowners_carrier': self.homeowners_carrier,
            'homeowners_dwelling_limit': self.homeowners_dwelling_limit,
            'homeowners_liability_limit': self.homeowners_liability_limit,
            'homeowners_renewal_date': self.homeowners_renewal_date.isoformat() if self.homeowners_renewal_date else None,
            'homeowners_premium': format_premium(self.homeowners_premium),
            'homeowners_outstanding_item': self.homeowners_outstanding_item,
            'homeowners_remarks': self.homeowners_remarks,
            'homeowners_policies_list': [p.to_dict() for p in sorted(self.homeowners_policies, key=lambda x: x.policy_number)],
            # Personal Umbrella
            'personal_umbrella_carrier': self.personal_umbrella_carrier,
            'personal_umbrella_liability_limit': self.personal_umbrella_liability_limit,
            'personal_umbrella_deductible': format_premium(self.personal_umbrella_deductible),
            'personal_umbrella_renewal_date': self.personal_umbrella_renewal_date.isoformat() if self.personal_umbrella_renewal_date else None,
            'personal_umbrella_premium': format_premium(self.personal_umbrella_premium),
            'personal_umbrella_outstanding_item': self.personal_umbrella_outstanding_item,
            'personal_umbrella_remarks': self.personal_umbrella_remarks,
            # Event Insurance
            'event_carrier': self.event_carrier,
            'event_type': self.event_type,
            'event_location': self.event_location,
            'event_start_date': self.event_start_date.isoformat() if self.event_start_date else None,
            'event_end_date': self.event_end_date.isoformat() if self.event_end_date else None,
            'event_entry_fee': format_premium(self.event_entry_fee),
            'event_audience_count': self.event_audience_count,
            'event_premium': format_premium(self.event_premium),
            'event_outstanding_item': self.event_outstanding_item,
            'event_remarks': self.event_remarks,
            # Visitors Medical
            'visitors_medical_carrier': self.visitors_medical_carrier,
            'visitors_medical_start_date': self.visitors_medical_start_date.isoformat() if self.visitors_medical_start_date else None,
            'visitors_medical_end_date': self.visitors_medical_end_date.isoformat() if self.visitors_medical_end_date else None,
            'visitors_medical_destination_country': self.visitors_medical_destination_country,
            'visitors_medical_premium': format_premium(self.visitors_medical_premium),
            'visitors_medical_outstanding_item': self.visitors_medical_outstanding_item,
            'visitors_medical_remarks': self.visitors_medical_remarks,
        }


PERSONAL_INSURANCE_PRODUCTS = [
    'personal_auto', 'homeowners', 'personal_umbrella', 'event', 'visitors_medical'
]


class Feedback(db.Model):
    __tablename__ = 'feedback'

    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(50), nullable=False, default='Bug')
    subject = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(50), nullable=False, default='New')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'type': self.type,
            'subject': self.subject,
            'description': self.description,
            'status': self.status,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }


class InvoiceSequence(db.Model):
    __tablename__ = 'invoice_sequence'

    id = db.Column(db.Integer, primary_key=True)
    last_number = db.Column(db.Integer, nullable=False, default=536658)

    @staticmethod
    def next_number(session):
        """Get and increment the next invoice number. Thread-safe via DB."""
        seq = session.query(InvoiceSequence).first()
        if not seq:
            seq = InvoiceSequence(last_number=536658)
            session.add(seq)
            session.flush()
        seq.last_number += 1
        session.flush()
        return seq.last_number


# ===========================================================================
# UTILITY FUNCTIONS
# ===========================================================================

def parse_date(date_str):
    """Parse date string to date object."""
    if not date_str or date_str == '':
        return None
    try:
        return parse(date_str).date()
    except:
        return None


def save_benefit_plans(session, benefit, plans_data):
    """Save multi-plan child records for a benefit. Deletes existing plans first."""
    # Delete existing plans for this benefit
    session.query(BenefitPlan).filter_by(employee_benefit_id=benefit.id).delete()
    session.flush()

    for plan_type in MULTI_PLAN_TYPES:
        for idx, plan_info in enumerate(plans_data.get(plan_type, []), 1):
            carrier = plan_info.get('carrier')
            renewal = plan_info.get('renewal_date')
            if carrier or renewal:
                plan = BenefitPlan(
                    employee_benefit_id=benefit.id,
                    plan_type=plan_type,
                    plan_number=idx,
                    carrier=carrier,
                    renewal_date=parse_date(renewal),
                    waiting_period=plan_info.get('waiting_period') or None,
                    remarks=plan_info.get('remarks') or None,
                    outstanding_item=plan_info.get('outstanding_item') or None
                )
                session.add(plan)

    # Also update flat fields from first plan for backward compat
    for plan_type in MULTI_PLAN_TYPES:
        plans_for_type = plans_data.get(plan_type, [])
        first = plans_for_type[0] if plans_for_type else {}
        if plan_type == 'medical':
            benefit.current_carrier = first.get('carrier') or None
            benefit.renewal_date = parse_date(first.get('renewal_date'))
        else:
            setattr(benefit, f'{plan_type}_carrier', first.get('carrier') or None)
            setattr(benefit, f'{plan_type}_renewal_date', parse_date(first.get('renewal_date')))


def save_commercial_plans(session, commercial, plans_data):
    """Save multi-plan child records for commercial insurance. Deletes existing plans first."""
    session.query(CommercialPlan).filter_by(commercial_insurance_id=commercial.id).delete()
    session.flush()

    for plan_type in MULTI_PLAN_COMMERCIAL_TYPES:
        for idx, plan_info in enumerate(plans_data.get(plan_type, []), 1):
            carrier = plan_info.get('carrier')
            agency = plan_info.get('agency')
            policy_number = plan_info.get('policy_number')
            renewal = plan_info.get('renewal_date')
            occ_limit_val = plan_info.get('occ_limit')
            agg_limit_val = plan_info.get('agg_limit')
            premium_val = plan_info.get('premium')
            if carrier or renewal or occ_limit_val or agg_limit_val or premium_val:
                plan = CommercialPlan(
                    commercial_insurance_id=commercial.id,
                    plan_type=plan_type,
                    plan_number=idx,
                    carrier=carrier,
                    agency=agency,
                    policy_number=policy_number or None,
                    coverage_occ_limit=occ_limit_val,
                    coverage_agg_limit=agg_limit_val,
                    premium=parse_premium(premium_val),
                    renewal_date=parse_date(renewal),
                    remarks=plan_info.get('remarks') or None,
                    outstanding_item=plan_info.get('outstanding_item') or None,
                    outstanding_item_due_date=parse_date(plan_info.get('outstanding_item_due_date')),
                    endorsement_tech_eo=bool(plan_info.get('endorsement_tech_eo')),
                    endorsement_allied_healthcare=bool(plan_info.get('endorsement_allied_healthcare')),
                    endorsement_staffing=bool(plan_info.get('endorsement_staffing')),
                    endorsement_medical_malpractice=bool(plan_info.get('endorsement_medical_malpractice'))
                )
                session.add(plan)

    # Also update flat fields from first plan for backward compat
    for plan_type in MULTI_PLAN_COMMERCIAL_TYPES:
        plans_for_type = plans_data.get(plan_type, [])
        first = plans_for_type[0] if plans_for_type else {}
        setattr(commercial, f'{plan_type}_carrier', first.get('carrier') or None)
        setattr(commercial, f'{plan_type}_agency', first.get('agency') or None)
        setattr(commercial, f'{plan_type}_policy_number', first.get('policy_number') or None)
        setattr(commercial, f'{plan_type}_occ_limit', first.get('occ_limit') or None)
        setattr(commercial, f'{plan_type}_agg_limit', first.get('agg_limit') or None)
        setattr(commercial, f'{plan_type}_premium', parse_premium(first.get('premium')))
        setattr(commercial, f'{plan_type}_renewal_date', parse_date(first.get('renewal_date')))


def save_homeowners_policies(session, personal, policies_data):
    """Save homeowners policy child records. Deletes existing first."""
    session.query(HomeownersPolicy).filter_by(personal_insurance_id=personal.id).delete()
    session.flush()
    for idx, p in enumerate(policies_data or [], 1):
        carrier = p.get('carrier')
        dwelling = p.get('dwelling_limit')
        liability = p.get('liability_limit')
        premium = p.get('premium')
        renewal = p.get('renewal_date')
        if carrier or dwelling or liability or premium or renewal:
            policy = HomeownersPolicy(
                personal_insurance_id=personal.id,
                policy_number=idx,
                carrier=carrier,
                dwelling_limit=dwelling,
                liability_limit=liability,
                premium=parse_premium(premium),
                renewal_date=parse_date(renewal),
                remarks=p.get('remarks') or None,
                outstanding_item=p.get('outstanding_item') or None,
                outstanding_item_due_date=parse_date(p.get('outstanding_item_due_date')),
                property_address_line_1=p.get('property_address_line_1') or None,
                property_address_line_2=p.get('property_address_line_2') or None,
                property_city=p.get('property_city') or None,
                property_state=p.get('property_state') or None,
                property_zip=p.get('property_zip') or None,
                is_primary_residence=bool(p.get('is_primary_residence'))
            )
            session.add(policy)
    # Update flat fields from first policy for backward compat
    first = policies_data[0] if policies_data else {}
    personal.homeowners_carrier = first.get('carrier') or None
    personal.homeowners_dwelling_limit = first.get('dwelling_limit') or None
    personal.homeowners_liability_limit = first.get('liability_limit') or None
    personal.homeowners_premium = parse_premium(first.get('premium'))
    personal.homeowners_renewal_date = parse_date(first.get('renewal_date'))
    personal.homeowners_remarks = first.get('remarks') or None
    personal.homeowners_outstanding_item = first.get('outstanding_item') or None


def parse_premium(val):
    """Parse premium value to float, returning None for empty/invalid values."""
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def format_limit(val):
    """Convert a limit value to millions for storage.
    '1,000,000' or 1000000 → '1'
    '2,000,000' or 2000000 → '2'
    '5,000,000' or 5000000 → '5'
    '500,000' or 500000 → '0.5'
    Already in millions (e.g. '1', '2', '5') → kept as-is.
    Non-numeric strings (e.g. 'N/A') → kept as-is."""
    if val is None or val == '':
        return None
    s = str(val).strip()
    if s.upper() == 'N/A':
        return None
    # Remove commas to get raw number
    raw = s.replace(',', '')
    try:
        num = float(raw)
        # If >= 100000, assume it's in raw dollars — convert to millions
        if num >= 100000:
            m = num / 1000000
            return str(int(m)) if m == int(m) else str(m)
        # Otherwise already in millions or a small value — keep as-is
        return str(int(num)) if num == int(num) else str(num)
    except (ValueError, TypeError):
        return s


# ===========================================================================
# CLIENT ENDPOINTS
# ===========================================================================

@app.route('/api/clients', methods=['GET'])
def get_clients():
    """Get all clients."""
    session = Session()
    try:
        clients = session.query(Client).all()
        return jsonify({
            'clients': [client.to_dict() for client in clients],
            'total': len(clients)
        }), 200
    except Exception as e:
        logging.error(f"Error fetching clients: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/clients', methods=['POST'])
def create_client():
    """Create a new client."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Validate tax_id format (##-#######)
        tax_id = data.get('tax_id', '')
        if tax_id and not re.match(r'^\d{2}-\d{7}$', tax_id):
            return jsonify({'error': 'Tax ID must be in ##-####### format'}), 400

        # Validate zip_code (5 digits)
        zip_code = data.get('zip_code', '')
        if zip_code and not re.match(r'^\d{5}$', str(zip_code)):
            return jsonify({'error': 'Zip code must be exactly 5 digits'}), 400

        client = Client(
            tax_id=data.get('tax_id'),
            client_name=data.get('client_name'),
            dba=data.get('dba'),
            industry=data.get('industry'),
            contact_person=data.get('contact_person'),
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            address_line_1=data.get('address_line_1'),
            address_line_2=data.get('address_line_2'),
            city=data.get('city'),
            state=data.get('state'),
            zip_code=data.get('zip_code'),
            status=data.get('status', 'Active'),
            gross_revenue=data.get('gross_revenue'),
            total_ees=data.get('total_ees')
        )

        session.add(client)
        session.flush()

        # Save contacts
        contacts = data.get('contacts', [])
        for i, c in enumerate(contacts):
            if c.get('contact_person') or c.get('email') or c.get('phone_number') or c.get('address_line_1'):
                contact = ClientContact(
                    client_id=client.id,
                    contact_person=c.get('contact_person', ''),
                    email=c.get('email', ''),
                    phone_number=c.get('phone_number', ''),
                    phone_extension=c.get('phone_extension', ''),
                    address_line_1=c.get('address_line_1', ''),
                    address_line_2=c.get('address_line_2', ''),
                    city=c.get('city', ''),
                    state=c.get('state', ''),
                    zip_code=c.get('zip_code', ''),
                    sort_order=i
                )
                session.add(contact)

        session.commit()

        return jsonify({
            'message': 'Client created successfully',
            'client': client.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error creating client: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/clients/<int:client_id>', methods=['PUT'])
def update_client(client_id):
    """Update a client."""
    session = Session()
    try:
        client = session.query(Client).filter_by(id=client_id).first()
        if not client:
            return jsonify({'error': 'Client not found'}), 404

        data = request.get_json()

        # Validate tax_id format (##-#######)
        tax_id = data.get('tax_id', client.tax_id)
        if tax_id and not re.match(r'^\d{2}-\d{7}$', tax_id):
            return jsonify({'error': 'Tax ID must be in ##-####### format'}), 400

        # Validate zip_code (5 digits)
        zip_code = data.get('zip_code', client.zip_code)
        if zip_code and not re.match(r'^\d{5}$', str(zip_code)):
            return jsonify({'error': 'Zip code must be exactly 5 digits'}), 400

        old_tax_id = client.tax_id
        new_tax_id = data.get('tax_id', client.tax_id)

        # If tax_id changed, update all associated policies
        if new_tax_id != old_tax_id:
            # Check new tax_id isn't already taken by another client
            existing = session.query(Client).filter(
                Client.tax_id == new_tax_id, Client.id != client_id
            ).first()
            if existing:
                return jsonify({'error': f'Tax ID {new_tax_id} is already assigned to another client'}), 400

            # Update tax_id in associated employee benefits
            session.query(EmployeeBenefit).filter_by(tax_id=old_tax_id).update(
                {'tax_id': new_tax_id}, synchronize_session='fetch'
            )
            # Update tax_id in associated commercial insurance
            session.query(CommercialInsurance).filter_by(tax_id=old_tax_id).update(
                {'tax_id': new_tax_id}, synchronize_session='fetch'
            )

        client.tax_id = new_tax_id
        client.client_name = data.get('client_name', client.client_name)
        client.dba = data.get('dba', client.dba)
        client.contact_person = data.get('contact_person', client.contact_person)
        client.email = data.get('email', client.email)
        client.phone_number = data.get('phone_number', client.phone_number)
        client.address_line_1 = data.get('address_line_1', client.address_line_1)
        client.address_line_2 = data.get('address_line_2', client.address_line_2)
        client.city = data.get('city', client.city)
        client.state = data.get('state', client.state)
        client.zip_code = data.get('zip_code', client.zip_code)
        client.status = data.get('status', client.status)
        client.gross_revenue = data.get('gross_revenue', client.gross_revenue)
        client.total_ees = data.get('total_ees', client.total_ees)
        client.industry = data.get('industry', client.industry)

        # Update contacts if provided
        if 'contacts' in data:
            # Remove existing contacts
            session.query(ClientContact).filter_by(client_id=client.id).delete()
            # Add new contacts
            for i, c in enumerate(data['contacts']):
                if c.get('contact_person') or c.get('email') or c.get('phone_number'):
                    contact = ClientContact(
                        client_id=client.id,
                        contact_person=c.get('contact_person', ''),
                        email=c.get('email', ''),
                        phone_number=c.get('phone_number', ''),
                        phone_extension=c.get('phone_extension', ''),
                        sort_order=i
                    )
                    session.add(contact)

        session.commit()

        return jsonify({
            'message': 'Client updated successfully',
            'client': client.to_dict()
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error updating client: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    """Delete a client."""
    session = Session()
    try:
        client = session.query(Client).filter_by(id=client_id).first()
        if not client:
            return jsonify({'error': 'Client not found'}), 404

        session.delete(client)
        session.commit()

        return jsonify({'message': 'Client deleted successfully'}), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error deleting client: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/clients/<int:client_id>/clone', methods=['POST'])
def clone_client(client_id):
    """Clone a client."""
    session = Session()
    try:
        original = session.query(Client).filter_by(id=client_id).first()
        if not original:
            return jsonify({'error': 'Client not found'}), 404

        # Create new client with copied data
        new_client = Client(
            tax_id=original.tax_id + '_COPY',
            client_name=original.client_name + ' (Copy)',
            contact_person=original.contact_person,
            email=original.email,
            phone_number=original.phone_number,
            address_line_1=original.address_line_1,
            address_line_2=original.address_line_2,
            city=original.city,
            state=original.state,
            zip_code=original.zip_code,
            status=original.status,
            gross_revenue=original.gross_revenue,
            total_ees=original.total_ees
        )

        session.add(new_client)
        session.commit()

        return jsonify({
            'message': 'Client cloned successfully',
            'client': new_client.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error cloning client: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# EMPLOYEE BENEFITS ENDPOINTS
# ===========================================================================

@app.route('/api/benefits', methods=['GET'])
def get_benefits():
    """Get all employee benefits with client info."""
    session = Session()
    try:
        benefits = session.query(EmployeeBenefit).all()
        return jsonify({
            'benefits': [benefit.to_dict() for benefit in benefits],
            'total': len(benefits)
        }), 200
    except Exception as e:
        logging.error(f"Error fetching benefits: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits/<int:benefit_id>', methods=['GET'])
def get_benefit(benefit_id):
    """Get a single benefit record."""
    session = Session()
    try:
        benefit = session.query(EmployeeBenefit).filter_by(id=benefit_id).first()
        if not benefit:
            return jsonify({'error': 'Benefit not found'}), 404
        return jsonify(benefit.to_dict()), 200
    except Exception as e:
        logging.error(f"Error fetching benefit: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits', methods=['POST'])
def create_benefit():
    """Create new employee benefit record."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        benefit = EmployeeBenefit(
            tax_id=data.get('tax_id'),
            parent_client=data.get('parent_client'),
            form_fire_code=data.get('form_fire_code'),
            enrollment_poc=data.get('enrollment_poc'),
            renewal_date=parse_date(data.get('renewal_date')),
            funding=data.get('funding'),
            current_carrier=data.get('current_carrier'),
            num_employees_at_renewal=data.get('num_employees_at_renewal'),
            enrolled_ees=data.get('enrolled_ees'),
            waiting_period=data.get('waiting_period'),
            deductible_accumulation=data.get('deductible_accumulation'),
            previous_carrier=data.get('previous_carrier'),
            cobra_carrier=data.get('cobra_carrier'),
            employee_contribution=data.get('employee_contribution'),
            dental_renewal_date=parse_date(data.get('dental_renewal_date')),
            dental_carrier=data.get('dental_carrier'),
            vision_renewal_date=parse_date(data.get('vision_renewal_date')),
            vision_carrier=data.get('vision_carrier'),
            life_adnd_renewal_date=parse_date(data.get('life_adnd_renewal_date')),
            life_adnd_carrier=data.get('life_adnd_carrier'),
            ltd_renewal_date=parse_date(data.get('ltd_renewal_date')),
            ltd_carrier=data.get('ltd_carrier'),
            ltd_remarks=data.get('ltd_remarks'),
            ltd_outstanding_item=data.get('ltd_outstanding_item'),
            ltd_outstanding_item_due_date=parse_date(data.get('ltd_outstanding_item_due_date')),
            std_renewal_date=parse_date(data.get('std_renewal_date')),
            std_carrier=data.get('std_carrier'),
            std_remarks=data.get('std_remarks'),
            std_outstanding_item=data.get('std_outstanding_item'),
            std_outstanding_item_due_date=parse_date(data.get('std_outstanding_item_due_date')),
            k401_renewal_date=parse_date(data.get('k401_renewal_date')),
            k401_carrier=data.get('k401_carrier'),
            k401_remarks=data.get('k401_remarks'),
            k401_outstanding_item=data.get('k401_outstanding_item'),
            k401_outstanding_item_due_date=parse_date(data.get('k401_outstanding_item_due_date')),
            critical_illness_renewal_date=parse_date(data.get('critical_illness_renewal_date')),
            critical_illness_carrier=data.get('critical_illness_carrier'),
            critical_illness_remarks=data.get('critical_illness_remarks'),
            critical_illness_outstanding_item=data.get('critical_illness_outstanding_item'),
            critical_illness_outstanding_item_due_date=parse_date(data.get('critical_illness_outstanding_item_due_date')),
            accident_renewal_date=parse_date(data.get('accident_renewal_date')),
            accident_carrier=data.get('accident_carrier'),
            accident_remarks=data.get('accident_remarks'),
            accident_outstanding_item=data.get('accident_outstanding_item'),
            accident_outstanding_item_due_date=parse_date(data.get('accident_outstanding_item_due_date')),
            hospital_renewal_date=parse_date(data.get('hospital_renewal_date')),
            hospital_carrier=data.get('hospital_carrier'),
            hospital_remarks=data.get('hospital_remarks'),
            hospital_outstanding_item=data.get('hospital_outstanding_item'),
            hospital_outstanding_item_due_date=parse_date(data.get('hospital_outstanding_item_due_date')),
            voluntary_life_renewal_date=parse_date(data.get('voluntary_life_renewal_date')),
            voluntary_life_carrier=data.get('voluntary_life_carrier'),
            voluntary_life_remarks=data.get('voluntary_life_remarks'),
            voluntary_life_outstanding_item=data.get('voluntary_life_outstanding_item'),
            voluntary_life_outstanding_item_due_date=parse_date(data.get('voluntary_life_outstanding_item_due_date')),
            outstanding_item_due_date=parse_date(data.get('outstanding_item_due_date'))
        )

        session.add(benefit)
        session.flush()

        # Save multi-plan child records if provided
        if 'plans' in data:
            save_benefit_plans(session, benefit, data['plans'])

        session.commit()

        # Refresh to load plans relationship
        session.refresh(benefit)

        return jsonify({
            'message': 'Benefit created successfully',
            'benefit': benefit.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error creating benefit: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits/<int:benefit_id>', methods=['PUT'])
def update_benefit(benefit_id):
    """Update employee benefit record."""
    session = Session()
    try:
        benefit = session.query(EmployeeBenefit).filter_by(id=benefit_id).first()
        if not benefit:
            return jsonify({'error': 'Benefit not found'}), 404

        data = request.get_json()

        # Update all fields
        benefit.tax_id = data.get('tax_id', benefit.tax_id)
        benefit.parent_client = data.get('parent_client', benefit.parent_client)
        benefit.form_fire_code = data.get('form_fire_code', benefit.form_fire_code)
        benefit.enrollment_poc = data.get('enrollment_poc', benefit.enrollment_poc)
        benefit.renewal_date = parse_date(data.get('renewal_date')) if 'renewal_date' in data else benefit.renewal_date
        benefit.funding = data.get('funding', benefit.funding)
        benefit.current_carrier = data.get('current_carrier', benefit.current_carrier)
        benefit.num_employees_at_renewal = data.get('num_employees_at_renewal', benefit.num_employees_at_renewal)
        benefit.enrolled_ees = data.get('enrolled_ees', benefit.enrolled_ees)
        benefit.waiting_period = data.get('waiting_period', benefit.waiting_period)
        benefit.deductible_accumulation = data.get('deductible_accumulation', benefit.deductible_accumulation)
        benefit.previous_carrier = data.get('previous_carrier', benefit.previous_carrier)
        benefit.cobra_carrier = data.get('cobra_carrier', benefit.cobra_carrier)
        benefit.employee_contribution = data.get('employee_contribution', benefit.employee_contribution)

        # Update benefit plans
        if 'dental_renewal_date' in data:
            benefit.dental_renewal_date = parse_date(data.get('dental_renewal_date'))
        if 'dental_carrier' in data:
            benefit.dental_carrier = data.get('dental_carrier')
        if 'vision_renewal_date' in data:
            benefit.vision_renewal_date = parse_date(data.get('vision_renewal_date'))
        if 'vision_carrier' in data:
            benefit.vision_carrier = data.get('vision_carrier')
        if 'life_adnd_renewal_date' in data:
            benefit.life_adnd_renewal_date = parse_date(data.get('life_adnd_renewal_date'))
        if 'life_adnd_carrier' in data:
            benefit.life_adnd_carrier = data.get('life_adnd_carrier')
        if 'ltd_renewal_date' in data:
            benefit.ltd_renewal_date = parse_date(data.get('ltd_renewal_date'))
        if 'ltd_carrier' in data:
            benefit.ltd_carrier = data.get('ltd_carrier')
        if 'std_renewal_date' in data:
            benefit.std_renewal_date = parse_date(data.get('std_renewal_date'))
        if 'std_carrier' in data:
            benefit.std_carrier = data.get('std_carrier')
        if 'k401_renewal_date' in data:
            benefit.k401_renewal_date = parse_date(data.get('k401_renewal_date'))
        if 'k401_carrier' in data:
            benefit.k401_carrier = data.get('k401_carrier')
        if 'critical_illness_renewal_date' in data:
            benefit.critical_illness_renewal_date = parse_date(data.get('critical_illness_renewal_date'))
        if 'critical_illness_carrier' in data:
            benefit.critical_illness_carrier = data.get('critical_illness_carrier')
        if 'accident_renewal_date' in data:
            benefit.accident_renewal_date = parse_date(data.get('accident_renewal_date'))
        if 'accident_carrier' in data:
            benefit.accident_carrier = data.get('accident_carrier')
        if 'hospital_renewal_date' in data:
            benefit.hospital_renewal_date = parse_date(data.get('hospital_renewal_date'))
        if 'hospital_carrier' in data:
            benefit.hospital_carrier = data.get('hospital_carrier')
        if 'voluntary_life_renewal_date' in data:
            benefit.voluntary_life_renewal_date = parse_date(data.get('voluntary_life_renewal_date'))
        if 'voluntary_life_carrier' in data:
            benefit.voluntary_life_carrier = data.get('voluntary_life_carrier')

        # Update outstanding_item_due_date for core
        if 'outstanding_item_due_date' in data:
            benefit.outstanding_item_due_date = parse_date(data.get('outstanding_item_due_date'))

        # Update single-plan remarks, outstanding_item, and due_date columns
        for prefix in ['ltd', 'std', 'k401', 'critical_illness', 'accident', 'hospital', 'voluntary_life']:
            if f'{prefix}_remarks' in data:
                setattr(benefit, f'{prefix}_remarks', data.get(f'{prefix}_remarks'))
            if f'{prefix}_outstanding_item' in data:
                setattr(benefit, f'{prefix}_outstanding_item', data.get(f'{prefix}_outstanding_item'))
            if f'{prefix}_outstanding_item_due_date' in data:
                setattr(benefit, f'{prefix}_outstanding_item_due_date', parse_date(data.get(f'{prefix}_outstanding_item_due_date')))

        # Save multi-plan child records if provided
        if 'plans' in data:
            save_benefit_plans(session, benefit, data['plans'])

        session.commit()

        # Refresh to load updated plans relationship
        session.refresh(benefit)

        return jsonify({
            'message': 'Benefit updated successfully',
            'benefit': benefit.to_dict()
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error updating benefit: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits/<int:benefit_id>', methods=['DELETE'])
def delete_benefit(benefit_id):
    """Delete employee benefit record."""
    session = Session()
    try:
        benefit = session.query(EmployeeBenefit).filter_by(id=benefit_id).first()
        if not benefit:
            return jsonify({'error': 'Benefit not found'}), 404

        session.delete(benefit)
        session.commit()

        return jsonify({'message': 'Benefit deleted successfully'}), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error deleting benefit: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits/<int:benefit_id>/clone', methods=['POST'])
def clone_benefit(benefit_id):
    """Clone employee benefit record."""
    session = Session()
    try:
        original = session.query(EmployeeBenefit).filter_by(id=benefit_id).first()
        if not original:
            return jsonify({'error': 'Benefit not found'}), 404

        new_benefit = EmployeeBenefit(
            tax_id=original.tax_id,
            form_fire_code=original.form_fire_code,
            enrollment_poc=original.enrollment_poc,
            renewal_date=original.renewal_date,
            funding=original.funding,
            current_carrier=original.current_carrier,
            num_employees_at_renewal=original.num_employees_at_renewal,
            enrolled_ees=original.enrolled_ees,
            waiting_period=original.waiting_period,
            deductible_accumulation=original.deductible_accumulation,
            previous_carrier=original.previous_carrier,
            cobra_carrier=original.cobra_carrier,
            employee_contribution=original.employee_contribution,
            dental_renewal_date=original.dental_renewal_date,
            dental_carrier=original.dental_carrier,
            vision_renewal_date=original.vision_renewal_date,
            vision_carrier=original.vision_carrier,
            life_adnd_renewal_date=original.life_adnd_renewal_date,
            life_adnd_carrier=original.life_adnd_carrier,
            ltd_renewal_date=original.ltd_renewal_date,
            ltd_carrier=original.ltd_carrier,
            ltd_remarks=original.ltd_remarks,
            ltd_outstanding_item=original.ltd_outstanding_item,
            std_renewal_date=original.std_renewal_date,
            std_carrier=original.std_carrier,
            std_remarks=original.std_remarks,
            std_outstanding_item=original.std_outstanding_item,
            k401_renewal_date=original.k401_renewal_date,
            k401_carrier=original.k401_carrier,
            k401_remarks=original.k401_remarks,
            k401_outstanding_item=original.k401_outstanding_item,
            critical_illness_renewal_date=original.critical_illness_renewal_date,
            critical_illness_carrier=original.critical_illness_carrier,
            critical_illness_remarks=original.critical_illness_remarks,
            critical_illness_outstanding_item=original.critical_illness_outstanding_item,
            accident_renewal_date=original.accident_renewal_date,
            accident_carrier=original.accident_carrier,
            accident_remarks=original.accident_remarks,
            accident_outstanding_item=original.accident_outstanding_item,
            hospital_renewal_date=original.hospital_renewal_date,
            hospital_carrier=original.hospital_carrier,
            hospital_remarks=original.hospital_remarks,
            hospital_outstanding_item=original.hospital_outstanding_item,
            voluntary_life_renewal_date=original.voluntary_life_renewal_date,
            voluntary_life_carrier=original.voluntary_life_carrier,
            voluntary_life_remarks=original.voluntary_life_remarks,
            voluntary_life_outstanding_item=original.voluntary_life_outstanding_item
        )

        session.add(new_benefit)
        session.flush()

        # Clone child BenefitPlan records
        for plan in original.plans:
            new_plan = BenefitPlan(
                employee_benefit_id=new_benefit.id,
                plan_type=plan.plan_type,
                plan_number=plan.plan_number,
                carrier=plan.carrier,
                renewal_date=plan.renewal_date,
                remarks=plan.remarks,
                outstanding_item=plan.outstanding_item
            )
            session.add(new_plan)

        session.commit()

        session.refresh(new_benefit)

        return jsonify({
            'message': 'Benefit cloned successfully',
            'benefit': new_benefit.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error cloning benefit: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# COMMERCIAL INSURANCE ENDPOINTS
# ===========================================================================

@app.route('/api/commercial', methods=['GET'])
def get_commercial():
    """Get all commercial insurance records with client info."""
    session = Session()
    try:
        commercial = session.query(CommercialInsurance).all()
        return jsonify({
            'commercial': [c.to_dict() for c in commercial],
            'total': len(commercial)
        }), 200
    except Exception as e:
        logging.error(f"Error fetching commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/commercial/<int:commercial_id>', methods=['GET'])
def get_commercial_single(commercial_id):
    """Get a single commercial insurance record."""
    session = Session()
    try:
        commercial = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not commercial:
            return jsonify({'error': 'Commercial insurance not found'}), 404
        return jsonify(commercial.to_dict()), 200
    except Exception as e:
        logging.error(f"Error fetching commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/commercial', methods=['POST'])
def create_commercial():
    """Create new commercial insurance record."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Single-plan product types (not in MULTI_PLAN_COMMERCIAL_TYPES)
        single_plan_products = [
            'general_liability', 'property', 'bop', 'workers_comp', 'auto',
            'epli', 'nydbl', 'surety', 'product_liability', 'flood',
            'directors_officers', 'fiduciary', 'inland_marine'
        ]

        commercial = CommercialInsurance(
            tax_id=data.get('tax_id'),
            parent_client=data.get('parent_client'),
            assigned_to=data.get('assigned_to') or None,
        )

        # Set single-plan product fields
        for product in single_plan_products:
            setattr(commercial, f'{product}_carrier', data.get(f'{product}_carrier') or None)
            setattr(commercial, f'{product}_agency', data.get(f'{product}_agency') or None)
            setattr(commercial, f'{product}_policy_number', data.get(f'{product}_policy_number') or None)
            setattr(commercial, f'{product}_occ_limit', data.get(f'{product}_occ_limit') or None)
            setattr(commercial, f'{product}_agg_limit', data.get(f'{product}_agg_limit') or None)
            setattr(commercial, f'{product}_premium', parse_premium(data.get(f'{product}_premium')))
            setattr(commercial, f'{product}_renewal_date', parse_date(data.get(f'{product}_renewal_date')))
            setattr(commercial, f'{product}_remarks', data.get(f'{product}_remarks') or None)
            setattr(commercial, f'{product}_outstanding_item', data.get(f'{product}_outstanding_item') or None)
            setattr(commercial, f'{product}_outstanding_item_due_date', parse_date(data.get(f'{product}_outstanding_item_due_date')))

        # GL endorsements
        for endorsement in ['bop', 'marine', 'foreign', 'molestation', 'staffing']:
            setattr(commercial, f'general_liability_endorsement_{endorsement}',
                    bool(data.get(f'general_liability_endorsement_{endorsement}')))

        # BOP property coverage fields
        commercial.bop_building_limit = parse_premium(data.get('bop_building_limit'))
        commercial.bop_personal_property = parse_premium(data.get('bop_personal_property'))

        # Multi-plan type flat fields (backward compat - will be overwritten by save_commercial_plans)
        for product in MULTI_PLAN_COMMERCIAL_TYPES:
            setattr(commercial, f'{product}_carrier', data.get(f'{product}_carrier') or None)
            setattr(commercial, f'{product}_agency', data.get(f'{product}_agency') or None)
            setattr(commercial, f'{product}_policy_number', data.get(f'{product}_policy_number') or None)
            setattr(commercial, f'{product}_occ_limit', data.get(f'{product}_occ_limit') or None)
            setattr(commercial, f'{product}_agg_limit', data.get(f'{product}_agg_limit') or None)
            setattr(commercial, f'{product}_premium', parse_premium(data.get(f'{product}_premium')))
            setattr(commercial, f'{product}_renewal_date', parse_date(data.get(f'{product}_renewal_date')))

        session.add(commercial)
        session.flush()

        # Save multi-plan child records if plans data provided
        if 'plans' in data:
            save_commercial_plans(session, commercial, data['plans'])

        session.commit()
        session.refresh(commercial)

        return jsonify({
            'message': 'Commercial insurance created successfully',
            'commercial': commercial.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error creating commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/commercial/<int:commercial_id>', methods=['PUT'])
def update_commercial(commercial_id):
    """Update commercial insurance record."""
    session = Session()
    try:
        commercial = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not commercial:
            return jsonify({'error': 'Commercial insurance not found'}), 404

        data = request.get_json()

        # Update core fields
        commercial.tax_id = data.get('tax_id', commercial.tax_id)
        commercial.parent_client = data.get('parent_client', commercial.parent_client)
        if 'assigned_to' in data:
            commercial.assigned_to = data.get('assigned_to') or None

        # Update single-plan insurance products
        single_plan_products = [
            'general_liability', 'property', 'bop', 'workers_comp', 'auto',
            'epli', 'nydbl', 'surety', 'product_liability', 'flood',
            'directors_officers', 'fiduciary', 'inland_marine'
        ]
        for product in single_plan_products:
            if f'{product}_carrier' in data:
                setattr(commercial, f'{product}_carrier', data.get(f'{product}_carrier') or None)
            if f'{product}_agency' in data:
                setattr(commercial, f'{product}_agency', data.get(f'{product}_agency') or None)
            if f'{product}_policy_number' in data:
                setattr(commercial, f'{product}_policy_number', data.get(f'{product}_policy_number') or None)
            if f'{product}_occ_limit' in data:
                setattr(commercial, f'{product}_occ_limit', data.get(f'{product}_occ_limit') or None)
            if f'{product}_agg_limit' in data:
                setattr(commercial, f'{product}_agg_limit', data.get(f'{product}_agg_limit') or None)
            if f'{product}_premium' in data:
                setattr(commercial, f'{product}_premium', parse_premium(data.get(f'{product}_premium')))
            if f'{product}_renewal_date' in data:
                setattr(commercial, f'{product}_renewal_date', parse_date(data.get(f'{product}_renewal_date')))
            if f'{product}_remarks' in data:
                setattr(commercial, f'{product}_remarks', data.get(f'{product}_remarks') or None)
            if f'{product}_outstanding_item' in data:
                setattr(commercial, f'{product}_outstanding_item', data.get(f'{product}_outstanding_item') or None)
            if f'{product}_outstanding_item_due_date' in data:
                setattr(commercial, f'{product}_outstanding_item_due_date', parse_date(data.get(f'{product}_outstanding_item_due_date')))

        # Update GL endorsements
        for endorsement in ['bop', 'marine', 'foreign', 'molestation', 'staffing']:
            key = f'general_liability_endorsement_{endorsement}'
            if key in data:
                setattr(commercial, key, bool(data.get(key)))

        # BOP property coverage fields
        if 'bop_building_limit' in data:
            commercial.bop_building_limit = parse_premium(data.get('bop_building_limit'))
        if 'bop_personal_property' in data:
            commercial.bop_personal_property = parse_premium(data.get('bop_personal_property'))

        # Update multi-plan type flat fields (backward compat)
        for product in MULTI_PLAN_COMMERCIAL_TYPES:
            if f'{product}_carrier' in data:
                setattr(commercial, f'{product}_carrier', data.get(f'{product}_carrier') or None)
            if f'{product}_agency' in data:
                setattr(commercial, f'{product}_agency', data.get(f'{product}_agency') or None)
            if f'{product}_policy_number' in data:
                setattr(commercial, f'{product}_policy_number', data.get(f'{product}_policy_number') or None)
            if f'{product}_occ_limit' in data:
                setattr(commercial, f'{product}_occ_limit', data.get(f'{product}_occ_limit') or None)
            if f'{product}_agg_limit' in data:
                setattr(commercial, f'{product}_agg_limit', data.get(f'{product}_agg_limit') or None)
            if f'{product}_premium' in data:
                setattr(commercial, f'{product}_premium', parse_premium(data.get(f'{product}_premium')))
            if f'{product}_renewal_date' in data:
                setattr(commercial, f'{product}_renewal_date', parse_date(data.get(f'{product}_renewal_date')))

        # Save multi-plan child records if plans data provided
        if 'plans' in data:
            save_commercial_plans(session, commercial, data['plans'])

        session.commit()

        return jsonify({
            'message': 'Commercial insurance updated successfully',
            'commercial': commercial.to_dict()
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error updating commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/commercial/<int:commercial_id>', methods=['DELETE'])
def delete_commercial(commercial_id):
    """Delete commercial insurance record."""
    session = Session()
    try:
        commercial = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not commercial:
            return jsonify({'error': 'Commercial insurance not found'}), 404

        session.delete(commercial)
        session.commit()

        return jsonify({'message': 'Commercial insurance deleted successfully'}), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error deleting commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/commercial/<int:commercial_id>/clone', methods=['POST'])
def clone_commercial(commercial_id):
    """Clone commercial insurance record."""
    session = Session()
    try:
        original = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not original:
            return jsonify({'error': 'Commercial insurance not found'}), 404

        new_commercial = CommercialInsurance(
            tax_id=original.tax_id,
            # Copy all product fields
            general_liability_carrier=original.general_liability_carrier,
            general_liability_agency=original.general_liability_agency,
            general_liability_occ_limit=original.general_liability_occ_limit,
            general_liability_agg_limit=original.general_liability_agg_limit,
            general_liability_premium=original.general_liability_premium,
            general_liability_renewal_date=original.general_liability_renewal_date,
            property_carrier=original.property_carrier,
            property_agency=original.property_agency,
            property_occ_limit=original.property_occ_limit,
            property_agg_limit=original.property_agg_limit,
            property_premium=original.property_premium,
            property_renewal_date=original.property_renewal_date,
            bop_carrier=original.bop_carrier,
            bop_agency=original.bop_agency,
            bop_occ_limit=original.bop_occ_limit,
            bop_agg_limit=original.bop_agg_limit,
            bop_premium=original.bop_premium,
            bop_renewal_date=original.bop_renewal_date,
            bop_building_limit=original.bop_building_limit,
            bop_personal_property=original.bop_personal_property,
            umbrella_carrier=original.umbrella_carrier,
            umbrella_agency=original.umbrella_agency,
            umbrella_occ_limit=original.umbrella_occ_limit,
            umbrella_agg_limit=original.umbrella_agg_limit,
            umbrella_premium=original.umbrella_premium,
            umbrella_renewal_date=original.umbrella_renewal_date,
            workers_comp_carrier=original.workers_comp_carrier,
            workers_comp_agency=original.workers_comp_agency,
            workers_comp_occ_limit=original.workers_comp_occ_limit,
            workers_comp_agg_limit=original.workers_comp_agg_limit,
            workers_comp_premium=original.workers_comp_premium,
            workers_comp_renewal_date=original.workers_comp_renewal_date,
            professional_eo_carrier=original.professional_eo_carrier,
            professional_eo_agency=original.professional_eo_agency,
            professional_eo_occ_limit=original.professional_eo_occ_limit,
            professional_eo_agg_limit=original.professional_eo_agg_limit,
            professional_eo_premium=original.professional_eo_premium,
            professional_eo_renewal_date=original.professional_eo_renewal_date,
            cyber_carrier=original.cyber_carrier,
            cyber_agency=original.cyber_agency,
            cyber_occ_limit=original.cyber_occ_limit,
            cyber_agg_limit=original.cyber_agg_limit,
            cyber_premium=original.cyber_premium,
            cyber_renewal_date=original.cyber_renewal_date,
            auto_carrier=original.auto_carrier,
            auto_agency=original.auto_agency,
            auto_occ_limit=original.auto_occ_limit,
            auto_agg_limit=original.auto_agg_limit,
            auto_premium=original.auto_premium,
            auto_renewal_date=original.auto_renewal_date,
            epli_carrier=original.epli_carrier,
            epli_agency=original.epli_agency,
            epli_occ_limit=original.epli_occ_limit,
            epli_agg_limit=original.epli_agg_limit,
            epli_premium=original.epli_premium,
            epli_renewal_date=original.epli_renewal_date,
            nydbl_carrier=original.nydbl_carrier,
            nydbl_agency=original.nydbl_agency,
            nydbl_occ_limit=original.nydbl_occ_limit,
            nydbl_agg_limit=original.nydbl_agg_limit,
            nydbl_premium=original.nydbl_premium,
            nydbl_renewal_date=original.nydbl_renewal_date,
            surety_carrier=original.surety_carrier,
            surety_agency=original.surety_agency,
            surety_occ_limit=original.surety_occ_limit,
            surety_agg_limit=original.surety_agg_limit,
            surety_premium=original.surety_premium,
            surety_renewal_date=original.surety_renewal_date,
            product_liability_carrier=original.product_liability_carrier,
            product_liability_agency=original.product_liability_agency,
            product_liability_occ_limit=original.product_liability_occ_limit,
            product_liability_agg_limit=original.product_liability_agg_limit,
            product_liability_premium=original.product_liability_premium,
            product_liability_renewal_date=original.product_liability_renewal_date,
            flood_carrier=original.flood_carrier,
            flood_agency=original.flood_agency,
            flood_occ_limit=original.flood_occ_limit,
            flood_agg_limit=original.flood_agg_limit,
            flood_premium=original.flood_premium,
            flood_renewal_date=original.flood_renewal_date,
            crime_carrier=original.crime_carrier,
            crime_agency=original.crime_agency,
            crime_occ_limit=original.crime_occ_limit,
            crime_agg_limit=original.crime_agg_limit,
            crime_premium=original.crime_premium,
            crime_renewal_date=original.crime_renewal_date,
            directors_officers_carrier=original.directors_officers_carrier,
            directors_officers_agency=original.directors_officers_agency,
            directors_officers_occ_limit=original.directors_officers_occ_limit,
            directors_officers_agg_limit=original.directors_officers_agg_limit,
            directors_officers_premium=original.directors_officers_premium,
            directors_officers_renewal_date=original.directors_officers_renewal_date,
            fiduciary_carrier=original.fiduciary_carrier,
            fiduciary_agency=original.fiduciary_agency,
            fiduciary_occ_limit=original.fiduciary_occ_limit,
            fiduciary_agg_limit=original.fiduciary_agg_limit,
            fiduciary_premium=original.fiduciary_premium,
            fiduciary_renewal_date=original.fiduciary_renewal_date,
            inland_marine_carrier=original.inland_marine_carrier,
            inland_marine_agency=original.inland_marine_agency,
            inland_marine_occ_limit=original.inland_marine_occ_limit,
            inland_marine_agg_limit=original.inland_marine_agg_limit,
            inland_marine_premium=original.inland_marine_premium,
            inland_marine_renewal_date=original.inland_marine_renewal_date
        )

        # Copy remarks, outstanding_item, and agency columns for single-plan types
        for product in ['general_liability', 'property', 'bop', 'workers_comp', 'auto',
                       'epli', 'nydbl', 'surety', 'product_liability', 'flood',
                       'directors_officers', 'fiduciary', 'inland_marine']:
            setattr(new_commercial, f'{product}_remarks', getattr(original, f'{product}_remarks', None))
            setattr(new_commercial, f'{product}_outstanding_item', getattr(original, f'{product}_outstanding_item', None))

        session.add(new_commercial)
        session.flush()

        # Clone child CommercialPlan records
        for plan in original.commercial_plans:
            new_plan = CommercialPlan(
                commercial_insurance_id=new_commercial.id,
                plan_type=plan.plan_type,
                plan_number=plan.plan_number,
                carrier=plan.carrier,
                coverage_occ_limit=plan.coverage_occ_limit,
                coverage_agg_limit=plan.coverage_agg_limit,
                premium=plan.premium,
                renewal_date=plan.renewal_date,
                remarks=plan.remarks,
                outstanding_item=plan.outstanding_item
            )
            session.add(new_plan)

        session.commit()
        session.refresh(new_commercial)

        return jsonify({
            'message': 'Commercial insurance cloned successfully',
            'commercial': new_commercial.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error cloning commercial: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# INVOICE ENDPOINTS
# ===========================================================================

@app.route('/api/invoice/preview', methods=['POST'])
def invoice_preview():
    """Generate an invoice PDF and return it for preview."""
    session = Session()
    try:
        data = request.get_json()
        commercial_id = data.get('commercial_id')
        policy_types = data.get('policy_types', [])
        invoice_date = data.get('invoice_date', datetime.now().strftime('%Y-%m-%d'))

        if not commercial_id or not policy_types:
            return jsonify({'error': 'commercial_id and policy_types are required'}), 400

        commercial = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not commercial:
            return jsonify({'error': 'Commercial record not found'}), 404

        client = commercial.client
        if not client:
            return jsonify({'error': 'Client not found for this commercial record'}), 404

        commercial_data = commercial.to_dict()
        line_items = _collect_line_items(commercial_data, policy_types)

        if not line_items:
            return jsonify({'error': 'No active policies found for selected types'}), 400

        invoice_number = InvoiceSequence.next_number(session)

        # Build client address
        addr_parts = [client.address_line_1 or '']
        if client.address_line_2:
            addr_parts.append(client.address_line_2)
        city_state_zip = ', '.join(filter(None, [client.city, client.state]))
        if client.zip_code:
            city_state_zip += f' {client.zip_code}'
        addr_parts.append(city_state_zip)
        client_address = '\n'.join(filter(None, addr_parts))

        pdf_buf = generate_invoice_pdf(
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            client_name=client.client_name or '',
            client_address=client_address,
            client_tax_id=client.tax_id or '',
            line_items=line_items,
        )

        session.commit()

        return send_file(
            pdf_buf,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'Invoice_{invoice_number}_{(client.client_name or "Client").replace(" ", "_")}.pdf'
        )
    except Exception as e:
        session.rollback()
        logging.error(f"Error generating invoice preview: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/invoice/send', methods=['POST'])
def invoice_send():
    """Generate an invoice PDF and email it to the client."""
    session = Session()
    try:
        data = request.get_json()
        commercial_id = data.get('commercial_id')
        policy_types = data.get('policy_types', [])
        invoice_date = data.get('invoice_date', datetime.now().strftime('%Y-%m-%d'))
        to_email = data.get('to_email')
        cc_email = data.get('cc_email', '')
        subject = data.get('subject', '')

        if not commercial_id or not policy_types:
            return jsonify({'error': 'commercial_id and policy_types are required'}), 400
        if not to_email:
            return jsonify({'error': 'to_email is required'}), 400

        commercial = session.query(CommercialInsurance).filter_by(id=commercial_id).first()
        if not commercial:
            return jsonify({'error': 'Commercial record not found'}), 404

        client = commercial.client
        if not client:
            return jsonify({'error': 'Client not found for this commercial record'}), 404

        commercial_data = commercial.to_dict()
        line_items = _collect_line_items(commercial_data, policy_types)

        if not line_items:
            return jsonify({'error': 'No active policies found for selected types'}), 400

        invoice_number = InvoiceSequence.next_number(session)

        if not subject:
            subject = f'Invoice #{invoice_number} — Edison General Insurance Service'

        # Build client address
        addr_parts = [client.address_line_1 or '']
        if client.address_line_2:
            addr_parts.append(client.address_line_2)
        city_state_zip = ', '.join(filter(None, [client.city, client.state]))
        if client.zip_code:
            city_state_zip += f' {client.zip_code}'
        addr_parts.append(city_state_zip)
        client_address = '\n'.join(filter(None, addr_parts))

        pdf_buf = generate_invoice_pdf(
            invoice_number=invoice_number,
            invoice_date=invoice_date,
            client_name=client.client_name or '',
            client_address=client_address,
            client_tax_id=client.tax_id or '',
            line_items=line_items,
        )

        # Send email
        if SMTP_USE_TLS and (not SMTP_USERNAME or not SMTP_PASSWORD):
            return jsonify({'error': 'SMTP credentials not configured. Set SMTP_USERNAME and SMTP_PASSWORD environment variables.'}), 500

        client_name_clean = (client.client_name or 'Client').replace(' ', '_')
        filename = f'Invoice_{invoice_number}_{client_name_clean}.pdf'

        msg = MIMEMultipart()
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        if cc_email:
            msg['Cc'] = cc_email
        msg['Subject'] = subject

        body = (
            f"Dear {client.client_name or 'Valued Client'},\n\n"
            f"Please find attached your invoice #{invoice_number} from Edison General Insurance Service.\n\n"
            f"If you have any questions regarding this invoice, please contact us at 732-548-8700 "
            f"or email info@njgroups.com.\n\n"
            f"Thank you for your business.\n\n"
            f"Best regards,\n"
            f"Edison General Insurance Service\n"
            f"22 Meridian Road, Suite 16\n"
            f"Edison, NJ 08820"
        )
        msg.attach(MIMEText(body, 'plain'))

        # Attach PDF
        pdf_data = pdf_buf.read()
        attachment = MIMEBase('application', 'pdf')
        attachment.set_payload(pdf_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(attachment)

        # Send via SMTP
        recipients = [to_email]
        if cc_email:
            recipients.append(cc_email)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            if SMTP_USE_TLS:
                server.starttls()
            if SMTP_USERNAME and SMTP_PASSWORD:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, recipients, msg.as_string())

        session.commit()

        return jsonify({
            'message': 'Invoice sent successfully',
            'invoice_number': invoice_number
        }), 200

    except smtplib.SMTPException as e:
        session.rollback()
        logging.error(f"SMTP error sending invoice: {e}")
        return jsonify({'error': f'Email sending failed: {str(e)}'}), 500
    except Exception as e:
        session.rollback()
        logging.error(f"Error sending invoice: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# INDIVIDUAL ENDPOINTS
# ===========================================================================

def generate_individual_id(session):
    """Generate next individual ID in format 00-#######."""
    last = session.query(Individual).order_by(Individual.id.desc()).first()
    if last and last.individual_id:
        try:
            num = int(last.individual_id.split('-')[1]) + 1
        except (IndexError, ValueError):
            num = 1
    else:
        num = 1
    return f"00-{num:07d}"


@app.route('/api/individuals', methods=['GET'])
def get_individuals():
    """Get all individuals."""
    session = Session()
    try:
        individuals = session.query(Individual).all()
        return jsonify({
            'individuals': [i.to_dict() for i in individuals],
            'total': len(individuals)
        })
    except Exception as e:
        logging.error(f"Error fetching individuals: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/individuals', methods=['POST'])
def create_individual():
    """Create a new individual with auto-generated ID."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        zip_code = data.get('zip_code', '')
        if zip_code and not re.match(r'^\d{5}$', str(zip_code)):
            return jsonify({'error': 'Zip code must be exactly 5 digits'}), 400

        individual = Individual(
            individual_id=generate_individual_id(session),
            first_name=data.get('first_name'),
            last_name=data.get('last_name'),
            email=data.get('email'),
            phone_number=data.get('phone_number'),
            address_line_1=data.get('address_line_1'),
            address_line_2=data.get('address_line_2'),
            city=data.get('city'),
            state=data.get('state'),
            zip_code=data.get('zip_code'),
            status=data.get('status', 'Active')
        )

        session.add(individual)
        session.commit()

        return jsonify({
            'message': 'Individual created successfully',
            'individual': individual.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error creating individual: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/individuals/<int:individual_id>', methods=['PUT'])
def update_individual(individual_id):
    """Update an individual."""
    session = Session()
    try:
        individual = session.query(Individual).filter_by(id=individual_id).first()
        if not individual:
            return jsonify({'error': 'Individual not found'}), 404

        data = request.get_json()

        zip_code = data.get('zip_code', individual.zip_code)
        if zip_code and not re.match(r'^\d{5}$', str(zip_code)):
            return jsonify({'error': 'Zip code must be exactly 5 digits'}), 400

        individual.first_name = data.get('first_name', individual.first_name)
        individual.last_name = data.get('last_name', individual.last_name)
        individual.email = data.get('email', individual.email)
        individual.phone_number = data.get('phone_number', individual.phone_number)
        individual.address_line_1 = data.get('address_line_1', individual.address_line_1)
        individual.address_line_2 = data.get('address_line_2', individual.address_line_2)
        individual.city = data.get('city', individual.city)
        individual.state = data.get('state', individual.state)
        individual.zip_code = data.get('zip_code', individual.zip_code)
        individual.status = data.get('status', individual.status)

        session.commit()

        return jsonify({
            'message': 'Individual updated successfully',
            'individual': individual.to_dict()
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error updating individual: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/individuals/<int:individual_id>', methods=['DELETE'])
def delete_individual(individual_id):
    """Delete an individual."""
    session = Session()
    try:
        individual = session.query(Individual).filter_by(id=individual_id).first()
        if not individual:
            return jsonify({'error': 'Individual not found'}), 404

        session.delete(individual)
        session.commit()

        return jsonify({'message': 'Individual deleted successfully'}), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error deleting individual: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# PERSONAL INSURANCE ENDPOINTS
# ===========================================================================

# Field definitions per personal insurance product type
PERSONAL_FIELDS = {
    'personal_auto': {
        'text': ['carrier'],
        'limit': ['bi_occ_limit', 'bi_agg_limit', 'pd_limit'],
        'premium': ['premium'],
        'date': ['renewal_date'],
        'other': ['outstanding_item', 'remarks']
    },
    'homeowners': {
        'text': ['carrier'],
        'limit': ['dwelling_limit', 'liability_limit'],
        'premium': ['premium'],
        'date': ['renewal_date'],
        'other': ['outstanding_item', 'remarks']
    },
    'personal_umbrella': {
        'text': ['carrier'],
        'limit': ['liability_limit'],
        'premium': ['deductible', 'premium'],
        'date': ['renewal_date'],
        'other': ['outstanding_item', 'remarks']
    },
    'event': {
        'text': ['carrier', 'type', 'location'],
        'limit': [],
        'premium': ['entry_fee', 'premium'],
        'date': ['start_date', 'end_date'],
        'integer': ['audience_count'],
        'other': ['outstanding_item', 'remarks']
    },
    'visitors_medical': {
        'text': ['carrier', 'destination_country'],
        'limit': [],
        'premium': ['premium'],
        'date': ['start_date', 'end_date'],
        'other': ['outstanding_item', 'remarks']
    }
}


@app.route('/api/personal', methods=['GET'])
def get_personal():
    """Get all personal insurance records."""
    session = Session()
    try:
        records = session.query(PersonalInsurance).all()
        return jsonify({
            'personal': [r.to_dict() for r in records],
            'total': len(records)
        })
    except Exception as e:
        logging.error(f"Error fetching personal insurance: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/personal/<int:personal_id>', methods=['GET'])
def get_personal_by_id(personal_id):
    """Get single personal insurance record."""
    session = Session()
    try:
        record = session.query(PersonalInsurance).filter_by(id=personal_id).first()
        if not record:
            return jsonify({'error': 'Personal insurance not found'}), 404
        return jsonify({'personal': record.to_dict()})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/personal', methods=['POST'])
def create_personal():
    """Create new personal insurance record."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        personal = PersonalInsurance(
            individual_id=data.get('individual_id'),
        )

        # Set fields for each product type
        for product in PERSONAL_INSURANCE_PRODUCTS:
            fields = PERSONAL_FIELDS[product]
            for f in fields.get('text', []) + fields.get('limit', []):
                setattr(personal, f'{product}_{f}', data.get(f'{product}_{f}') or None)
            for f in fields.get('premium', []):
                setattr(personal, f'{product}_{f}', parse_premium(data.get(f'{product}_{f}')))
            for f in fields.get('date', []):
                setattr(personal, f'{product}_{f}', parse_date(data.get(f'{product}_{f}')))
            for f in fields.get('integer', []):
                val = data.get(f'{product}_{f}')
                setattr(personal, f'{product}_{f}', int(val) if val else None)
            for f in fields.get('other', []):
                setattr(personal, f'{product}_{f}', data.get(f'{product}_{f}') or None)

        session.add(personal)
        session.flush()

        # Save homeowners policies if provided
        if 'homeowners_policies_list' in data:
            save_homeowners_policies(session, personal, data['homeowners_policies_list'])

        session.commit()
        session.refresh(personal)

        return jsonify({
            'message': 'Personal insurance created successfully',
            'personal': personal.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error creating personal insurance: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/personal/<int:personal_id>', methods=['PUT'])
def update_personal(personal_id):
    """Update personal insurance record."""
    session = Session()
    try:
        personal = session.query(PersonalInsurance).filter_by(id=personal_id).first()
        if not personal:
            return jsonify({'error': 'Personal insurance not found'}), 404

        data = request.get_json()

        personal.individual_id = data.get('individual_id', personal.individual_id)

        for product in PERSONAL_INSURANCE_PRODUCTS:
            fields = PERSONAL_FIELDS[product]
            for f in fields.get('text', []) + fields.get('limit', []):
                key = f'{product}_{f}'
                if key in data:
                    setattr(personal, key, data.get(key) or None)
            for f in fields.get('premium', []):
                key = f'{product}_{f}'
                if key in data:
                    setattr(personal, key, parse_premium(data.get(key)))
            for f in fields.get('date', []):
                key = f'{product}_{f}'
                if key in data:
                    setattr(personal, key, parse_date(data.get(key)))
            for f in fields.get('integer', []):
                key = f'{product}_{f}'
                if key in data:
                    val = data.get(key)
                    setattr(personal, key, int(val) if val else None)
            for f in fields.get('other', []):
                key = f'{product}_{f}'
                if key in data:
                    setattr(personal, key, data.get(key) or None)

        # Save homeowners policies if provided
        if 'homeowners_policies_list' in data:
            save_homeowners_policies(session, personal, data['homeowners_policies_list'])

        session.commit()
        session.refresh(personal)

        return jsonify({
            'message': 'Personal insurance updated successfully',
            'personal': personal.to_dict()
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error updating personal insurance: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/personal/<int:personal_id>', methods=['DELETE'])
def delete_personal(personal_id):
    """Delete personal insurance record."""
    session = Session()
    try:
        personal = session.query(PersonalInsurance).filter_by(id=personal_id).first()
        if not personal:
            return jsonify({'error': 'Personal insurance not found'}), 404

        session.delete(personal)
        session.commit()

        return jsonify({'message': 'Personal insurance deleted successfully'}), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error deleting personal insurance: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/personal/<int:personal_id>/clone', methods=['POST'])
def clone_personal(personal_id):
    """Clone personal insurance record."""
    session = Session()
    try:
        original = session.query(PersonalInsurance).filter_by(id=personal_id).first()
        if not original:
            return jsonify({'error': 'Personal insurance not found'}), 404

        new_personal = PersonalInsurance(individual_id=original.individual_id)

        # Copy all product fields
        for product in PERSONAL_INSURANCE_PRODUCTS:
            fields = PERSONAL_FIELDS[product]
            all_fields = (fields.get('text', []) + fields.get('limit', []) +
                         fields.get('premium', []) + fields.get('date', []) +
                         fields.get('integer', []) + fields.get('other', []))
            for f in all_fields:
                setattr(new_personal, f'{product}_{f}', getattr(original, f'{product}_{f}', None))

        session.add(new_personal)
        session.commit()
        session.refresh(new_personal)

        return jsonify({
            'message': 'Personal insurance cloned successfully',
            'personal': new_personal.to_dict()
        }), 201
    except Exception as e:
        session.rollback()
        logging.error(f"Error cloning personal insurance: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# DASHBOARD ANALYTICS ENDPOINTS
# ===========================================================================

@app.route('/api/dashboard/renewals', methods=['GET'])
def get_dashboard_renewals():
    """Get renewal data for dashboard (next 12 months from both benefits and commercial)."""
    session = Session()
    try:
        today = datetime.now().date()
        twelve_months = today + timedelta(days=365)

        renewals = []

        # Get employee benefits renewals
        multi_plan_type_names = {'medical': 'Medical', 'dental': 'Dental', 'vision': 'Vision', 'life_adnd': 'Life & AD&D'}
        benefits = session.query(EmployeeBenefit).all()
        for benefit in benefits:
            # Multi-plan types: read from benefit_plans child table
            for plan in benefit.plans:
                if plan.renewal_date and today <= plan.renewal_date <= twelve_months:
                    type_name = multi_plan_type_names.get(plan.plan_type, plan.plan_type)
                    label = f"{type_name} Plan {plan.plan_number}" if plan.plan_number > 1 else type_name
                    renewals.append({
                        'type': 'benefits',
                        'policy_type': label,
                        'renewal_date': plan.renewal_date.isoformat(),
                        'client_name': benefit.client.client_name if benefit.client else None,
                        'tax_id': benefit.tax_id,
                        'carrier': plan.carrier
                    })

            # Single-plan types: read from flat fields
            single_plan_fields = [
                ('ltd_renewal_date', 'LTD'),
                ('std_renewal_date', 'STD'),
                ('k401_renewal_date', '401K'),
                ('critical_illness_renewal_date', 'Critical Illness'),
                ('accident_renewal_date', 'Accident'),
                ('hospital_renewal_date', 'Hospital'),
                ('voluntary_life_renewal_date', 'Voluntary Life')
            ]

            for field_name, policy_type in single_plan_fields:
                renewal_date = getattr(benefit, field_name)
                if renewal_date and today <= renewal_date <= twelve_months:
                    carrier = getattr(benefit, field_name.replace('_renewal_date', '_carrier'), None)
                    renewals.append({
                        'type': 'benefits',
                        'policy_type': policy_type,
                        'renewal_date': renewal_date.isoformat(),
                        'client_name': benefit.client.client_name if benefit.client else None,
                        'tax_id': benefit.tax_id,
                        'carrier': carrier
                    })

        # Get commercial renewals
        multi_plan_commercial_names = {'umbrella': 'Umbrella', 'professional_eo': 'Professional E&O', 'cyber': 'Cyber', 'crime': 'Crime'}
        commercial = session.query(CommercialInsurance).all()
        for comm in commercial:
            # Multi-plan types: read from commercial_plans child table
            for plan in comm.commercial_plans:
                if plan.renewal_date and today <= plan.renewal_date <= twelve_months:
                    type_name = multi_plan_commercial_names.get(plan.plan_type, plan.plan_type)
                    label = f"{type_name} Plan {plan.plan_number}" if plan.plan_number > 1 else type_name
                    renewals.append({
                        'type': 'commercial',
                        'policy_type': label,
                        'renewal_date': plan.renewal_date.isoformat(),
                        'client_name': comm.client.client_name if comm.client else None,
                        'tax_id': comm.tax_id,
                        'carrier': plan.carrier
                    })

            # Single-plan types: read from flat fields
            single_plan_fields = [
                ('general_liability_renewal_date', 'General Liability'),
                ('property_renewal_date', 'Property'),
                ('bop_renewal_date', 'BOP'),
                ('workers_comp_renewal_date', 'Workers Comp'),
                ('auto_renewal_date', 'Auto'),
                ('epli_renewal_date', 'EPLI'),
                ('nydbl_renewal_date', 'NYDBL'),
                ('surety_renewal_date', 'Surety'),
                ('product_liability_renewal_date', 'Product Liability'),
                ('flood_renewal_date', 'Flood'),
                ('directors_officers_renewal_date', 'D&O'),
                ('fiduciary_renewal_date', 'Fiduciary'),
                ('inland_marine_renewal_date', 'Inland Marine')
            ]

            for field_name, policy_type in single_plan_fields:
                renewal_date = getattr(comm, field_name)
                if renewal_date and today <= renewal_date <= twelve_months:
                    renewals.append({
                        'type': 'commercial',
                        'policy_type': policy_type,
                        'renewal_date': renewal_date.isoformat(),
                        'client_name': comm.client.client_name if comm.client else None,
                        'tax_id': comm.tax_id,
                        'carrier': getattr(comm, field_name.replace('_renewal_date', '_carrier'), None)
                    })

        # Get personal insurance renewals
        personal_renewal_fields = [
            ('personal_auto_renewal_date', 'Personal Auto'),
            ('homeowners_renewal_date', 'Homeowners'),
            ('personal_umbrella_renewal_date', 'Personal Umbrella'),
            ('event_start_date', 'Event Insurance'),
            ('visitors_medical_start_date', 'Visitors Medical')
        ]
        personal_records = session.query(PersonalInsurance).all()
        for pers in personal_records:
            for field_name, policy_type in personal_renewal_fields:
                renewal_date = getattr(pers, field_name)
                if renewal_date and today <= renewal_date <= twelve_months:
                    carrier_field = field_name.replace('_renewal_date', '_carrier').replace('_start_date', '_carrier')
                    ind = pers.individual
                    ind_name = f"{ind.first_name or ''} {ind.last_name or ''}".strip() if ind else None
                    renewals.append({
                        'type': 'personal',
                        'policy_type': policy_type,
                        'renewal_date': renewal_date.isoformat(),
                        'client_name': ind_name,
                        'tax_id': pers.individual_id,
                        'carrier': getattr(pers, carrier_field, None)
                    })

        # Sort by renewal date
        renewals.sort(key=lambda x: x['renewal_date'])

        return jsonify({
            'renewals': renewals,
            'total': len(renewals)
        }), 200
    except Exception as e:
        logging.error(f"Error fetching dashboard renewals: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/dashboard/cross-sell', methods=['GET'])
def get_cross_sell_opportunities():
    """Get cross-sell opportunities (clients with only one type of insurance)."""
    session = Session()
    try:
        all_clients = session.query(Client).all()

        clients_with_benefits_only = []
        clients_with_commercial_only = []

        for client in all_clients:
            has_benefits = len(client.employee_benefits) > 0
            has_commercial = len(client.commercial_insurance) > 0

            if has_benefits and not has_commercial:
                clients_with_benefits_only.append({
                    'tax_id': client.tax_id,
                    'client_name': client.client_name,
                    'contact_person': client.contact_person,
                    'email': client.email
                })
            elif has_commercial and not has_benefits:
                clients_with_commercial_only.append({
                    'tax_id': client.tax_id,
                    'client_name': client.client_name,
                    'contact_person': client.contact_person,
                    'email': client.email
                })

        return jsonify({
            'benefits_only': clients_with_benefits_only,
            'commercial_only': clients_with_commercial_only,
            'total_opportunities': len(clients_with_benefits_only) + len(clients_with_commercial_only)
        }), 200
    except Exception as e:
        logging.error(f"Error fetching cross-sell opportunities: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# EXCEL EXPORT/IMPORT ENDPOINTS
# ===========================================================================

@app.route('/api/export', methods=['GET'])
def export_to_excel():
    """Export all data to Excel in the same format as Data Sheet.xlsx."""
    session = Session()
    try:
        wb = Workbook()

        # Header styling
        header_font = Font(bold=True)
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== CLIENTS SHEET ==========
        ws_clients = wb.active
        ws_clients.title = "Clients"

        # Determine max number of contacts across all clients
        clients = session.query(Client).all()
        max_contacts = 1
        for client in clients:
            count = len(client.contacts) if client.contacts else 0
            if count > max_contacts:
                max_contacts = count
        # Fall back to legacy flat fields if no contacts exist
        if max_contacts == 0:
            max_contacts = 1

        # Row 1: Section headers
        client_fixed_cols = 7  # Tax ID, Client Name, DBA, Industry, Status, Gross Revenue, Total EEs
        contact_cols_per = 9   # Contact Person, Email, Phone, Ext, Addr1, Addr2, City, State, Zip
        contact_start = client_fixed_cols + 1
        for i in range(max_contacts):
            start_col = contact_start + i * contact_cols_per
            end_col = start_col + contact_cols_per - 1
            ws_clients.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws_clients.cell(row=1, column=start_col, value=f'Contact {i + 1}' if max_contacts > 1 else 'Contact Info')
            cell.font = section_font
            cell.fill = section_fill

        # Row 2: Column headers
        client_headers = ['Tax ID', 'Client Name', 'DBA', 'Industry', 'Status', 'Gross Revenue', 'Total EEs']
        contact_headers = ['Contact Person', 'Email', 'Phone Number', 'Ext', 'Address Line 1', 'Address Line 2', 'City', 'State', 'Zip Code']
        for i in range(max_contacts):
            client_headers.extend(contact_headers)
        for col, header in enumerate(client_headers, 1):
            cell = ws_clients.cell(row=2, column=col, value=header)
            cell.font = header_font

        # Client data
        for row_idx, client in enumerate(clients, 3):
            ws_clients.cell(row=row_idx, column=1, value=client.tax_id)
            ws_clients.cell(row=row_idx, column=2, value=client.client_name)
            ws_clients.cell(row=row_idx, column=3, value=client.dba)
            ws_clients.cell(row=row_idx, column=4, value=client.industry)
            ws_clients.cell(row=row_idx, column=5, value=client.status)
            ws_clients.cell(row=row_idx, column=6, value=float(client.gross_revenue) if client.gross_revenue else None)
            ws_clients.cell(row=row_idx, column=7, value=client.total_ees)

            # Write contacts
            contact_list = list(client.contacts) if client.contacts else []
            # Backward compat: if no contacts, use flat fields
            if not contact_list and (client.contact_person or client.email or client.phone_number):
                contact_list = [type('obj', (object,), {
                    'contact_person': client.contact_person, 'email': client.email,
                    'phone_number': client.phone_number, 'phone_extension': None,
                    'address_line_1': client.address_line_1, 'address_line_2': client.address_line_2,
                    'city': client.city, 'state': client.state, 'zip_code': client.zip_code
                })()]
            for ci, contact in enumerate(contact_list):
                base = contact_start + ci * contact_cols_per
                ws_clients.cell(row=row_idx, column=base, value=getattr(contact, 'contact_person', None))
                ws_clients.cell(row=row_idx, column=base + 1, value=getattr(contact, 'email', None))
                ws_clients.cell(row=row_idx, column=base + 2, value=getattr(contact, 'phone_number', None))
                ws_clients.cell(row=row_idx, column=base + 3, value=getattr(contact, 'phone_extension', None))
                ws_clients.cell(row=row_idx, column=base + 4, value=getattr(contact, 'address_line_1', None))
                ws_clients.cell(row=row_idx, column=base + 5, value=getattr(contact, 'address_line_2', None))
                ws_clients.cell(row=row_idx, column=base + 6, value=getattr(contact, 'city', None))
                ws_clients.cell(row=row_idx, column=base + 7, value=getattr(contact, 'state', None))
                zip_cell = ws_clients.cell(row=row_idx, column=base + 8, value=getattr(contact, 'zip_code', None))
                zip_cell.number_format = '@'

        # ========== INDIVIDUALS SHEET ==========
        ws_individuals = wb.create_sheet("Individuals")

        individuals_list = session.query(Individual).all()

        individual_headers = ['Individual ID', 'First Name', 'Last Name', 'Email',
                              'Phone Number', 'Address Line 1', 'Address Line 2',
                              'City', 'State', 'Zip Code', 'Status']

        # Row 1: empty (consistent with other sheets that have section headers)
        # Row 2: Column headers
        for col, header in enumerate(individual_headers, 1):
            cell = ws_individuals.cell(row=2, column=col, value=header)
            cell.font = header_font

        for row_idx, ind in enumerate(individuals_list, 3):
            ws_individuals.cell(row=row_idx, column=1, value=ind.individual_id)
            ws_individuals.cell(row=row_idx, column=2, value=ind.first_name)
            ws_individuals.cell(row=row_idx, column=3, value=ind.last_name)
            ws_individuals.cell(row=row_idx, column=4, value=ind.email)
            ws_individuals.cell(row=row_idx, column=5, value=ind.phone_number)
            ws_individuals.cell(row=row_idx, column=6, value=ind.address_line_1)
            ws_individuals.cell(row=row_idx, column=7, value=ind.address_line_2)
            ws_individuals.cell(row=row_idx, column=8, value=ind.city)
            ws_individuals.cell(row=row_idx, column=9, value=ind.state)
            zip_cell = ws_individuals.cell(row=row_idx, column=10, value=ind.zip_code)
            zip_cell.number_format = '@'
            ws_individuals.cell(row=row_idx, column=11, value=ind.status)

        # ========== EMPLOYEE BENEFITS SHEET ==========
        ws_benefits = wb.create_sheet("Employee Benefits")

        # Load benefits and determine max plan counts for dynamic columns
        benefits = session.query(EmployeeBenefit).all()

        multi_plan_defs = [
            ('medical', 'MEDICAL'),
            ('dental', 'DENTAL'),
            ('vision', 'VISION'),
            ('life_adnd', 'Life & AD&D')
        ]

        # Determine max number of plans per type across all benefits
        max_plans = {}
        for plan_type, _ in multi_plan_defs:
            max_count = 1  # At least 1 column per type
            for benefit in benefits:
                count = len([p for p in benefit.plans if p.plan_type == plan_type])
                if count > max_count:
                    max_count = count
            max_plans[plan_type] = max_count

        # Build dynamic headers and section spans
        # Fixed columns: Tax ID, Client Name, Status, Outstanding Item, Remarks (cols 1-5)
        # Medical global fields: Form Fire Code, Enrollment POC, Other Broker, Funding,
        #   # of Emp at renewal, Waiting Period, Deductible Accumulation, Previous Carrier, Cobra Administrator (9 cols)
        # Then dynamic multi-plan columns (carrier + renewal per plan)
        # Then single-plan types (7 types x 2 cols = 14 cols)
        # Then 1095 (2 cols)

        benefit_headers = ['Tax ID', 'Client Name ', 'Parent Client',
                           'Form Fire Code', 'Assigned To', 'Other Broker', 'Funding',
                           '# of Emp at renewal', 'Enrolled EEs', 'Waiting Period', 'Deductible Accumulation',
                           'Previous Carrier', 'Cobra Administrator']
        # Track col position (1-based) — after fixed cols
        col_pos = len(benefit_headers) + 1  # Next col to use

        benefit_sections = [(1, 3, ''), (4, 13, 'MEDICAL GLOBAL')]

        # Multi-plan type dynamic columns (carrier, renewal, waiting_period, remarks, outstanding_item per plan)
        multi_plan_col_map = {}  # plan_type -> start_col (for data writing)
        for plan_type, label in multi_plan_defs:
            n = max_plans[plan_type]
            start = col_pos
            for i in range(1, n + 1):
                suffix = f' {i}' if n > 1 else ''
                benefit_headers.append(f'{label} Carrier{suffix}')
                benefit_headers.append(f'{label} Renewal Date{suffix}')
                benefit_headers.append(f'{label} Waiting Period{suffix}')
                benefit_headers.append(f'{label} Remarks{suffix}')
                benefit_headers.append(f'{label} Outstanding Item{suffix}')
            multi_plan_col_map[plan_type] = start
            end = col_pos + n * 5 - 1
            benefit_sections.append((start, end, f'{label} PLANS'))
            col_pos += n * 5

        # Single-plan types (renewal, carrier, remarks, outstanding_item per type)
        single_plan_types = [
            ('ltd', 'LTD'), ('std', 'STD'), ('k401', '401K'),
            ('critical_illness', 'Critical Illness'), ('accident', 'Accident'),
            ('hospital', 'Hospital'), ('voluntary_life', 'Voluntary Life')
        ]
        single_plan_col_start = col_pos
        for prefix, label in single_plan_types:
            benefit_headers.append(f'{label} Renewal Date')
            benefit_headers.append(f'{label} Carrier')
            benefit_headers.append(f'{label} Remarks')
            benefit_headers.append(f'{label} Outstanding Item')
            benefit_sections.append((col_pos, col_pos + 3, f'{label.upper()} PLANS'))
            col_pos += 4

        # 1095
        benefit_headers.extend(['Employer Contribution %', 'Employee Contribution %'])
        benefit_sections.append((col_pos, col_pos + 1, '1095'))

        # Write section headers (Row 1)
        for start_col, end_col, title in benefit_sections:
            if title:
                if start_col != end_col:
                    ws_benefits.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws_benefits.cell(row=1, column=start_col, value=title)
                cell.font = section_font
                cell.fill = section_fill

        # Write column headers (Row 2)
        for col, header in enumerate(benefit_headers, 1):
            cell = ws_benefits.cell(row=2, column=col, value=header)
            cell.font = header_font

        # Write benefits data
        for row_idx, benefit in enumerate(benefits, 3):
            client_name = benefit.client.client_name if benefit.client else None
            c = 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.tax_id); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=client_name); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.parent_client); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.form_fire_code); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.enrollment_poc); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=None); c += 1  # Other Broker
            ws_benefits.cell(row=row_idx, column=c, value=benefit.funding); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.num_employees_at_renewal); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.enrolled_ees); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.waiting_period); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.deductible_accumulation); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.previous_carrier); c += 1
            ws_benefits.cell(row=row_idx, column=c, value=benefit.cobra_carrier); c += 1

            # Multi-plan types: write dynamic columns (carrier, renewal, waiting_period, remarks, outstanding_item per plan)
            plans_by_type = {}
            for plan in benefit.plans:
                plans_by_type.setdefault(plan.plan_type, []).append(plan)
            for plan_type, _ in multi_plan_defs:
                type_plans = sorted(plans_by_type.get(plan_type, []), key=lambda p: p.plan_number)
                start = multi_plan_col_map[plan_type]
                for i in range(max_plans[plan_type]):
                    if i < len(type_plans):
                        ws_benefits.cell(row=row_idx, column=start + i * 5, value=type_plans[i].carrier)
                        ws_benefits.cell(row=row_idx, column=start + i * 5 + 1, value=type_plans[i].renewal_date)
                        ws_benefits.cell(row=row_idx, column=start + i * 5 + 2, value=type_plans[i].waiting_period)
                        ws_benefits.cell(row=row_idx, column=start + i * 5 + 3, value=type_plans[i].remarks)
                        ws_benefits.cell(row=row_idx, column=start + i * 5 + 4, value=type_plans[i].outstanding_item)
                    # else leave blank

            # Single-plan types (renewal, carrier, remarks, outstanding_item)
            sc = single_plan_col_start
            for prefix, label in single_plan_types:
                ws_benefits.cell(row=row_idx, column=sc, value=getattr(benefit, f'{prefix}_renewal_date'))
                ws_benefits.cell(row=row_idx, column=sc + 1, value=getattr(benefit, f'{prefix}_carrier'))
                ws_benefits.cell(row=row_idx, column=sc + 2, value=getattr(benefit, f'{prefix}_remarks', None))
                ws_benefits.cell(row=row_idx, column=sc + 3, value=getattr(benefit, f'{prefix}_outstanding_item', None))
                sc += 4

            # 1095
            ws_benefits.cell(row=row_idx, column=sc, value=benefit.employer_contribution)
            ws_benefits.cell(row=row_idx, column=sc + 1, value=benefit.employee_contribution)

        # ========== COMMERCIAL SHEET ==========
        ws_commercial = wb.create_sheet("Commercial")

        commercial_records = session.query(CommercialInsurance).all()

        # Single-plan types (flat fields, 5 cols each: carrier, limit, premium, renewal, flag)
        commercial_single_plan_defs = [
            ('general_liability', 'Commercial General Liability'),
            ('property', 'Commercial Property'),
            ('bop', 'Business Owners Policy'),
            ('workers_comp', 'Workers Compensation'),
            ('auto', 'Commercial Auto'),
            ('epli', 'EPLI'),
            ('nydbl', 'NYDBL'),
            ('surety', 'Surety Bond'),
            ('product_liability', 'Product Liability'),
            ('flood', 'Flood'),
            ('directors_officers', 'Directors & Officers'),
            ('fiduciary', 'Fiduciary Bond'),
            ('inland_marine', 'Inland Marine')
        ]

        # Multi-plan types (dynamic cols from CommercialPlan child records)
        commercial_multi_plan_defs = [
            ('umbrella', 'Umbrella Liability'),
            ('professional_eo', 'Professional or E&O'),
            ('cyber', 'Cyber Liability'),
            ('crime', 'Crime or Fidelity Bond')
        ]

        # Determine max number of plans per multi-plan type
        comm_max_plans = {}
        for plan_type, _ in commercial_multi_plan_defs:
            max_count = 1
            for comm in commercial_records:
                count = len([p for p in comm.commercial_plans if p.plan_type == plan_type])
                if count > max_count:
                    max_count = count
            comm_max_plans[plan_type] = max_count

        # Build headers dynamically
        commercial_headers = ['Tax ID', 'Client Name ', 'Parent Client', 'Assigned To']
        commercial_sections = [(1, 4, '')]
        col_pos = 5

        # Single-plan type columns
        GL_ENDORSEMENT_COLS = ['Endorsement BOP', 'Endorsement Marine', 'Endorsement Foreign', 'Endorsement Molestation', 'Endorsement Staffing', 'Endorsement Accidental & Medical', 'Endorsement Liquor Liability']
        BOP_PROPERTY_COLS = ['Building Limit', 'Personal Property']
        comm_single_col_start = col_pos
        comm_single_col_sizes = {}  # prefix -> number of cols
        for prefix, label in commercial_single_plan_defs:
            if prefix == 'property':
                base_cols = ['Carrier', 'Agency', 'Policy Number', 'Building Limit', 'Personal Property', 'Premium', 'Renewal Date', 'Remarks', 'Outstanding Item']
            else:
                base_cols = ['Carrier', 'Agency', 'Policy Number', 'Occ Limit (M)', 'Agg Limit (M)', 'Premium', 'Renewal Date', 'Remarks', 'Outstanding Item']
            if prefix == 'general_liability':
                base_cols.extend(GL_ENDORSEMENT_COLS)
            elif prefix == 'bop':
                base_cols.extend(BOP_PROPERTY_COLS)
            commercial_headers.extend(base_cols)
            ncols = len(base_cols)
            commercial_sections.append((col_pos, col_pos + ncols - 1, label))
            comm_single_col_sizes[prefix] = ncols
            col_pos += ncols

        # Multi-plan type columns
        EO_ENDORSEMENT_COLS = ['Endorsement Tech E&O', 'Endorsement Staffing', 'Endorsement Allied Healthcare', 'Endorsement Medical Malpractice']
        comm_multi_col_map = {}
        comm_multi_plan_col_size = {}  # plan_type -> cols per plan
        for plan_type, label in commercial_multi_plan_defs:
            n = comm_max_plans[plan_type]
            start = col_pos
            base_plan_cols = ['Carrier', 'Agency', 'Policy Number', 'Occ Limit', 'Agg Limit', 'Premium', 'Renewal Date', 'Remarks', 'Outstanding Item']
            if plan_type == 'professional_eo':
                base_plan_cols.extend(EO_ENDORSEMENT_COLS)
            cols_per_plan = len(base_plan_cols)
            comm_multi_plan_col_size[plan_type] = cols_per_plan
            for i in range(1, n + 1):
                suffix = f' {i}' if n > 1 else ''
                commercial_headers.extend([f'{h}{suffix}' for h in base_plan_cols])
            comm_multi_col_map[plan_type] = start
            end = col_pos + n * cols_per_plan - 1
            commercial_sections.append((start, end, label))
            col_pos += n * cols_per_plan

        # Write section headers (Row 1)
        for start_col, end_col, title in commercial_sections:
            if title:
                if start_col != end_col:
                    ws_commercial.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws_commercial.cell(row=1, column=start_col, value=title)
                cell.font = section_font
                cell.fill = section_fill

        # Write column headers (Row 2)
        for col, header in enumerate(commercial_headers, 1):
            cell = ws_commercial.cell(row=2, column=col, value=header)
            cell.font = header_font

        # Write commercial data
        for row_idx, comm in enumerate(commercial_records, 3):
            client_name = comm.client.client_name if comm.client else None
            ws_commercial.cell(row=row_idx, column=1, value=comm.tax_id)
            ws_commercial.cell(row=row_idx, column=2, value=client_name)
            ws_commercial.cell(row=row_idx, column=3, value=comm.parent_client)
            ws_commercial.cell(row=row_idx, column=4, value=comm.assigned_to)

            # Single-plan types
            sc = comm_single_col_start
            for prefix, label in commercial_single_plan_defs:
                ws_commercial.cell(row=row_idx, column=sc, value=getattr(comm, f'{prefix}_carrier', None) or '')
                ws_commercial.cell(row=row_idx, column=sc + 1, value=getattr(comm, f'{prefix}_agency', None))
                ws_commercial.cell(row=row_idx, column=sc + 2, value=getattr(comm, f'{prefix}_policy_number', None))
                ws_commercial.cell(row=row_idx, column=sc + 3, value=getattr(comm, f'{prefix}_occ_limit', None) or '')
                ws_commercial.cell(row=row_idx, column=sc + 4, value=getattr(comm, f'{prefix}_agg_limit', None) or '')
                premium = getattr(comm, f'{prefix}_premium', None)
                ws_commercial.cell(row=row_idx, column=sc + 5, value=float(premium) if premium else 0)
                ws_commercial.cell(row=row_idx, column=sc + 6, value=getattr(comm, f'{prefix}_renewal_date', None) or '')
                ws_commercial.cell(row=row_idx, column=sc + 7, value=getattr(comm, f'{prefix}_remarks', None))
                ws_commercial.cell(row=row_idx, column=sc + 8, value=getattr(comm, f'{prefix}_outstanding_item', None))
                if prefix == 'general_liability':
                    ws_commercial.cell(row=row_idx, column=sc + 9, value='Yes' if getattr(comm, 'general_liability_endorsement_bop', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 10, value='Yes' if getattr(comm, 'general_liability_endorsement_marine', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 11, value='Yes' if getattr(comm, 'general_liability_endorsement_foreign', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 12, value='Yes' if getattr(comm, 'general_liability_endorsement_molestation', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 13, value='Yes' if getattr(comm, 'general_liability_endorsement_staffing', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 14, value='Yes' if getattr(comm, 'general_liability_endorsement_accidental_medical', False) else '')
                    ws_commercial.cell(row=row_idx, column=sc + 15, value='Yes' if getattr(comm, 'general_liability_endorsement_liquor_liability', False) else '')
                elif prefix == 'bop':
                    bl = getattr(comm, 'bop_building_limit', None)
                    pp = getattr(comm, 'bop_personal_property', None)
                    ws_commercial.cell(row=row_idx, column=sc + 9, value=float(bl) if bl else None)
                    ws_commercial.cell(row=row_idx, column=sc + 10, value=float(pp) if pp else None)
                sc += comm_single_col_sizes[prefix]

            # Multi-plan types
            plans_by_type = {}
            for plan in comm.commercial_plans:
                plans_by_type.setdefault(plan.plan_type, []).append(plan)
            for plan_type, _ in commercial_multi_plan_defs:
                type_plans = sorted(plans_by_type.get(plan_type, []), key=lambda p: p.plan_number)
                start = comm_multi_col_map[plan_type]
                cps = comm_multi_plan_col_size[plan_type]
                for i in range(comm_max_plans[plan_type]):
                    base = start + i * cps
                    if i < len(type_plans):
                        p = type_plans[i]
                        ws_commercial.cell(row=row_idx, column=base, value=p.carrier)
                        ws_commercial.cell(row=row_idx, column=base + 1, value=p.agency)
                        ws_commercial.cell(row=row_idx, column=base + 2, value=p.policy_number)
                        ws_commercial.cell(row=row_idx, column=base + 3, value=p.coverage_occ_limit)
                        ws_commercial.cell(row=row_idx, column=base + 4, value=p.coverage_agg_limit)
                        ws_commercial.cell(row=row_idx, column=base + 5, value=float(p.premium) if p.premium else 0)
                        ws_commercial.cell(row=row_idx, column=base + 6, value=p.renewal_date)
                        ws_commercial.cell(row=row_idx, column=base + 7, value=p.remarks)
                        ws_commercial.cell(row=row_idx, column=base + 8, value=p.outstanding_item)
                        if plan_type == 'professional_eo':
                            ws_commercial.cell(row=row_idx, column=base + 9, value='Yes' if p.endorsement_tech_eo else '')
                            ws_commercial.cell(row=row_idx, column=base + 10, value='Yes' if p.endorsement_staffing else '')
                            ws_commercial.cell(row=row_idx, column=base + 11, value='Yes' if p.endorsement_allied_healthcare else '')
                            ws_commercial.cell(row=row_idx, column=base + 12, value='Yes' if p.endorsement_medical_malpractice else '')

        # ========== PERSONAL INSURANCE SHEET ==========
        ws_personal = wb.create_sheet("Personal")

        personal_records = session.query(PersonalInsurance).all()

        personal_product_defs = [
            ('personal_auto', 'Personal Auto', ['Carrier', 'BI Occ Limit', 'BI Agg Limit', 'PD Limit', 'Premium', 'Renewal Date', 'Outstanding Item', 'Remarks']),
            ('homeowners', 'Homeowners', ['Carrier', 'Dwelling Limit', 'Liability Limit', 'Premium', 'Renewal Date', 'Outstanding Item', 'Remarks']),
            ('personal_umbrella', 'Personal Umbrella', ['Carrier', 'Liability Limit', 'Deductible', 'Premium', 'Renewal Date', 'Outstanding Item', 'Remarks']),
            ('event', 'Event Insurance', ['Carrier', 'Type of Event', 'Event Location', 'Start Date', 'End Date', 'Entry Fee', 'Audience Count', 'Premium', 'Outstanding Item', 'Remarks']),
            ('visitors_medical', 'Visitors Medical', ['Carrier', 'Start Date', 'End Date', 'Destination Country', 'Premium', 'Outstanding Item', 'Remarks']),
        ]

        # Field name mapping (column label -> db field suffix)
        personal_field_map = {
            'personal_auto': ['carrier', 'bi_occ_limit', 'bi_agg_limit', 'pd_limit', 'premium', 'renewal_date', 'outstanding_item', 'remarks'],
            'homeowners': ['carrier', 'dwelling_limit', 'liability_limit', 'premium', 'renewal_date', 'outstanding_item', 'remarks'],
            'personal_umbrella': ['carrier', 'liability_limit', 'deductible', 'premium', 'renewal_date', 'outstanding_item', 'remarks'],
            'event': ['carrier', 'type', 'location', 'start_date', 'end_date', 'entry_fee', 'audience_count', 'premium', 'outstanding_item', 'remarks'],
            'visitors_medical': ['carrier', 'start_date', 'end_date', 'destination_country', 'premium', 'outstanding_item', 'remarks'],
        }

        personal_headers = ['Individual ID', 'Individual Name']
        personal_sections = [(1, 2, '')]
        pcol = 3

        for prefix, label, cols in personal_product_defs:
            personal_headers.extend(cols)
            personal_sections.append((pcol, pcol + len(cols) - 1, label))
            pcol += len(cols)

        # Write section headers (Row 1)
        for start_col, end_col, title in personal_sections:
            if title:
                if start_col != end_col:
                    ws_personal.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws_personal.cell(row=1, column=start_col, value=title)
                cell.font = section_font
                cell.fill = section_fill

        # Write column headers (Row 2)
        for col, header in enumerate(personal_headers, 1):
            cell = ws_personal.cell(row=2, column=col, value=header)
            cell.font = header_font

        # Write personal data
        for row_idx, rec in enumerate(personal_records, 3):
            ind = rec.individual
            ind_name = f"{ind.first_name or ''} {ind.last_name or ''}".strip() if ind else None
            ws_personal.cell(row=row_idx, column=1, value=rec.individual_id)
            ws_personal.cell(row=row_idx, column=2, value=ind_name)

            pc = 3
            for prefix, label, cols in personal_product_defs:
                field_names = personal_field_map[prefix]
                for fi, field_suffix in enumerate(field_names):
                    val = getattr(rec, f'{prefix}_{field_suffix}', None)
                    if val is not None and hasattr(val, 'isoformat'):
                        pass  # date stays as-is for Excel
                    elif isinstance(val, (int, float)):
                        val = float(val) if val else 0
                    ws_personal.cell(row=row_idx, column=pc + fi, value=val)
                pc += len(field_names)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Client_Data_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/import', methods=['POST'])
def import_from_excel():
    """Import data from an Excel file matching the Data Sheet.xlsx format."""
    session = Session()
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400

        wb = load_workbook(file)

        stats = {
            'clients_created': 0,
            'individuals_created': 0,
            'benefits_created': 0,
            'commercial_created': 0,
            'personal_created': 0,
            'errors': []
        }

        # Error row collectors for errors.xlsx generation
        error_rows_clients = []      # [(row_tuple, error_msg), ...]
        error_rows_individuals = []
        error_rows_benefits = []
        error_rows_commercial = []
        error_rows_personal = []

        # ========== DELETE ALL EXISTING DATA ==========
        # Delete child records first to respect FK constraints
        session.query(HomeownersPolicy).delete()
        session.query(BenefitPlan).delete()
        session.query(CommercialPlan).delete()
        session.query(PersonalInsurance).delete()
        session.query(EmployeeBenefit).delete()
        session.query(CommercialInsurance).delete()
        session.query(ClientContact).delete()
        session.query(Individual).delete()
        session.query(Client).delete()
        session.flush()

        # ========== IMPORT CLIENTS ==========
        imported_clients = {}  # tax_id -> Client object (for handling duplicates in xlsx)
        if 'Clients' in wb.sheetnames:
            ws_clients = wb['Clients']
            for row_idx, row in enumerate(ws_clients.iter_rows(min_row=3, values_only=True), start=3):
                if not row[0]:  # Skip empty rows
                    continue
                try:
                    tax_id = str(row[0]).strip() if row[0] else None
                    if not tax_id:
                        continue

                    # New column order: Tax ID(0), Client Name(1), DBA(2), Industry(3), Status(4), Gross Revenue(5), Total EEs(6)
                    # Then contacts starting at col 7, each contact has 9 columns
                    contact_cols_per = 9
                    contact_start_col = 7

                    def parse_contacts_from_row(r):
                        contacts_list = []
                        ci = contact_start_col
                        sort = 0
                        while ci + contact_cols_per - 1 < len(r):
                            cp = r[ci] if r[ci] else None
                            em = r[ci+1] if len(r) > ci+1 and r[ci+1] else None
                            ph = str(r[ci+2]) if len(r) > ci+2 and r[ci+2] else None
                            ext = str(r[ci+3]) if len(r) > ci+3 and r[ci+3] else None
                            a1 = r[ci+4] if len(r) > ci+4 else None
                            a2 = r[ci+5] if len(r) > ci+5 else None
                            ct = r[ci+6] if len(r) > ci+6 else None
                            st = r[ci+7] if len(r) > ci+7 else None
                            zp = str(int(r[ci+8])).zfill(5) if len(r) > ci+8 and r[ci+8] else None
                            if cp or em or ph or a1:
                                contacts_list.append({'contact_person': cp, 'email': em, 'phone_number': ph, 'phone_extension': ext, 'address_line_1': a1, 'address_line_2': a2, 'city': ct, 'state': st, 'zip_code': zp, 'sort_order': sort})
                                sort += 1
                            ci += contact_cols_per
                        return contacts_list

                    parsed_contacts = parse_contacts_from_row(row)
                    fc = parsed_contacts[0] if parsed_contacts else {}

                    if tax_id in imported_clients:
                        # Duplicate tax_id in xlsx — update the existing record
                        existing = imported_clients[tax_id]
                        existing.client_name = row[1] if len(row) > 1 else existing.client_name
                        existing.dba = row[2] if len(row) > 2 else existing.dba
                        existing.industry = row[3] if len(row) > 3 else existing.industry
                        existing.status = row[4] if len(row) > 4 and row[4] else existing.status
                        existing.gross_revenue = float(row[5]) if len(row) > 5 and row[5] else existing.gross_revenue
                        existing.total_ees = int(row[6]) if len(row) > 6 and row[6] else existing.total_ees
                        if parsed_contacts:
                            existing.contact_person = fc.get('contact_person')
                            existing.email = fc.get('email')
                            existing.phone_number = fc.get('phone_number')
                            existing.address_line_1 = fc.get('address_line_1')
                            existing.address_line_2 = fc.get('address_line_2')
                            existing.city = fc.get('city')
                            existing.state = fc.get('state')
                            existing.zip_code = fc.get('zip_code')
                            # Add new contacts (append to existing)
                            for cd in parsed_contacts:
                                session.add(ClientContact(client_id=existing.id, **cd))
                        stats['clients_created'] += 1
                        continue

                    client = Client(
                        tax_id=tax_id,
                        client_name=row[1] if len(row) > 1 else None,
                        dba=row[2] if len(row) > 2 else None,
                        industry=row[3] if len(row) > 3 else None,
                        status=row[4] if len(row) > 4 and row[4] else 'Active',
                        gross_revenue=float(row[5]) if len(row) > 5 and row[5] else None,
                        total_ees=int(row[6]) if len(row) > 6 and row[6] else None,
                        contact_person=fc.get('contact_person'),
                        email=fc.get('email'),
                        phone_number=fc.get('phone_number'),
                        address_line_1=fc.get('address_line_1'),
                        address_line_2=fc.get('address_line_2'),
                        city=fc.get('city'),
                        state=fc.get('state'),
                        zip_code=fc.get('zip_code')
                    )
                    session.add(client)
                    session.flush()
                    imported_clients[tax_id] = client
                    for cd in parsed_contacts:
                        session.add(ClientContact(client_id=client.id, **cd))
                    stats['clients_created'] += 1
                except Exception as e:
                    error_rows_clients.append((row, str(e)))
                    stats['errors'].append(f"Clients row {row_idx}: {str(e)}")

        session.flush()  # Flush to ensure clients are available for FK references

        # ========== IMPORT INDIVIDUALS ==========
        if 'Individuals' in wb.sheetnames:
            ws_individuals = wb['Individuals']
            for row_idx, row in enumerate(ws_individuals.iter_rows(min_row=3, values_only=True), start=3):
                if not row[0]:
                    continue
                try:
                    individual_id = str(row[0]).strip() if row[0] else None
                    if not individual_id:
                        continue

                    zip_code = str(int(row[9])).zfill(5) if row[9] else None

                    individual = Individual(
                        individual_id=individual_id,
                        first_name=row[1] if len(row) > 1 else None,
                        last_name=row[2] if len(row) > 2 else None,
                        email=row[3] if len(row) > 3 else None,
                        phone_number=str(row[4]) if len(row) > 4 and row[4] else None,
                        address_line_1=row[5] if len(row) > 5 else None,
                        address_line_2=row[6] if len(row) > 6 else None,
                        city=row[7] if len(row) > 7 else None,
                        state=row[8] if len(row) > 8 else None,
                        zip_code=zip_code,
                        status=row[10] if len(row) > 10 and row[10] else 'Active'
                    )
                    session.add(individual)
                    stats['individuals_created'] += 1
                except Exception as e:
                    error_rows_individuals.append((row, str(e)))
                    stats['errors'].append(f"Individuals row {row_idx}: {str(e)}")

        session.flush()  # Flush to ensure individuals are available for FK references

        # ========== IMPORT EMPLOYEE BENEFITS ==========
        if 'Employee Benefits' in wb.sheetnames:
            ws_benefits = wb['Employee Benefits']

            # Read headers from row 2 to detect multi-plan columns dynamically
            headers = []
            for cell in ws_benefits[2]:
                headers.append(str(cell.value).strip() if cell.value else '')

            # Detect multi-plan column positions by header pattern
            multi_plan_header_map = {
                'medical': 'MEDICAL',
                'dental': 'DENTAL',
                'vision': 'VISION',
                'life_adnd': 'Life & AD&D'
            }

            # Find column indices for each multi-plan type (carrier/renewal/waiting_period/remarks/outstanding_item groups)
            multi_plan_cols = {}  # plan_type -> [(carrier_col, renewal_col, wp_col, remarks_col, outstanding_item_col), ...]
            for plan_type, label in multi_plan_header_map.items():
                cols = []
                for i, h in enumerate(headers):
                    if h and label.upper() in h.upper() and 'CARRIER' in h.upper():
                        renewal_col = i + 1 if i + 1 < len(headers) and 'RENEWAL' in headers[i + 1].upper() else None
                        wp_col = i + 2 if i + 2 < len(headers) and 'WAITING' in headers[i + 2].upper() else None
                        remarks_col = i + 3 if i + 3 < len(headers) and 'REMARKS' in headers[i + 3].upper() else None
                        oi_col = i + 4 if i + 4 < len(headers) and 'OUTSTANDING' in headers[i + 4].upper() else None
                        cols.append((i, renewal_col, wp_col, remarks_col, oi_col))
                multi_plan_cols[plan_type] = cols

            # Find single-plan type columns by header
            single_plan_col_map = {}  # prefix -> (renewal_col, carrier_col, remarks_col, outstanding_item_col)
            single_plan_labels = {
                'ltd': 'LTD', 'std': 'STD', 'k401': '401K',
                'critical_illness': 'Critical Illness', 'accident': 'Accident',
                'hospital': 'Hospital', 'voluntary_life': 'Voluntary Life'
            }
            for prefix, label in single_plan_labels.items():
                for i, h in enumerate(headers):
                    if h and label.upper() in h.upper() and 'RENEWAL' in h.upper():
                        is_multi = any(label.upper() == ml.upper() for ml in multi_plan_header_map.values())
                        if not is_multi:
                            carrier_col = i + 1 if i + 1 < len(headers) and 'CARRIER' in headers[i + 1].upper() else None
                            remarks_col = i + 2 if i + 2 < len(headers) and 'REMARKS' in headers[i + 2].upper() else None
                            oi_col = i + 3 if i + 3 < len(headers) and 'OUTSTANDING' in headers[i + 3].upper() else None
                            single_plan_col_map[prefix] = (i, carrier_col, remarks_col, oi_col)
                            break

            # Find fixed column indices
            def find_col(name):
                name_upper = name.upper()
                for i, h in enumerate(headers):
                    if h.upper().strip() == name_upper:
                        return i
                return None

            col_employer_contribution = None
            col_employee_contribution = None
            for i, h in enumerate(headers):
                if h and 'EMPLOYER CONTRIBUTION' in h.upper():
                    col_employer_contribution = i
                elif h and 'EMPLOYEE CONTRIBUTION' in h.upper():
                    col_employee_contribution = i

            for row_idx, row in enumerate(ws_benefits.iter_rows(min_row=3, values_only=True), start=3):
                if not row[0]:
                    continue
                try:
                    tax_id = str(row[0]).strip() if row[0] else None
                    if not tax_id:
                        continue

                    client = session.query(Client).filter_by(tax_id=tax_id).first()
                    if not client:
                        error_rows_benefits.append((row, f"Client with tax_id {tax_id} not found"))
                        stats['errors'].append(f"Benefits row {row_idx}: Client with tax_id {tax_id} not found")
                        continue

                    def parse_excel_date(val):
                        if val is None or val == '' or val == 'N/A':
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        try:
                            return parse(str(val)).date()
                        except:
                            return None

                    def safe_int(val):
                        if val is None or val == '':
                            return None
                        try:
                            return int(val)
                        except:
                            return None

                    def safe_val(idx):
                        return row[idx] if len(row) > idx and row[idx] else None

                    benefit_data = {
                        'tax_id': tax_id,
                        'parent_client': safe_val(2),
                        'form_fire_code': safe_val(3),
                        'enrollment_poc': safe_val(4),
                        'funding': safe_val(6),
                        'num_employees_at_renewal': safe_int(safe_val(7)),
                        'enrolled_ees': safe_int(safe_val(8)),
                        'waiting_period': safe_val(9),
                        'deductible_accumulation': safe_val(10),
                        'previous_carrier': safe_val(11),
                        'cobra_carrier': safe_val(12),
                        'employer_contribution': str(row[col_employer_contribution]) if col_employer_contribution is not None and len(row) > col_employer_contribution and row[col_employer_contribution] else None,
                        'employee_contribution': str(row[col_employee_contribution]) if col_employee_contribution is not None and len(row) > col_employee_contribution and row[col_employee_contribution] else None
                    }

                    # Single-plan types
                    for prefix, (renewal_col, carrier_col, remarks_col, oi_col) in single_plan_col_map.items():
                        if renewal_col is not None:
                            benefit_data[f'{prefix}_renewal_date'] = parse_excel_date(safe_val(renewal_col))
                        if carrier_col is not None:
                            benefit_data[f'{prefix}_carrier'] = safe_val(carrier_col)
                        if remarks_col is not None:
                            benefit_data[f'{prefix}_remarks'] = safe_val(remarks_col)
                        if oi_col is not None:
                            benefit_data[f'{prefix}_outstanding_item'] = safe_val(oi_col)

                    benefit_obj = EmployeeBenefit(**benefit_data)
                    session.add(benefit_obj)
                    session.flush()
                    stats['benefits_created'] += 1

                    # Multi-plan types: create BenefitPlan child records (deduplicate by carrier)
                    for plan_type, cols_list in multi_plan_cols.items():
                        seen_carriers = set()
                        actual_plan_num = 0
                        for _, (carrier_col, renewal_col, wp_col, remarks_col, oi_col) in enumerate(cols_list, 1):
                            carrier = safe_val(carrier_col)
                            renewal = parse_excel_date(safe_val(renewal_col)) if renewal_col is not None else None
                            wp_val = safe_val(wp_col) if wp_col is not None else None
                            remarks_val = safe_val(remarks_col) if remarks_col is not None else None
                            oi_val = safe_val(oi_col) if oi_col is not None else None
                            if carrier or renewal:
                                dedup_key = str(carrier or '').strip().lower()
                                if dedup_key in seen_carriers:
                                    continue
                                seen_carriers.add(dedup_key)
                                actual_plan_num += 1
                                plan = BenefitPlan(
                                    employee_benefit_id=benefit_obj.id,
                                    plan_type=plan_type,
                                    plan_number=actual_plan_num,
                                    carrier=carrier,
                                    renewal_date=renewal,
                                    waiting_period=wp_val,
                                    remarks=remarks_val,
                                    outstanding_item=oi_val
                                )
                                session.add(plan)
                                # Also set flat fields from first plan
                                if actual_plan_num == 1:
                                    if plan_type == 'medical':
                                        benefit_obj.current_carrier = carrier
                                        benefit_obj.renewal_date = renewal
                                    else:
                                        setattr(benefit_obj, f'{plan_type}_carrier', carrier)
                                        setattr(benefit_obj, f'{plan_type}_renewal_date', renewal)

                except Exception as e:
                    error_rows_benefits.append((row, str(e)))
                    stats['errors'].append(f"Benefits row {row_idx}: {str(e)}")

        # ========== IMPORT COMMERCIAL INSURANCE ==========
        if 'Commercial' in wb.sheetnames:
            ws_commercial = wb['Commercial']

            # Read headers from row 2 to detect column positions dynamically
            comm_headers = []
            for cell in ws_commercial[2]:
                comm_headers.append(str(cell.value).strip() if cell.value else '')

            # Read section headers from row 1 to identify product type boundaries
            section_headers = []
            for cell in ws_commercial[1]:
                section_headers.append(str(cell.value).strip() if cell.value else '')

            # Single-plan types: 8 cols each (Carrier, Agency, Occ Limit, Agg Limit, Premium, Renewal Date, Remarks, Outstanding Item)
            commercial_single_import_defs = [
                ('general_liability', 'Commercial General Liability'),
                ('property', 'Commercial Property'),
                ('bop', 'Business Owners Policy'),
                ('workers_comp', 'Workers Compensation'),
                ('auto', 'Commercial Auto'),
                ('epli', 'EPLI'),
                ('nydbl', 'NYDBL'),
                ('surety', 'Surety Bond'),
                ('product_liability', 'Product Liability'),
                ('flood', 'Flood'),
                ('directors_officers', 'Directors & Officers'),
                ('fiduciary', 'Fiduciary Bond'),
                ('inland_marine', 'Inland Marine')
            ]

            # Multi-plan types: dynamic cols (Carrier, Agency, Occ Limit, Agg Limit, Premium, Renewal Date, Remarks, Outstanding Item per plan)
            commercial_multi_import_defs = [
                ('umbrella', 'Umbrella Liability'),
                ('professional_eo', 'Professional or E&O'),
                ('cyber', 'Cyber Liability'),
                ('crime', 'Crime or Fidelity Bond')
            ]

            # Find section start columns from row 1
            section_col_map = {}
            for i, sh in enumerate(section_headers):
                if sh and sh != 'None':
                    section_col_map[sh] = i

            # Build column maps for single-plan types
            comm_single_col_map = {}  # prefix -> start_col (0-based)
            for prefix, label in commercial_single_import_defs:
                if label in section_col_map:
                    comm_single_col_map[prefix] = section_col_map[label]

            # Build column maps for multi-plan types — detect how many plans per type
            comm_multi_col_map = {}  # plan_type -> [(carrier_col, agency_col, policy_col, occ_limit_col, agg_limit_col, premium_col, renewal_col, remarks_col, oi_col, endorsement_cols), ...]
            for plan_type, label in commercial_multi_import_defs:
                if label in section_col_map:
                    start = section_col_map[label]
                    # Count columns belonging to this section (until next section or end)
                    plans = []
                    i = start
                    while i < len(comm_headers):
                        h = comm_headers[i].upper()
                        if 'CARRIER' in h:
                            j = i + 1
                            agency_col = j if j < len(comm_headers) and 'AGENCY' in comm_headers[j].upper() else None
                            if agency_col is not None:
                                j += 1
                            policy_col = j if j < len(comm_headers) and 'POLICY' in comm_headers[j].upper() else None
                            if policy_col is not None:
                                j += 1
                            occ_limit_col = j if j < len(comm_headers) and 'OCC' in comm_headers[j].upper() else None
                            if occ_limit_col is not None:
                                j += 1
                            agg_limit_col = j if j < len(comm_headers) and 'AGG' in comm_headers[j].upper() else None
                            if agg_limit_col is not None:
                                j += 1
                            premium_col = j if j < len(comm_headers) and 'PREMIUM' in comm_headers[j].upper() else None
                            if premium_col is not None:
                                j += 1
                            renewal_col = j if j < len(comm_headers) and 'RENEWAL' in comm_headers[j].upper() else None
                            if renewal_col is not None:
                                j += 1
                            remarks_col = j if j < len(comm_headers) and 'REMARKS' in comm_headers[j].upper() else None
                            if remarks_col is not None:
                                j += 1
                            oi_col = j if j < len(comm_headers) and 'OUTSTANDING' in comm_headers[j].upper() else None
                            if oi_col is not None:
                                j += 1
                            # Detect endorsement columns for professional_eo
                            endorsement_cols = {}
                            if plan_type == 'professional_eo':
                                for ek, ekey in [('TECH', 'endorsement_tech_eo'), ('STAFFING', 'endorsement_staffing'), ('ALLIED', 'endorsement_allied_healthcare'), ('MEDICAL MALPRACTICE', 'endorsement_medical_malpractice')]:
                                    if j < len(comm_headers) and 'ENDORSEMENT' in comm_headers[j].upper() and ek in comm_headers[j].upper():
                                        endorsement_cols[ekey] = j
                                        j += 1
                            plans.append((i, agency_col, policy_col, occ_limit_col, agg_limit_col, premium_col, renewal_col, remarks_col, oi_col, endorsement_cols))
                            i = j
                        else:
                            break
                    comm_multi_col_map[plan_type] = plans

            for row_idx, row in enumerate(ws_commercial.iter_rows(min_row=3, values_only=True), start=3):
                if not row[0]:
                    continue
                try:
                    tax_id = str(row[0]).strip() if row[0] else None
                    if not tax_id:
                        continue

                    client = session.query(Client).filter_by(tax_id=tax_id).first()
                    if not client:
                        error_rows_commercial.append((row, f"Client with tax_id {tax_id} not found"))
                        stats['errors'].append(f"Commercial row {row_idx}: Client with tax_id {tax_id} not found")
                        continue

                    def parse_excel_date(val):
                        if val is None or val == '' or val == 'N/A':
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        try:
                            return parse(str(val)).date()
                        except:
                            return None

                    def safe_decimal(val):
                        if val is None or val == '' or val == 'N/A':
                            return None
                        try:
                            return float(val)
                        except:
                            return None

                    def safe_val(idx):
                        return row[idx] if len(row) > idx and row[idx] else None

                    commercial_data = {
                        'tax_id': tax_id,
                        'parent_client': row[2] if len(row) > 2 else None,
                        'assigned_to': row[3] if len(row) > 3 else None
                    }

                    # Single-plan types (9 cols each: carrier, agency, policy_number, occ_limit, agg_limit, premium, renewal, remarks, outstanding_item + endorsements for GL)
                    for prefix, label in commercial_single_import_defs:
                        if prefix in comm_single_col_map:
                            sc = comm_single_col_map[prefix]
                            carrier = safe_val(sc)
                            if carrier and carrier == 'None':
                                carrier = None
                            commercial_data[f'{prefix}_carrier'] = carrier
                            commercial_data[f'{prefix}_agency'] = safe_val(sc + 1)
                            commercial_data[f'{prefix}_policy_number'] = safe_val(sc + 2)
                            occ_limit_val = safe_val(sc + 3)
                            if occ_limit_val and str(occ_limit_val) == 'N/A':
                                occ_limit_val = None
                            agg_limit_val = safe_val(sc + 4)
                            if agg_limit_val and str(agg_limit_val) == 'N/A':
                                agg_limit_val = None
                            # Property uses absolute dollar amounts (Building Limit, Personal Property), not millions
                            if prefix == 'property':
                                commercial_data[f'{prefix}_occ_limit'] = str(occ_limit_val) if occ_limit_val is not None else None
                                commercial_data[f'{prefix}_agg_limit'] = str(agg_limit_val) if agg_limit_val is not None else None
                            else:
                                commercial_data[f'{prefix}_occ_limit'] = format_limit(occ_limit_val)
                                commercial_data[f'{prefix}_agg_limit'] = format_limit(agg_limit_val)
                            commercial_data[f'{prefix}_premium'] = safe_decimal(safe_val(sc + 5))
                            commercial_data[f'{prefix}_renewal_date'] = parse_excel_date(safe_val(sc + 6))
                            commercial_data[f'{prefix}_remarks'] = safe_val(sc + 7)
                            commercial_data[f'{prefix}_outstanding_item'] = safe_val(sc + 8)
                            # GL endorsements (7 extra columns after the base 9)
                            if prefix == 'general_liability':
                                commercial_data['general_liability_endorsement_bop'] = str(safe_val(sc + 9)).strip().upper() == 'YES' if safe_val(sc + 9) else False
                                commercial_data['general_liability_endorsement_marine'] = str(safe_val(sc + 10)).strip().upper() == 'YES' if safe_val(sc + 10) else False
                                commercial_data['general_liability_endorsement_foreign'] = str(safe_val(sc + 11)).strip().upper() == 'YES' if safe_val(sc + 11) else False
                                commercial_data['general_liability_endorsement_molestation'] = str(safe_val(sc + 12)).strip().upper() == 'YES' if safe_val(sc + 12) else False
                                commercial_data['general_liability_endorsement_staffing'] = str(safe_val(sc + 13)).strip().upper() == 'YES' if safe_val(sc + 13) else False
                                commercial_data['general_liability_endorsement_accidental_medical'] = str(safe_val(sc + 14)).strip().upper() == 'YES' if safe_val(sc + 14) else False
                                commercial_data['general_liability_endorsement_liquor_liability'] = str(safe_val(sc + 15)).strip().upper() == 'YES' if safe_val(sc + 15) else False
                            # BOP property coverage (2 extra columns after the base 9)
                            elif prefix == 'bop':
                                commercial_data['bop_building_limit'] = safe_decimal(safe_val(sc + 9))
                                commercial_data['bop_personal_property'] = safe_decimal(safe_val(sc + 10))

                    comm_obj = CommercialInsurance(**commercial_data)
                    session.add(comm_obj)
                    session.flush()
                    stats['commercial_created'] += 1

                    # Multi-plan types: create CommercialPlan child records (deduplicate by carrier+policy_number)
                    for plan_type, cols_list in comm_multi_col_map.items():
                        seen_plans = set()  # track (carrier, policy_number) to skip duplicates
                        actual_plan_num = 0
                        for _, (carrier_col, agency_col, policy_col, occ_limit_col, agg_limit_col, premium_col, renewal_col, remarks_col, oi_col, endorsement_cols) in enumerate(cols_list, 1):
                            carrier = safe_val(carrier_col)
                            agency = safe_val(agency_col) if agency_col is not None else None
                            policy_number = safe_val(policy_col) if policy_col is not None else None
                            occ_limit_val = format_limit(safe_val(occ_limit_col)) if occ_limit_col is not None else None
                            agg_limit_val = format_limit(safe_val(agg_limit_col)) if agg_limit_col is not None else None
                            premium = safe_decimal(safe_val(premium_col)) if premium_col is not None else None
                            renewal = parse_excel_date(safe_val(renewal_col)) if renewal_col is not None else None
                            remarks_val = safe_val(remarks_col) if remarks_col is not None else None
                            oi_val = safe_val(oi_col) if oi_col is not None else None
                            if carrier or occ_limit_val or agg_limit_val or premium or renewal:
                                # Skip duplicate plans (same carrier and policy number)
                                dedup_key = (str(carrier or '').strip().lower(), str(policy_number or '').strip().lower())
                                if dedup_key in seen_plans:
                                    continue
                                seen_plans.add(dedup_key)
                                actual_plan_num += 1
                                plan = CommercialPlan(
                                    commercial_insurance_id=comm_obj.id,
                                    plan_type=plan_type,
                                    plan_number=actual_plan_num,
                                    carrier=carrier,
                                    agency=agency,
                                    policy_number=policy_number,
                                    coverage_occ_limit=occ_limit_val if occ_limit_val and str(occ_limit_val) != 'N/A' else None,
                                    coverage_agg_limit=agg_limit_val if agg_limit_val and str(agg_limit_val) != 'N/A' else None,
                                    premium=premium,
                                    renewal_date=renewal,
                                    remarks=remarks_val,
                                    outstanding_item=oi_val
                                )
                                # Professional E&O endorsements
                                if plan_type == 'professional_eo' and endorsement_cols:
                                    for ekey, ecol in endorsement_cols.items():
                                        setattr(plan, ekey, str(safe_val(ecol)).strip().upper() == 'YES' if safe_val(ecol) else False)
                                session.add(plan)
                                # Set flat fields from first plan for backward compat
                                if actual_plan_num == 1:
                                    setattr(comm_obj, f'{plan_type}_carrier', carrier)
                                    setattr(comm_obj, f'{plan_type}_agency', agency)
                                    setattr(comm_obj, f'{plan_type}_policy_number', policy_number)
                                    setattr(comm_obj, f'{plan_type}_occ_limit', occ_limit_val if occ_limit_val and occ_limit_val != 'N/A' else None)
                                    setattr(comm_obj, f'{plan_type}_agg_limit', agg_limit_val if agg_limit_val and agg_limit_val != 'N/A' else None)
                                    setattr(comm_obj, f'{plan_type}_premium', premium)
                                    setattr(comm_obj, f'{plan_type}_renewal_date', renewal)

                except Exception as e:
                    error_rows_commercial.append((row, str(e)))
                    stats['errors'].append(f"Commercial row {row_idx}: {str(e)}")

        # ========== IMPORT PERSONAL INSURANCE ==========
        if 'Personal' in wb.sheetnames:
            ws_personal = wb['Personal']

            # Read headers from row 2
            pers_headers = []
            for cell in ws_personal[2]:
                pers_headers.append(str(cell.value).strip() if cell.value else '')

            # Read section headers from row 1
            pers_section_headers = []
            for cell in ws_personal[1]:
                pers_section_headers.append(str(cell.value).strip() if cell.value else '')

            # Map section label -> start column (0-based)
            pers_section_col_map = {}
            for i, sh in enumerate(pers_section_headers):
                if sh and sh != 'None':
                    pers_section_col_map[sh] = i

            # Product defs: (prefix, section_label, [db_field_suffixes])
            pers_import_defs = [
                ('personal_auto', 'Personal Auto', ['carrier', 'bi_occ_limit', 'bi_agg_limit', 'pd_limit', 'premium', 'renewal_date', 'outstanding_item', 'remarks']),
                ('homeowners', 'Homeowners', ['carrier', 'dwelling_limit', 'liability_limit', 'premium', 'renewal_date', 'outstanding_item', 'remarks']),
                ('personal_umbrella', 'Personal Umbrella', ['carrier', 'liability_limit', 'deductible', 'premium', 'renewal_date', 'outstanding_item', 'remarks']),
                ('event', 'Event Insurance', ['carrier', 'type', 'location', 'start_date', 'end_date', 'entry_fee', 'audience_count', 'premium', 'outstanding_item', 'remarks']),
                ('visitors_medical', 'Visitors Medical', ['carrier', 'start_date', 'end_date', 'destination_country', 'premium', 'outstanding_item', 'remarks']),
            ]

            # Determine column start for each product from section headers
            pers_product_col_map = {}
            for prefix, label, fields in pers_import_defs:
                if label in pers_section_col_map:
                    pers_product_col_map[prefix] = (pers_section_col_map[label], fields)

            premium_fields = {'premium', 'deductible', 'entry_fee'}
            date_fields = {'renewal_date', 'start_date', 'end_date'}
            integer_fields = {'audience_count'}

            for row_idx, row in enumerate(ws_personal.iter_rows(min_row=3, values_only=True), start=3):
                if not row[0]:
                    continue
                try:
                    ind_id = str(row[0]).strip() if row[0] else None
                    if not ind_id:
                        continue

                    ind = session.query(Individual).filter_by(individual_id=ind_id).first()
                    if not ind:
                        error_rows_personal.append((row, f"Individual with id {ind_id} not found"))
                        stats['errors'].append(f"Personal row {row_idx}: Individual with id {ind_id} not found")
                        continue

                    def parse_excel_date_p(val):
                        if val is None or val == '' or val == 'N/A':
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        try:
                            return parse(str(val)).date()
                        except:
                            return None

                    def safe_decimal_p(val):
                        if val is None or val == '' or val == 'N/A':
                            return None
                        try:
                            return float(val)
                        except:
                            return None

                    def safe_val_p(idx):
                        return row[idx] if len(row) > idx and row[idx] else None

                    personal_data = {
                        'individual_id': ind_id,
                    }

                    for prefix, label, fields in pers_import_defs:
                        if prefix in pers_product_col_map:
                            start_col, field_list = pers_product_col_map[prefix]
                            for fi, field_suffix in enumerate(field_list):
                                val = safe_val_p(start_col + fi)
                                key = f'{prefix}_{field_suffix}'
                                if field_suffix in date_fields:
                                    personal_data[key] = parse_excel_date_p(val)
                                elif field_suffix in premium_fields:
                                    personal_data[key] = safe_decimal_p(val)
                                elif field_suffix in integer_fields:
                                    personal_data[key] = int(val) if val else None
                                else:
                                    personal_data[key] = val

                    pers_obj = PersonalInsurance(**personal_data)
                    session.add(pers_obj)
                    stats['personal_created'] += 1

                except Exception as e:
                    error_rows_personal.append((row, str(e)))
                    stats['errors'].append(f"Personal row {row_idx}: {str(e)}")

        session.commit()

        # ========== BUILD ERRORS WORKBOOK ==========
        response_data = {
            'message': 'Import completed successfully',
            'stats': stats
        }

        has_errors = error_rows_clients or error_rows_individuals or error_rows_benefits or error_rows_commercial or error_rows_personal
        if has_errors:
            error_wb = Workbook()
            error_wb.remove(error_wb.active)  # Remove default sheet

            # Helper: copy headers from source sheet to error sheet, append "Error" column
            def copy_headers_and_write_errors(source_sheet_name, error_rows):
                if not error_rows:
                    return
                src_ws = wb[source_sheet_name]
                err_ws = error_wb.create_sheet(source_sheet_name)

                # Find the max column used in row 2 (column headers)
                max_col = 0
                for cell in src_ws[2]:
                    if cell.value is not None:
                        max_col = cell.column

                # Copy row 1 (section headers) with merged cells
                for cell in src_ws[1]:
                    if cell.value is not None:
                        err_ws.cell(row=1, column=cell.column, value=cell.value)
                        err_ws.cell(row=1, column=cell.column).font = Font(bold=True, size=11)
                        err_ws.cell(row=1, column=cell.column).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

                # Copy merged cells from row 1
                for merged_range in src_ws.merged_cells.ranges:
                    if merged_range.min_row == 1 and merged_range.max_row == 1:
                        err_ws.merge_cells(start_row=1, start_column=merged_range.min_col,
                                           end_row=1, end_column=merged_range.max_col)

                # Copy row 2 (column headers)
                for cell in src_ws[2]:
                    if cell.value is not None:
                        err_ws.cell(row=2, column=cell.column, value=cell.value)
                        err_ws.cell(row=2, column=cell.column).font = Font(bold=True)

                # Append "Error" column header
                error_col = max_col + 1
                err_ws.cell(row=2, column=error_col, value='Error')
                err_ws.cell(row=2, column=error_col).font = Font(bold=True, color="FF0000")

                # Write errored rows
                for data_row_idx, (row_data, error_msg) in enumerate(error_rows, 3):
                    for col_idx, val in enumerate(row_data, 1):
                        # Convert date/datetime objects to string for safe writing
                        if isinstance(val, datetime):
                            val = val.strftime('%m/%d/%Y')
                        elif hasattr(val, 'strftime'):
                            val = val.strftime('%m/%d/%Y')
                        err_ws.cell(row=data_row_idx, column=col_idx, value=val)
                    err_ws.cell(row=data_row_idx, column=error_col, value=error_msg)

            # Build error sheets for each tab that has errors
            if 'Clients' in wb.sheetnames:
                copy_headers_and_write_errors('Clients', error_rows_clients)
            if 'Individuals' in wb.sheetnames:
                copy_headers_and_write_errors('Individuals', error_rows_individuals)
            if 'Employee Benefits' in wb.sheetnames:
                copy_headers_and_write_errors('Employee Benefits', error_rows_benefits)
            if 'Commercial' in wb.sheetnames:
                copy_headers_and_write_errors('Commercial', error_rows_commercial)
                if 'Personal' in wb.sheetnames:
                    copy_headers_and_write_errors('Personal', error_rows_personal)

            # Encode as base64
            error_output = io.BytesIO()
            error_wb.save(error_output)
            error_output.seek(0)
            errors_b64 = base64.b64encode(error_output.read()).decode('utf-8')

            response_data['errors_file'] = errors_b64
            response_data['errors_filename'] = f'Import_Errors_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return jsonify(response_data), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error importing from Excel: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# FEEDBACK ENDPOINTS
# ===========================================================================

@app.route('/api/feedback', methods=['GET'])
def get_feedback():
    """Get all feedback items."""
    try:
        items = Feedback.query.order_by(Feedback.created_at.desc()).all()
        return jsonify({'feedback': [item.to_dict() for item in items]}), 200
    except Exception as e:
        logging.error(f"Error fetching feedback: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/feedback', methods=['POST'])
def create_feedback():
    """Create a new feedback item."""
    try:
        data = request.get_json()
        item = Feedback(
            type=data.get('type', 'Bug'),
            subject=data.get('subject', ''),
            description=data.get('description', ''),
            status=data.get('status', 'New')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify(item.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating feedback: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/feedback/<int:feedback_id>', methods=['PUT'])
def update_feedback(feedback_id):
    """Update a feedback item."""
    try:
        item = Feedback.query.get(feedback_id)
        if not item:
            return jsonify({'error': 'Feedback not found'}), 404

        data = request.get_json()
        if 'type' in data:
            item.type = data['type']
        if 'subject' in data:
            item.subject = data['subject']
        if 'description' in data:
            item.description = data['description']
        if 'status' in data:
            item.status = data['status']
        item.updated_at = datetime.utcnow()

        db.session.commit()
        return jsonify(item.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating feedback: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/feedback/<int:feedback_id>', methods=['DELETE'])
def delete_feedback(feedback_id):
    """Delete a feedback item."""
    try:
        item = Feedback.query.get(feedback_id)
        if not item:
            return jsonify({'error': 'Feedback not found'}), 404

        db.session.delete(item)
        db.session.commit()
        return jsonify({'message': 'Feedback deleted'}), 200
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error deleting feedback: {e}")
        return jsonify({'error': str(e)}), 500


# ===========================================================================
# HEALTH CHECK
# ===========================================================================

@app.route('/api/health')
def health_check():
    """Quick health check — verifies DB connectivity."""
    try:
        with app.app_context():
            db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'ok',
            'database': db_uri.split('?')[0],  # show path without query params
            'static_folder': app.static_folder or 'not configured'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'database': db_uri.split('?')[0],
            'error': str(e)
        }), 500


# ===========================================================================
# POC MANAGEMENT
# ===========================================================================

@app.route('/api/benefits/poc-summary', methods=['GET'])
def get_poc_summary():
    """Get unique enrollment PoCs with their record counts."""
    session = Session()
    try:
        results = (
            session.query(
                EmployeeBenefit.enrollment_poc,
                func.count(EmployeeBenefit.id).label('count')
            )
            .filter(
                EmployeeBenefit.enrollment_poc.isnot(None),
                EmployeeBenefit.enrollment_poc != ''
            )
            .group_by(EmployeeBenefit.enrollment_poc)
            .order_by(EmployeeBenefit.enrollment_poc)
            .all()
        )

        unassigned_count = (
            session.query(func.count(EmployeeBenefit.id))
            .filter(
                or_(
                    EmployeeBenefit.enrollment_poc.is_(None),
                    EmployeeBenefit.enrollment_poc == ''
                )
            )
            .scalar()
        )

        poc_list = [{'poc': r[0], 'count': r[1]} for r in results]

        return jsonify({
            'pocs': poc_list,
            'unassigned_count': unassigned_count
        }), 200
    except Exception as e:
        logging.error(f"Error fetching PoC summary: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@app.route('/api/benefits/poc-reassign', methods=['PUT'])
def reassign_poc():
    """Bulk reassign enrollment PoC. Supports full or partial reassignment via record_ids."""
    session = Session()
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        to_poc = data.get('to_poc')
        record_ids = data.get('record_ids')  # optional list of IDs for partial reassign
        from_poc = data.get('from_poc')

        if not to_poc:
            return jsonify({'error': 'to_poc is required'}), 400

        if record_ids:
            # Partial reassignment by specific record IDs
            query = session.query(EmployeeBenefit).filter(
                EmployeeBenefit.id.in_(record_ids)
            )
        elif from_poc:
            # Full reassignment by source PoC name
            if from_poc == to_poc:
                return jsonify({'error': 'Source and target PoC cannot be the same'}), 400
            query = session.query(EmployeeBenefit).filter(
                EmployeeBenefit.enrollment_poc == from_poc
            )
        else:
            return jsonify({'error': 'Either from_poc or record_ids is required'}), 400

        updated_count = query.update({EmployeeBenefit.enrollment_poc: to_poc})
        session.commit()

        return jsonify({
            'message': f'Successfully reassigned {updated_count} record(s) to "{to_poc}"',
            'updated_count': updated_count
        }), 200
    except Exception as e:
        session.rollback()
        logging.error(f"Error reassigning PoC: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# CHAT ENDPOINT
# ===========================================================================

@app.route('/api/chat', methods=['POST'])
def chat_endpoint():
    """Handle AI chat messages with tool-calling support."""
    session = Session()
    try:
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({'error': 'message is required'}), 400

        message = data['message']
        history = data.get('history', [])

        models = {
            'Client': Client,
            'EmployeeBenefit': EmployeeBenefit,
            'CommercialInsurance': CommercialInsurance,
            'PersonalInsurance': PersonalInsurance,
            'Individual': Individual,
        }

        result = chat_with_ollama(message, history, session, models)
        return jsonify(result), 200

    except Exception as e:
        logging.error(f"Chat endpoint error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


# ===========================================================================
# SERVE REACT APP (production mode)
# ===========================================================================

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    """Serve React build files. API routes take priority (registered first)."""
    if app.static_folder and os.path.isdir(app.static_folder):
        file_path = os.path.join(app.static_folder, path)
        if path and os.path.isfile(file_path):
            return send_file(file_path)
        index = os.path.join(app.static_folder, 'index.html')
        if os.path.isfile(index):
            return send_file(index)
    return jsonify({'error': 'Not found'}), 404


# ===========================================================================
# DATABASE INITIALIZATION
# ===========================================================================

# Create tables (runs on import so schema applies regardless of entry point)
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    # Run app (host/port configurable via env vars)
    host = os.environ.get('API_HOST', '127.0.0.1')
    port = int(os.environ.get('API_PORT', '5001'))
    debug = os.environ.get('API_DEBUG', 'true').lower() == 'true'
    app.run(debug=debug, host=host, port=port)
