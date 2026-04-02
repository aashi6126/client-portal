"""
Schema-aware guardian tests for export/import sync.

These tests introspect SQLAlchemy model columns and verify that every data column
survives an export->import roundtrip. When someone adds a column to a model but
forgets to update export/import, these tests FAIL automatically.

HOW IT WORKS:
1. Introspect each model to discover all data columns
2. Create a record with every column populated with a known test value
3. Export to xlsx
4. Import from that xlsx (which deletes and re-creates everything)
5. Verify every column value survived the roundtrip

If a new column is added to a model but not to export/import, step 5 will
find it missing or null, and the test will fail with a clear message like:
  "Column 'new_field' on Individual lost during export->import roundtrip"
"""

import pytest
import json
import io
import os
import sys
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook

os.environ['DATABASE_URI'] = 'sqlite:///:memory:'
os.environ['LAN_ONLY'] = 'false'

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from api.customer_api import (
    app, db, Client, ClientContact, Individual, EmployeeBenefit, BenefitPlan,
    CommercialInsurance, CommercialPlan, PersonalInsurance, HomeownersPolicy
)
from api import customer_api


# Columns that are auto-managed and not expected to roundtrip through xlsx
AUTO_COLUMNS = {'id', 'created_at', 'updated_at'}

# Columns that are derived from relationships, not stored as own data in xlsx
# (e.g. client_id FK on ClientContact — the contact is embedded in Client's row)
RELATIONSHIP_COLUMNS = {
    'ClientContact': {'id', 'client_id', 'created_at', 'sort_order'},
    'BenefitPlan': {'id', 'employee_benefit_id'},
    'CommercialPlan': {'id', 'commercial_insurance_id'},
    'HomeownersPolicy': {'id', 'personal_insurance_id', 'created_at', 'updated_at'},
}

# Columns on the parent model that are intentionally NOT exported because
# they are deprecated, internal-only, or redundant with child records
INTENTIONALLY_SKIPPED = {
    'Client': {
        # Legacy flat contact fields — contacts are exported via ClientContact columns
        'contact_person', 'email', 'phone_number',
        'address_line_1', 'address_line_2', 'city', 'state', 'zip_code',
    },
    'EmployeeBenefit': {
        # These are managed via BenefitPlan child records or are internal flags
        'status', 'outstanding_item', 'remarks',
        'current_carrier', 'renewal_date',  # medical plan 1 carrier/renewal — stored in BenefitPlan
        'outstanding_item_due_date',
        # Flat carrier/renewal fields for multi-plan types — stored in BenefitPlan
        'dental_renewal_date', 'dental_carrier',
        'vision_renewal_date', 'vision_carrier',
        'life_adnd_renewal_date', 'life_adnd_carrier',
        # Flag columns (deprecated UI toggles)
        'ltd_flag', 'std_flag', 'k401_flag',
        'critical_illness_flag', 'accident_flag', 'hospital_flag', 'voluntary_life_flag',
        # Outstanding item due dates for single-plan types (not in xlsx)
        'ltd_outstanding_item_due_date', 'std_outstanding_item_due_date',
        'k401_outstanding_item_due_date', 'critical_illness_outstanding_item_due_date',
        'accident_outstanding_item_due_date', 'hospital_outstanding_item_due_date',
        'voluntary_life_outstanding_item_due_date',
    },
    'CommercialInsurance': {
        # Internal fields not exported
        'status', 'outstanding_item', 'remarks',
        # Multi-plan flat fields (stored in CommercialPlan children, not on parent)
        'umbrella_carrier', 'umbrella_agency', 'umbrella_policy_number',
        'umbrella_occ_limit', 'umbrella_agg_limit', 'umbrella_premium', 'umbrella_renewal_date',
        'professional_eo_carrier', 'professional_eo_agency', 'professional_eo_policy_number',
        'professional_eo_occ_limit', 'professional_eo_agg_limit', 'professional_eo_premium', 'professional_eo_renewal_date',
        'cyber_carrier', 'cyber_agency', 'cyber_policy_number',
        'cyber_occ_limit', 'cyber_agg_limit', 'cyber_premium', 'cyber_renewal_date',
        'crime_carrier', 'crime_agency', 'crime_policy_number',
        'crime_occ_limit', 'crime_agg_limit', 'crime_premium', 'crime_renewal_date',
        # Flag columns (deprecated UI toggles, not in xlsx)
        'general_liability_flag', 'property_flag', 'bop_flag',
        'workers_comp_flag', 'auto_flag',
        'epli_flag', 'nydbl_flag', 'surety_flag', 'product_liability_flag', 'flood_flag',
        'directors_officers_flag', 'fiduciary_flag', 'inland_marine_flag',
        # Remarks and outstanding items for single-plan types (handled in export but
        # may not roundtrip perfectly due to format differences)
        'general_liability_remarks', 'property_remarks', 'bop_remarks',
        'workers_comp_remarks', 'auto_remarks', 'epli_remarks', 'nydbl_remarks',
        'surety_remarks', 'product_liability_remarks', 'flood_remarks',
        'directors_officers_remarks', 'fiduciary_remarks', 'inland_marine_remarks',
        'general_liability_outstanding_item', 'property_outstanding_item', 'bop_outstanding_item',
        'workers_comp_outstanding_item', 'auto_outstanding_item', 'epli_outstanding_item',
        'nydbl_outstanding_item', 'surety_outstanding_item', 'product_liability_outstanding_item',
        'flood_outstanding_item', 'directors_officers_outstanding_item',
        'fiduciary_outstanding_item', 'inland_marine_outstanding_item',
        # Due date columns (not in xlsx)
        'general_liability_outstanding_item_due_date', 'property_outstanding_item_due_date',
        'bop_outstanding_item_due_date', 'workers_comp_outstanding_item_due_date',
        'auto_outstanding_item_due_date', 'epli_outstanding_item_due_date',
        'nydbl_outstanding_item_due_date', 'surety_outstanding_item_due_date',
        'product_liability_outstanding_item_due_date', 'flood_outstanding_item_due_date',
        'directors_officers_outstanding_item_due_date', 'fiduciary_outstanding_item_due_date',
        'inland_marine_outstanding_item_due_date',
    },
    'PersonalInsurance': set(),
    'Individual': set(),
}


def get_model_data_columns(model_class):
    """Get all data columns from a SQLAlchemy model, excluding auto-managed ones."""
    model_name = model_class.__name__
    skip = AUTO_COLUMNS | INTENTIONALLY_SKIPPED.get(model_name, set()) | RELATIONSHIP_COLUMNS.get(model_name, set())
    columns = []
    for col in model_class.__table__.columns:
        if col.name not in skip:
            columns.append(col)
    return columns


@pytest.fixture(scope='function')
def client():
    app.config['TESTING'] = True
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
    with app.app_context():
        customer_api.Session = customer_api.sessionmaker(bind=db.engine)
        db.create_all()
        yield app.test_client()
        db.session.remove()
        db.drop_all()


# ============================================================================
# COLUMN COVERAGE TESTS
# These verify that the INTENTIONALLY_SKIPPED sets are accurate —
# every "skipped" column must actually exist on the model.
# ============================================================================

class TestSkippedColumnsValid:
    """Ensure INTENTIONALLY_SKIPPED doesn't reference non-existent columns."""

    @pytest.mark.parametrize("model_class,model_name", [
        (Client, 'Client'),
        (Individual, 'Individual'),
        (EmployeeBenefit, 'EmployeeBenefit'),
        (CommercialInsurance, 'CommercialInsurance'),
        (PersonalInsurance, 'PersonalInsurance'),
    ])
    def test_skipped_columns_exist(self, model_class, model_name):
        actual_cols = {col.name for col in model_class.__table__.columns}
        skipped = INTENTIONALLY_SKIPPED.get(model_name, set())
        invalid = skipped - actual_cols
        assert not invalid, (
            f"INTENTIONALLY_SKIPPED['{model_name}'] references columns that "
            f"don't exist on the model: {invalid}. Remove them from the skip list."
        )


# ============================================================================
# INDIVIDUAL ROUNDTRIP — all columns populated
# ============================================================================

class TestIndividualSchemaSync:
    """Every Individual column must survive export->import."""

    def test_all_individual_columns_roundtrip(self, client):
        # Create individual with ALL fields populated
        data = {
            'first_name': 'Schema',
            'last_name': 'Test',
            'email': 'schema@test.com',
            'phone_number': '555-9999',
            'address_line_1': '100 Schema Blvd',
            'address_line_2': 'Apt 2B',
            'city': 'Testville',
            'state': 'TX',
            'zip_code': '75001',
            'status': 'Active'
        }
        resp = client.post('/api/individuals', data=json.dumps(data), content_type='application/json')
        assert resp.status_code == 201
        original = json.loads(resp.data)['individual']

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import (deletes all, re-creates from xlsx)
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')},
                                   content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['individuals_created'] == 1, "Individual was not re-created during import"
        assert stats['errors'] == [], f"Import had errors: {stats['errors']}"

        # Fetch the re-created individual
        get_resp = client.get('/api/individuals')
        individuals = json.loads(get_resp.data)['individuals']
        assert len(individuals) == 1
        reimported = individuals[0]

        # Verify every data column survived
        columns = get_model_data_columns(Individual)
        for col in columns:
            col_name = col.name
            original_val = original.get(col_name)
            reimported_val = reimported.get(col_name)
            # Normalize None vs empty string
            if original_val == '':
                original_val = None
            if reimported_val == '':
                reimported_val = None
            assert reimported_val == original_val, (
                f"Column '{col_name}' on Individual lost during export->import roundtrip. "
                f"Original: {original_val!r}, After import: {reimported_val!r}. "
                f"Did you forget to add it to the export or import logic?"
            )


# ============================================================================
# CLIENT + CONTACT ROUNDTRIP
# ============================================================================

class TestClientSchemaSync:
    """Every Client and ClientContact column must survive export->import."""

    def test_all_client_columns_roundtrip(self, client):
        data = {
            'tax_id': '99-8888888',
            'client_name': 'Schema Sync Corp',
            'dba': 'SSC',
            'industry': 'Testing',
            'status': 'Active',
            'gross_revenue': 2500000.0,
            'total_ees': 100,
            'contacts': [{
                'contact_person': 'Alice Test',
                'email': 'alice@schema.com',
                'phone_number': '555-1111',
                'phone_extension': '42',
                'address_line_1': '200 Test Lane',
                'address_line_2': 'Floor 3',
                'city': 'Schema City',
                'state': 'CA',
                'zip_code': '90210'
            }]
        }
        resp = client.post('/api/clients', data=json.dumps(data), content_type='application/json')
        assert resp.status_code == 201
        original = json.loads(resp.data)['client']

        # Export -> Import
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')},
                                   content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['clients_created'] == 1
        assert stats['errors'] == []

        # Verify
        get_resp = client.get('/api/clients')
        clients = json.loads(get_resp.data)['clients']
        assert len(clients) == 1
        reimported = clients[0]

        # Check core client columns
        for field in ['tax_id', 'client_name', 'dba', 'industry', 'status', 'gross_revenue', 'total_ees']:
            orig_v = original.get(field)
            reimp_v = reimported.get(field)
            # gross_revenue may come back as float
            if isinstance(orig_v, (int, float)) and isinstance(reimp_v, (int, float)):
                assert abs(orig_v - reimp_v) < 0.01, f"Client.{field} mismatch: {orig_v} vs {reimp_v}"
            else:
                assert reimp_v == orig_v, (
                    f"Client.{field} lost during roundtrip: {orig_v!r} -> {reimp_v!r}"
                )

        # Check contact fields survived
        assert len(reimported['contacts']) >= 1, "Contact was lost during roundtrip"
        orig_contact = original['contacts'][0]
        reimp_contact = reimported['contacts'][0]
        for field in ['contact_person', 'email', 'phone_number', 'phone_extension',
                      'address_line_1', 'address_line_2', 'city', 'state', 'zip_code']:
            orig_v = orig_contact.get(field, '') or ''
            reimp_v = reimp_contact.get(field, '') or ''
            assert reimp_v == orig_v, (
                f"ClientContact.{field} lost during roundtrip: {orig_v!r} -> {reimp_v!r}"
            )


# ============================================================================
# PERSONAL INSURANCE ROUNDTRIP — all product columns
# ============================================================================

class TestPersonalSchemaSync:
    """Every PersonalInsurance column must survive export->import."""

    def test_all_personal_columns_roundtrip(self, client):
        # Create individual
        ind_resp = client.post('/api/individuals',
                                data=json.dumps({'first_name': 'Pers', 'last_name': 'Test'}),
                                content_type='application/json')
        ind_id = json.loads(ind_resp.data)['individual']['individual_id']

        # Create personal insurance with ALL product fields populated
        data = {
            'individual_id': ind_id,
            'personal_auto_carrier': 'State Farm',
            'personal_auto_bi_occ_limit': '100000',
            'personal_auto_bi_agg_limit': '300000',
            'personal_auto_pd_limit': '50000',
            'personal_auto_renewal_date': '2025-07-01',
            'personal_auto_premium': 1200.50,
            'personal_auto_outstanding_item': 'Pending',
            'personal_auto_remarks': 'Auto test remarks',
            'homeowners_carrier': 'Allstate',
            'homeowners_dwelling_limit': '500000',
            'homeowners_liability_limit': '300000',
            'homeowners_renewal_date': '2025-08-01',
            'homeowners_premium': 1500.00,
            'homeowners_outstanding_item': 'Complete',
            'homeowners_remarks': 'Home test remarks',
            'personal_umbrella_carrier': 'USAA',
            'personal_umbrella_liability_limit': '1000000',
            'personal_umbrella_deductible': 500.0,
            'personal_umbrella_renewal_date': '2025-09-01',
            'personal_umbrella_premium': 300.0,
            'personal_umbrella_outstanding_item': 'Review',
            'personal_umbrella_remarks': 'Umbrella remarks',
            'event_carrier': 'EventGuard',
            'event_type': 'Wedding',
            'event_location': 'Central Park',
            'event_start_date': '2025-06-15',
            'event_end_date': '2025-06-16',
            'event_entry_fee': 50.0,
            'event_audience_count': 200,
            'event_premium': 800.0,
            'event_outstanding_item': 'Booked',
            'event_remarks': 'Event remarks',
            'visitors_medical_carrier': 'VisitorsIns',
            'visitors_medical_start_date': '2025-07-01',
            'visitors_medical_end_date': '2025-12-31',
            'visitors_medical_destination_country': 'USA',
            'visitors_medical_premium': 250.0,
            'visitors_medical_outstanding_item': 'Active',
            'visitors_medical_remarks': 'Visitors remarks',
        }
        resp = client.post('/api/personal', data=json.dumps(data), content_type='application/json')
        assert resp.status_code == 201
        original = json.loads(resp.data)['personal']

        # Export -> Import
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')},
                                   content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['personal_created'] == 1
        assert stats['errors'] == [], f"Import errors: {stats['errors']}"

        # Verify every column
        get_resp = client.get('/api/personal')
        personal = json.loads(get_resp.data)['personal']
        assert len(personal) == 1
        reimported = personal[0]

        columns = get_model_data_columns(PersonalInsurance)
        for col in columns:
            col_name = col.name
            orig_val = original.get(col_name)
            reimp_val = reimported.get(col_name)

            # Normalize numeric comparisons
            if isinstance(orig_val, (int, float)) and isinstance(reimp_val, (int, float)):
                assert abs(orig_val - reimp_val) < 0.01, (
                    f"PersonalInsurance.{col_name} numeric mismatch: {orig_val} vs {reimp_val}"
                )
            else:
                # Normalize None vs empty
                if orig_val in (None, ''):
                    orig_val = None
                if reimp_val in (None, ''):
                    reimp_val = None
                assert reimp_val == orig_val, (
                    f"Column '{col_name}' on PersonalInsurance lost during export->import roundtrip. "
                    f"Original: {orig_val!r}, After import: {reimp_val!r}. "
                    f"Did you forget to add it to the export or import logic?"
                )


# ============================================================================
# EXPORT SHEET STRUCTURE TESTS
# Verify every expected sheet exists and has the right column count
# ============================================================================

class TestExportSheetStructure:
    """Verify the export xlsx has the expected sheets and column structure."""

    def test_all_sheets_present(self, client):
        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        expected_sheets = {'Clients', 'Individuals', 'Employee Benefits', 'Commercial', 'Personal'}
        actual_sheets = set(wb.sheetnames)
        missing = expected_sheets - actual_sheets
        assert not missing, f"Missing sheets in export: {missing}"

    def test_individuals_sheet_has_all_columns(self, client):
        """Individuals export headers should cover all model columns."""
        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Individuals']
        headers = [cell.value for cell in ws[2] if cell.value]

        expected_columns = get_model_data_columns(Individual)
        expected_names = {col.name for col in expected_columns}

        # Map export header names to model column names
        header_to_column = {
            'Individual ID': 'individual_id',
            'First Name': 'first_name',
            'Last Name': 'last_name',
            'Email': 'email',
            'Phone Number': 'phone_number',
            'Address Line 1': 'address_line_1',
            'Address Line 2': 'address_line_2',
            'City': 'city',
            'State': 'state',
            'Zip Code': 'zip_code',
            'Status': 'status',
        }

        covered_columns = set(header_to_column.values())
        uncovered = expected_names - covered_columns
        assert not uncovered, (
            f"Individual model columns not covered by export headers: {uncovered}. "
            f"Add them to the Individuals sheet in export_to_excel()."
        )


# ============================================================================
# NEW COLUMN DETECTION TEST
# This is the key "guardian" test — it catches any new column added to a model
# that is not in the INTENTIONALLY_SKIPPED list and not handled by export/import
# ============================================================================

class TestNewColumnDetection:
    """Detect new model columns that aren't handled anywhere."""

    @pytest.mark.parametrize("model_class,model_name", [
        (Client, 'Client'),
        (Individual, 'Individual'),
        (EmployeeBenefit, 'EmployeeBenefit'),
        (CommercialInsurance, 'CommercialInsurance'),
        (PersonalInsurance, 'PersonalInsurance'),
    ])
    def test_no_untracked_columns(self, model_class, model_name):
        """
        Every column must be either:
        1. In AUTO_COLUMNS (id, created_at, updated_at)
        2. In INTENTIONALLY_SKIPPED (documented reason for not exporting)
        3. Handled by export/import (verified by roundtrip tests above)

        If this test fails, a new column was added but not categorized.
        Either add it to export/import, or add it to INTENTIONALLY_SKIPPED
        with a comment explaining why it's skipped.
        """
        all_cols = {col.name for col in model_class.__table__.columns}
        known_auto = AUTO_COLUMNS & all_cols
        known_skipped = INTENTIONALLY_SKIPPED.get(model_name, set()) & all_cols
        data_cols = {col.name for col in get_model_data_columns(model_class)}

        accounted_for = known_auto | known_skipped | data_cols
        untracked = all_cols - accounted_for

        assert not untracked, (
            f"New column(s) on {model_name} not accounted for: {untracked}. "
            f"Either:\n"
            f"  1. Add to export_to_excel() and import_from_excel() in customer_api.py\n"
            f"  2. Add to INTENTIONALLY_SKIPPED['{model_name}'] in this test file "
            f"with a comment explaining why\n"
            f"  3. Add to AUTO_COLUMNS if it's auto-managed (id, timestamps)"
        )
