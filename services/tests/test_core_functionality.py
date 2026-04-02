"""
Comprehensive tests for core API functionality: CRUD operations for all entities,
import/export roundtrip, and data integrity validation.
"""

import pytest
import json
import io
import os
import sys
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

# Ensure test DB is set before importing app
os.environ['DATABASE_URI'] = 'sqlite:///:memory:'
os.environ['LAN_ONLY'] = 'false'

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from api.customer_api import (
    app, db, Client, ClientContact, Individual, EmployeeBenefit, BenefitPlan,
    CommercialInsurance, CommercialPlan, PersonalInsurance, HomeownersPolicy, Feedback
)
from api import customer_api


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture(scope='function')
def client():
    """Create a test client with isolated in-memory database."""
    app.config['TESTING'] = True
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'

    with app.app_context():
        customer_api.Session = customer_api.sessionmaker(bind=db.engine)
        db.create_all()
        yield app.test_client()
        db.session.remove()
        db.drop_all()


@pytest.fixture
def sample_client_data():
    """Sample client data for testing."""
    return {
        'tax_id': '12-3456789',
        'client_name': 'Test Company LLC',
        'dba': 'Test Co',
        'industry': 'Technology',
        'status': 'Active',
        'gross_revenue': 1000000.0,
        'total_ees': 50,
        'contacts': [{
            'contact_person': 'Jane Smith',
            'email': 'jane@testco.com',
            'phone_number': '555-0100',
            'phone_extension': '101',
            'address_line_1': '123 Main St',
            'address_line_2': 'Suite 100',
            'city': 'New York',
            'state': 'NY',
            'zip_code': '10001'
        }]
    }


@pytest.fixture
def sample_individual_data():
    """Sample individual data for testing."""
    return {
        'first_name': 'John',
        'last_name': 'Doe',
        'email': 'john.doe@email.com',
        'phone_number': '555-0200',
        'address_line_1': '456 Oak Ave',
        'city': 'Chicago',
        'state': 'IL',
        'zip_code': '60601',
        'status': 'Active'
    }


@pytest.fixture
def sample_benefit_data():
    """Sample employee benefit data for testing."""
    return {
        'tax_id': '12-3456789',
        'parent_client': 'Parent Corp',
        'form_fire_code': 'FF100',
        'enrollment_poc': 'John Doe',
        'funding': 'Fully Insured',
        'num_employees_at_renewal': 50,
        'enrolled_ees': 45,
        'waiting_period': '30 days',
        'deductible_accumulation': 'Calendar Year',
        'previous_carrier': 'Aetna',
        'cobra_carrier': 'WageWorks',
        'employer_contribution': '80%',
        'employee_contribution': '20%',
        'plans': {
            'medical': [{'carrier': 'BlueCross', 'renewal_date': '2025-06-01'}],
            'dental': [{'carrier': 'Delta Dental', 'renewal_date': '2025-06-01'}],
        }
    }


@pytest.fixture
def sample_commercial_data():
    """Sample commercial insurance data for testing."""
    return {
        'tax_id': '12-3456789',
        'parent_client': 'Parent Corp',
        'assigned_to': 'Agent Smith',
        'general_liability_carrier': 'Hartford',
        'general_liability_premium': 5000.0,
        'general_liability_renewal_date': '2025-06-01',
        'general_liability_occ_limit': '1',
        'general_liability_agg_limit': '2',
        'property_carrier': 'Zurich',
        'property_premium': 3000.0,
    }


@pytest.fixture
def sample_personal_data():
    """Sample personal insurance data for testing."""
    return {
        'individual_id': None,  # Set dynamically
        'personal_auto_carrier': 'State Farm',
        'personal_auto_premium': 1200.0,
        'personal_auto_renewal_date': '2025-07-01',
        'personal_auto_bi_occ_limit': '100000',
        'personal_auto_bi_agg_limit': '300000',
        'personal_auto_pd_limit': '50000',
        'homeowners_carrier': 'Allstate',
        'homeowners_premium': 1500.0,
        'homeowners_renewal_date': '2025-08-01',
    }


def create_client(client, data=None):
    """Helper to create a client and return the response data."""
    if data is None:
        data = {
            'tax_id': '12-3456789',
            'client_name': 'Test Company',
            'status': 'Active',
            'contacts': [{'contact_person': 'Jane', 'email': 'jane@test.com'}]
        }
    resp = client.post('/api/clients', data=json.dumps(data), content_type='application/json')
    return json.loads(resp.data)


def create_individual(client, data=None):
    """Helper to create an individual and return the response data."""
    if data is None:
        data = {'first_name': 'John', 'last_name': 'Doe', 'email': 'john@test.com', 'status': 'Active'}
    resp = client.post('/api/individuals', data=json.dumps(data), content_type='application/json')
    return json.loads(resp.data)


# ============================================================================
# CLIENT CRUD TESTS
# ============================================================================

class TestClientCRUD:
    """Tests for Client CRUD operations."""

    def test_create_client(self, client, sample_client_data):
        resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        assert resp.status_code == 201
        data = json.loads(resp.data)
        assert data['client']['tax_id'] == '12-3456789'
        assert data['client']['client_name'] == 'Test Company LLC'

    def test_get_clients_empty(self, client):
        resp = client.get('/api/clients')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['clients'] == []

    def test_get_clients_with_data(self, client, sample_client_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        resp = client.get('/api/clients')
        data = json.loads(resp.data)
        assert len(data['clients']) == 1
        assert data['clients'][0]['client_name'] == 'Test Company LLC'

    def test_update_client(self, client, sample_client_data):
        create_resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client_id = json.loads(create_resp.data)['client']['id']

        update = {'client_name': 'Updated Company', 'status': 'Inactive'}
        resp = client.put(f'/api/clients/{client_id}', data=json.dumps(update), content_type='application/json')
        assert resp.status_code == 200

        get_resp = client.get('/api/clients')
        clients = json.loads(get_resp.data)['clients']
        assert clients[0]['client_name'] == 'Updated Company'

    def test_delete_client(self, client, sample_client_data):
        create_resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client_id = json.loads(create_resp.data)['client']['id']

        resp = client.delete(f'/api/clients/{client_id}')
        assert resp.status_code == 200

        get_resp = client.get('/api/clients')
        assert json.loads(get_resp.data)['clients'] == []

    def test_delete_nonexistent_client(self, client):
        resp = client.delete('/api/clients/99999')
        assert resp.status_code == 404

    def test_clone_client(self, client, sample_client_data):
        create_resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client_id = json.loads(create_resp.data)['client']['id']

        resp = client.post(f'/api/clients/{client_id}/clone')
        assert resp.status_code == 201

        get_resp = client.get('/api/clients')
        clients = json.loads(get_resp.data)['clients']
        assert len(clients) == 2

    def test_client_with_multiple_contacts(self, client):
        data = {
            'tax_id': '99-1111111',
            'client_name': 'Multi Contact Co',
            'contacts': [
                {'contact_person': 'Alice', 'email': 'alice@test.com', 'sort_order': 0},
                {'contact_person': 'Bob', 'email': 'bob@test.com', 'sort_order': 1},
            ]
        }
        resp = client.post('/api/clients', data=json.dumps(data), content_type='application/json')
        assert resp.status_code == 201
        result = json.loads(resp.data)
        assert len(result['client']['contacts']) == 2


# ============================================================================
# INDIVIDUAL CRUD TESTS
# ============================================================================

class TestIndividualCRUD:
    """Tests for Individual CRUD operations."""

    def test_create_individual(self, client, sample_individual_data):
        resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        assert resp.status_code == 201
        data = json.loads(resp.data)
        assert data['individual']['first_name'] == 'John'
        assert data['individual']['last_name'] == 'Doe'
        assert data['individual']['individual_id'] is not None

    def test_get_individuals_empty(self, client):
        resp = client.get('/api/individuals')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['individuals'] == []

    def test_get_individuals_with_data(self, client, sample_individual_data):
        client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        resp = client.get('/api/individuals')
        data = json.loads(resp.data)
        assert len(data['individuals']) == 1
        assert data['individuals'][0]['first_name'] == 'John'

    def test_update_individual(self, client, sample_individual_data):
        create_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(create_resp.data)['individual']['id']

        update = {'first_name': 'Jane', 'status': 'Inactive'}
        resp = client.put(f'/api/individuals/{ind_id}', data=json.dumps(update), content_type='application/json')
        assert resp.status_code == 200

    def test_delete_individual(self, client, sample_individual_data):
        create_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(create_resp.data)['individual']['id']

        resp = client.delete(f'/api/individuals/{ind_id}')
        assert resp.status_code == 200

        get_resp = client.get('/api/individuals')
        assert json.loads(get_resp.data)['individuals'] == []

    def test_individual_auto_id_generation(self, client):
        """Individual IDs should auto-generate in format 00-#######."""
        resp1 = client.post('/api/individuals', data=json.dumps({'first_name': 'A'}), content_type='application/json')
        resp2 = client.post('/api/individuals', data=json.dumps({'first_name': 'B'}), content_type='application/json')

        id1 = json.loads(resp1.data)['individual']['individual_id']
        id2 = json.loads(resp2.data)['individual']['individual_id']

        assert id1.startswith('00-')
        assert id2.startswith('00-')
        assert id1 != id2


# ============================================================================
# EMPLOYEE BENEFITS CRUD TESTS
# ============================================================================

class TestBenefitsCRUD:
    """Tests for Employee Benefits CRUD operations."""

    def test_create_benefit(self, client, sample_client_data, sample_benefit_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        resp = client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        assert resp.status_code == 201

    def test_get_benefits_empty(self, client):
        resp = client.get('/api/benefits')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['benefits'] == []

    def test_get_benefits_with_data(self, client, sample_client_data, sample_benefit_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')

        resp = client.get('/api/benefits')
        data = json.loads(resp.data)
        assert len(data['benefits']) == 1
        assert data['benefits'][0]['tax_id'] == '12-3456789'

    def test_update_benefit(self, client, sample_client_data, sample_benefit_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        benefit_id = json.loads(create_resp.data)['benefit']['id']

        update = {'funding': 'Self-Funded', 'enrolled_ees': 30}
        resp = client.put(f'/api/benefits/{benefit_id}', data=json.dumps(update), content_type='application/json')
        assert resp.status_code == 200

    def test_delete_benefit(self, client, sample_client_data, sample_benefit_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        benefit_id = json.loads(create_resp.data)['benefit']['id']

        resp = client.delete(f'/api/benefits/{benefit_id}')
        assert resp.status_code == 200

    def test_clone_benefit(self, client, sample_client_data, sample_benefit_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        benefit_id = json.loads(create_resp.data)['benefit']['id']

        resp = client.post(f'/api/benefits/{benefit_id}/clone')
        assert resp.status_code == 201

        get_resp = client.get('/api/benefits')
        assert len(json.loads(get_resp.data)['benefits']) == 2


# ============================================================================
# COMMERCIAL INSURANCE CRUD TESTS
# ============================================================================

class TestCommercialCRUD:
    """Tests for Commercial Insurance CRUD operations."""

    def test_create_commercial(self, client, sample_client_data, sample_commercial_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        resp = client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        assert resp.status_code == 201

    def test_get_commercial(self, client, sample_client_data, sample_commercial_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')

        resp = client.get('/api/commercial')
        data = json.loads(resp.data)
        assert len(data['commercial']) == 1
        assert data['commercial'][0]['general_liability_carrier'] == 'Hartford'

    def test_update_commercial(self, client, sample_client_data, sample_commercial_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        comm_id = json.loads(create_resp.data)['commercial']['id']

        update = {'general_liability_carrier': 'Travelers', 'general_liability_premium': 6000.0}
        resp = client.put(f'/api/commercial/{comm_id}', data=json.dumps(update), content_type='application/json')
        assert resp.status_code == 200

    def test_delete_commercial(self, client, sample_client_data, sample_commercial_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        comm_id = json.loads(create_resp.data)['commercial']['id']

        resp = client.delete(f'/api/commercial/{comm_id}')
        assert resp.status_code == 200

    def test_clone_commercial(self, client, sample_client_data, sample_commercial_data):
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        create_resp = client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        comm_id = json.loads(create_resp.data)['commercial']['id']

        resp = client.post(f'/api/commercial/{comm_id}/clone')
        assert resp.status_code == 201

        get_resp = client.get('/api/commercial')
        assert len(json.loads(get_resp.data)['commercial']) == 2


# ============================================================================
# PERSONAL INSURANCE CRUD TESTS
# ============================================================================

class TestPersonalCRUD:
    """Tests for Personal Insurance CRUD operations."""

    def _create_individual_and_personal(self, client, sample_individual_data, sample_personal_data):
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind = json.loads(ind_resp.data)['individual']
        sample_personal_data['individual_id'] = ind['individual_id']
        return ind

    def test_create_personal(self, client, sample_individual_data, sample_personal_data):
        self._create_individual_and_personal(client, sample_individual_data, sample_personal_data)
        resp = client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')
        assert resp.status_code == 201

    def test_get_personal(self, client, sample_individual_data, sample_personal_data):
        self._create_individual_and_personal(client, sample_individual_data, sample_personal_data)
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        resp = client.get('/api/personal')
        data = json.loads(resp.data)
        assert len(data['personal']) == 1
        assert data['personal'][0]['personal_auto_carrier'] == 'State Farm'

    def test_update_personal(self, client, sample_individual_data, sample_personal_data):
        self._create_individual_and_personal(client, sample_individual_data, sample_personal_data)
        create_resp = client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')
        pers_id = json.loads(create_resp.data)['personal']['id']

        update = {'personal_auto_carrier': 'Geico', 'personal_auto_premium': 1000.0}
        resp = client.put(f'/api/personal/{pers_id}', data=json.dumps(update), content_type='application/json')
        assert resp.status_code == 200

    def test_delete_personal(self, client, sample_individual_data, sample_personal_data):
        self._create_individual_and_personal(client, sample_individual_data, sample_personal_data)
        create_resp = client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')
        pers_id = json.loads(create_resp.data)['personal']['id']

        resp = client.delete(f'/api/personal/{pers_id}')
        assert resp.status_code == 200

    def test_clone_personal(self, client, sample_individual_data, sample_personal_data):
        self._create_individual_and_personal(client, sample_individual_data, sample_personal_data)
        create_resp = client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')
        pers_id = json.loads(create_resp.data)['personal']['id']

        resp = client.post(f'/api/personal/{pers_id}/clone')
        assert resp.status_code == 201


# ============================================================================
# EXPORT TESTS
# ============================================================================

class TestExport:
    """Tests for the Excel export functionality."""

    def test_export_empty_database(self, client):
        """Export with no data should return a valid xlsx with correct sheets."""
        resp = client.get('/api/export')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.content_type

        wb = load_workbook(io.BytesIO(resp.data))
        assert 'Clients' in wb.sheetnames
        assert 'Individuals' in wb.sheetnames
        assert 'Employee Benefits' in wb.sheetnames
        assert 'Commercial' in wb.sheetnames
        assert 'Personal' in wb.sheetnames

    def test_export_clients_sheet(self, client, sample_client_data):
        """Exported Clients sheet should have correct headers and data."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')

        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Clients']

        # Check headers in row 2
        headers = [cell.value for cell in ws[2]]
        assert 'Tax ID' in headers
        assert 'Client Name' in headers
        assert 'Status' in headers

        # Check data in row 3
        assert ws.cell(row=3, column=1).value == '12-3456789'
        assert ws.cell(row=3, column=2).value == 'Test Company LLC'

    def test_export_individuals_sheet(self, client, sample_individual_data):
        """Exported Individuals sheet should have correct headers and data."""
        client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')

        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Individuals']

        # Check headers in row 2
        headers = [cell.value for cell in ws[2]]
        assert 'Individual ID' in headers
        assert 'First Name' in headers
        assert 'Last Name' in headers
        assert 'Status' in headers

        # Check data
        assert ws.cell(row=3, column=2).value == 'John'
        assert ws.cell(row=3, column=3).value == 'Doe'

    def test_export_benefits_sheet(self, client, sample_client_data, sample_benefit_data):
        """Exported Benefits sheet should include plan data."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')

        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Employee Benefits']

        # Row 3 should have the benefit
        assert ws.cell(row=3, column=1).value == '12-3456789'

    def test_export_commercial_sheet(self, client, sample_client_data, sample_commercial_data):
        """Exported Commercial sheet should include insurance data."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')

        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Commercial']

        assert ws.cell(row=3, column=1).value == '12-3456789'

    def test_export_personal_sheet(self, client, sample_individual_data, sample_personal_data):
        """Exported Personal sheet should include individual insurance data."""
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(ind_resp.data)['individual']['individual_id']
        sample_personal_data['individual_id'] = ind_id
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        resp = client.get('/api/export')
        wb = load_workbook(io.BytesIO(resp.data))
        ws = wb['Personal']

        assert ws.cell(row=3, column=1).value == ind_id


# ============================================================================
# IMPORT TESTS
# ============================================================================

class TestImport:
    """Tests for the Excel import functionality."""

    def _build_import_workbook(self, clients=None, individuals=None, benefits=None,
                                commercial=None, personal=None):
        """Build an xlsx workbook in memory for import testing."""
        wb = Workbook()

        # Clients sheet
        ws = wb.active
        ws.title = 'Clients'
        ws.append([None])  # Row 1: section headers (empty for basic)
        ws.append(['Tax ID', 'Client Name', 'DBA', 'Industry', 'Status', 'Gross Revenue', 'Total EEs',
                    'Contact Person', 'Email', 'Phone Number', 'Ext', 'Address Line 1', 'Address Line 2',
                    'City', 'State', 'Zip Code'])
        for row in (clients or []):
            ws.append(row)

        # Individuals sheet
        ws_ind = wb.create_sheet('Individuals')
        ws_ind.append([None])  # Row 1: empty
        ws_ind.append(['Individual ID', 'First Name', 'Last Name', 'Email', 'Phone Number',
                        'Address Line 1', 'Address Line 2', 'City', 'State', 'Zip Code', 'Status'])
        for row in (individuals or []):
            ws_ind.append(row)

        # Employee Benefits sheet
        ws_ben = wb.create_sheet('Employee Benefits')
        ws_ben.append([None])  # Row 1
        ws_ben.append(['Tax ID', 'Client Name ', 'Parent Client',
                        'Form Fire Code', 'Assigned To', 'Other Broker', 'Funding',
                        '# of Emp at renewal', 'Enrolled EEs', 'Waiting Period',
                        'Deductible Accumulation', 'Previous Carrier', 'Cobra Administrator',
                        'MEDICAL Carrier', 'MEDICAL Renewal Date', 'MEDICAL Waiting Period',
                        'MEDICAL Remarks', 'MEDICAL Outstanding Item',
                        'DENTAL Carrier', 'DENTAL Renewal Date', 'DENTAL Waiting Period',
                        'DENTAL Remarks', 'DENTAL Outstanding Item',
                        'VISION Carrier', 'VISION Renewal Date', 'VISION Waiting Period',
                        'VISION Remarks', 'VISION Outstanding Item',
                        'Life & AD&D Carrier', 'Life & AD&D Renewal Date', 'Life & AD&D Waiting Period',
                        'Life & AD&D Remarks', 'Life & AD&D Outstanding Item',
                        'LTD Renewal Date', 'LTD Carrier', 'LTD Remarks', 'LTD Outstanding Item',
                        'STD Renewal Date', 'STD Carrier', 'STD Remarks', 'STD Outstanding Item',
                        '401K Renewal Date', '401K Carrier', '401K Remarks', '401K Outstanding Item',
                        'Critical Illness Renewal Date', 'Critical Illness Carrier', 'Critical Illness Remarks', 'Critical Illness Outstanding Item',
                        'Accident Renewal Date', 'Accident Carrier', 'Accident Remarks', 'Accident Outstanding Item',
                        'Hospital Renewal Date', 'Hospital Carrier', 'Hospital Remarks', 'Hospital Outstanding Item',
                        'Voluntary Life Renewal Date', 'Voluntary Life Carrier', 'Voluntary Life Remarks', 'Voluntary Life Outstanding Item',
                        'Employer Contribution %', 'Employee Contribution %'])
        for row in (benefits or []):
            ws_ben.append(row)

        # Commercial sheet
        ws_comm = wb.create_sheet('Commercial')
        ws_comm.append(['', '', '', '',
                        'Commercial General Liability', None, None, None, None, None, None, None, None])  # Row 1 section headers
        ws_comm.append(['Tax ID', 'Client Name ', 'Parent Client', 'Assigned To',
                        'Carrier', 'Agency', 'Policy Number', 'Occ Limit (M)', 'Agg Limit (M)',
                        'Premium', 'Renewal Date', 'Remarks', 'Outstanding Item'])
        for row in (commercial or []):
            ws_comm.append(row)

        # Personal sheet
        ws_pers = wb.create_sheet('Personal')
        ws_pers.append(['', '',
                        'Personal Auto', None, None, None, None, None, None, None])
        ws_pers.append(['Individual ID', 'Individual Name',
                        'Carrier', 'BI Occ Limit', 'BI Agg Limit', 'PD Limit',
                        'Premium', 'Renewal Date', 'Outstanding Item', 'Remarks'])
        for row in (personal or []):
            ws_pers.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def test_import_empty_file(self, client):
        """Import with empty sheets should succeed with zero counts."""
        xlsx = self._build_import_workbook()
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['stats']['clients_created'] == 0
        assert data['stats']['individuals_created'] == 0

    def test_import_no_file(self, client):
        """Import with no file should return 400."""
        resp = client.post('/api/import', content_type='multipart/form-data')
        assert resp.status_code == 400

    def test_import_wrong_format(self, client):
        """Import with non-xlsx file should return 400."""
        resp = client.post('/api/import',
                          data={'file': (io.BytesIO(b'not excel'), 'test.csv')},
                          content_type='multipart/form-data')
        assert resp.status_code == 400

    def test_import_clients(self, client):
        """Import should create clients from xlsx data."""
        xlsx = self._build_import_workbook(
            clients=[
                ['11-1111111', 'Company A', 'DBA A', 'Tech', 'Active', 500000, 25,
                 'Alice', 'alice@a.com', '555-0001', None, '1 Main', None, 'NYC', 'NY', 10001],
                ['22-2222222', 'Company B', None, 'Finance', 'Active', 1000000, 50,
                 'Bob', 'bob@b.com', '555-0002', None, '2 Main', None, 'LA', 'CA', 90001],
            ]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['clients_created'] == 2

        # Verify in database
        get_resp = client.get('/api/clients')
        clients_data = json.loads(get_resp.data)['clients']
        assert len(clients_data) == 2

    def test_import_individuals(self, client):
        """Import should create individuals from xlsx data."""
        xlsx = self._build_import_workbook(
            individuals=[
                ['00-0000001', 'John', 'Doe', 'john@test.com', '555-0001',
                 '123 Main', None, 'NYC', 'NY', 10001, 'Active'],
                ['00-0000002', 'Jane', 'Smith', 'jane@test.com', '555-0002',
                 '456 Oak', None, 'LA', 'CA', 90001, 'Active'],
            ]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['individuals_created'] == 2

        get_resp = client.get('/api/individuals')
        individuals = json.loads(get_resp.data)['individuals']
        assert len(individuals) == 2
        assert individuals[0]['individual_id'] in ['00-0000001', '00-0000002']

    def test_import_preserves_individual_id(self, client):
        """Import should use the individual_id from the xlsx, not auto-generate."""
        xlsx = self._build_import_workbook(
            individuals=[['00-9999999', 'Test', 'User', None, None, None, None, None, None, None, 'Active']]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        assert json.loads(resp.data)['stats']['individuals_created'] == 1

        get_resp = client.get('/api/individuals')
        individuals = json.loads(get_resp.data)['individuals']
        assert individuals[0]['individual_id'] == '00-9999999'

    def test_import_deletes_existing_data(self, client, sample_client_data, sample_individual_data):
        """Import should delete all existing data before importing."""
        # Create existing data
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')

        # Verify data exists
        assert len(json.loads(client.get('/api/clients').data)['clients']) == 1
        assert len(json.loads(client.get('/api/individuals').data)['individuals']) == 1

        # Import empty file - should wipe everything
        xlsx = self._build_import_workbook()
        client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')

        # Verify all deleted
        assert json.loads(client.get('/api/clients').data)['clients'] == []
        assert json.loads(client.get('/api/individuals').data)['individuals'] == []

    def test_import_benefits_with_plans(self, client):
        """Import should create benefits with multi-plan child records."""
        xlsx = self._build_import_workbook(
            clients=[['11-1111111', 'Company A', None, None, 'Active', None, None,
                      'Alice', 'alice@a.com', None, None, None, None, None, None, None]],
            benefits=[['11-1111111', 'Company A', None, 'FF100', 'John', None, 'Fully Insured',
                       50, 45, '30 days', 'Calendar Year', 'Aetna', 'WageWorks',
                       'BlueCross', '2025-06-01', '30 days', None, None,
                       'Delta Dental', '2025-07-01', None, None, None,
                       None, None, None, None, None,
                       None, None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       '80%', '20%']]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['clients_created'] == 1
        assert data['stats']['benefits_created'] == 1

    def test_import_commercial(self, client):
        """Import should create commercial insurance records."""
        xlsx = self._build_import_workbook(
            clients=[['11-1111111', 'Company A', None, None, 'Active', None, None,
                      'Alice', 'alice@a.com', None, None, None, None, None, None, None]],
            commercial=[['11-1111111', 'Company A', None, 'Agent Smith',
                         'Hartford', 'ABC Agency', 'POL-001', '1', '2', 5000, '2025-06-01', None, None]]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['commercial_created'] == 1

    def test_import_personal_requires_individual(self, client):
        """Import personal should fail if individual not found."""
        xlsx = self._build_import_workbook(
            personal=[['00-9999999', 'Ghost Person', 'State Farm', '100/300', '300/300', '50', 1200, '2025-07-01', None, None]]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['personal_created'] == 0
        assert len(data['stats']['errors']) > 0

    def test_import_personal_with_individual(self, client):
        """Import personal should succeed when individual exists in the same import."""
        xlsx = self._build_import_workbook(
            individuals=[['00-0000001', 'John', 'Doe', None, None, None, None, None, None, None, 'Active']],
            personal=[['00-0000001', 'John Doe', 'State Farm', '100000', '300000', '50000', 1200, '2025-07-01', None, None]]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['individuals_created'] == 1
        assert data['stats']['personal_created'] == 1

    def test_import_replaces_all_data(self, client):
        """Second import should completely replace data from first import."""
        # First import: 2 clients
        xlsx1 = self._build_import_workbook(
            clients=[
                ['11-1111111', 'Company A', None, None, 'Active', None, None,
                 'Alice', 'alice@a.com', None, None, None, None, None, None, None],
                ['22-2222222', 'Company B', None, None, 'Active', None, None,
                 'Bob', 'bob@b.com', None, None, None, None, None, None, None],
            ]
        )
        client.post('/api/import', data={'file': (xlsx1, 'test.xlsx')}, content_type='multipart/form-data')
        assert len(json.loads(client.get('/api/clients').data)['clients']) == 2

        # Second import: 1 different client
        xlsx2 = self._build_import_workbook(
            clients=[['33-3333333', 'Company C', None, None, 'Active', None, None,
                      'Carol', 'carol@c.com', None, None, None, None, None, None, None]]
        )
        client.post('/api/import', data={'file': (xlsx2, 'test.xlsx')}, content_type='multipart/form-data')

        clients_data = json.loads(client.get('/api/clients').data)['clients']
        assert len(clients_data) == 1
        assert clients_data[0]['tax_id'] == '33-3333333'

    def test_import_error_rows_returned(self, client):
        """Import errors should be tracked and returned in stats."""
        xlsx = self._build_import_workbook(
            # Benefits row referencing non-existent client
            benefits=[['99-9999999', 'Ghost Corp', None, None, None, None, None,
                       None, None, None, None, None, None,
                       None, None, None, None, None,
                       None, None, None, None, None,
                       None, None, None, None, None,
                       None, None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None, None, None,
                       None, None]]
        )
        resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        data = json.loads(resp.data)
        assert data['stats']['benefits_created'] == 0
        assert len(data['stats']['errors']) > 0


# ============================================================================
# EXPORT -> IMPORT ROUNDTRIP TESTS
# ============================================================================

class TestExportImportRoundtrip:
    """Tests that data survives an export->import cycle intact."""

    def test_clients_roundtrip(self, client, sample_client_data):
        """Clients should survive export->import roundtrip."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import (this deletes and re-creates)
        client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')

        # Verify
        get_resp = client.get('/api/clients')
        clients_data = json.loads(get_resp.data)['clients']
        assert len(clients_data) == 1
        assert clients_data[0]['tax_id'] == '12-3456789'
        assert clients_data[0]['client_name'] == 'Test Company LLC'

    def test_individuals_roundtrip(self, client, sample_individual_data):
        """Individuals should survive export->import roundtrip."""
        create_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        original_ind_id = json.loads(create_resp.data)['individual']['individual_id']

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import
        client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')

        # Verify
        get_resp = client.get('/api/individuals')
        individuals = json.loads(get_resp.data)['individuals']
        assert len(individuals) == 1
        assert individuals[0]['individual_id'] == original_ind_id
        assert individuals[0]['first_name'] == 'John'
        assert individuals[0]['last_name'] == 'Doe'

    def test_benefits_roundtrip(self, client, sample_client_data, sample_benefit_data):
        """Benefits with plans should survive export->import roundtrip."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['clients_created'] == 1
        assert stats['benefits_created'] == 1

        # Verify benefit data
        get_resp = client.get('/api/benefits')
        benefits = json.loads(get_resp.data)['benefits']
        assert len(benefits) == 1
        assert benefits[0]['tax_id'] == '12-3456789'
        assert benefits[0]['funding'] == 'Fully Insured'

    def test_commercial_roundtrip(self, client, sample_client_data, sample_commercial_data):
        """Commercial insurance should survive export->import roundtrip."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['commercial_created'] == 1

    def test_personal_roundtrip(self, client, sample_individual_data, sample_personal_data):
        """Personal insurance should survive export->import roundtrip."""
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(ind_resp.data)['individual']['individual_id']
        sample_personal_data['individual_id'] = ind_id
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']
        assert stats['individuals_created'] == 1
        assert stats['personal_created'] == 1

    def test_full_roundtrip_all_entities(self, client, sample_client_data, sample_individual_data,
                                          sample_benefit_data, sample_commercial_data, sample_personal_data):
        """All entity types should survive a full export->import cycle."""
        # Create all entity types
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(ind_resp.data)['individual']['individual_id']
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        sample_personal_data['individual_id'] = ind_id
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        # Export
        export_resp = client.get('/api/export')
        xlsx = io.BytesIO(export_resp.data)

        # Import
        import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
        stats = json.loads(import_resp.data)['stats']

        assert stats['clients_created'] == 1
        assert stats['individuals_created'] == 1
        assert stats['benefits_created'] == 1
        assert stats['commercial_created'] == 1
        assert stats['personal_created'] == 1
        assert stats['errors'] == []

    def test_double_roundtrip(self, client, sample_client_data, sample_individual_data):
        """Data should survive two consecutive export->import cycles."""
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')

        for _ in range(2):
            export_resp = client.get('/api/export')
            xlsx = io.BytesIO(export_resp.data)
            import_resp = client.post('/api/import', data={'file': (xlsx, 'test.xlsx')}, content_type='multipart/form-data')
            stats = json.loads(import_resp.data)['stats']
            assert stats['clients_created'] == 1
            assert stats['individuals_created'] == 1
            assert stats['errors'] == []


# ============================================================================
# DASHBOARD TESTS
# ============================================================================

class TestDashboard:
    """Tests for dashboard endpoints."""

    def test_renewals_empty(self, client):
        resp = client.get('/api/dashboard/renewals')
        assert resp.status_code == 200

    def test_cross_sell_empty(self, client):
        resp = client.get('/api/dashboard/cross-sell')
        assert resp.status_code == 200


# ============================================================================
# HEALTH CHECK
# ============================================================================

class TestHealth:
    """Tests for health check endpoint."""

    def test_health_check(self, client):
        resp = client.get('/api/health')
        assert resp.status_code == 200


# ============================================================================
# FEEDBACK CRUD TESTS
# ============================================================================

class TestFeedbackCRUD:
    """Tests for Feedback CRUD operations."""

    def test_create_feedback(self, client):
        data = {'type': 'Bug', 'subject': 'Test bug', 'description': 'Something broke'}
        resp = client.post('/api/feedback', data=json.dumps(data), content_type='application/json')
        assert resp.status_code == 201

    def test_get_feedback(self, client):
        data = {'type': 'Feature', 'subject': 'New feature', 'description': 'Want this'}
        client.post('/api/feedback', data=json.dumps(data), content_type='application/json')

        resp = client.get('/api/feedback')
        assert resp.status_code == 200
        feedback = json.loads(resp.data)['feedback']
        assert len(feedback) == 1

    def test_delete_feedback(self, client):
        data = {'type': 'Bug', 'subject': 'Delete me', 'description': 'Gone'}
        create_resp = client.post('/api/feedback', data=json.dumps(data), content_type='application/json')
        fb_id = json.loads(create_resp.data)['id']

        resp = client.delete(f'/api/feedback/{fb_id}')
        assert resp.status_code == 200


# ============================================================================
# DATA INTEGRITY TESTS
# ============================================================================

class TestDataIntegrity:
    """Tests for data integrity and cascade behavior."""

    def test_delete_client_cascades_benefits(self, client, sample_client_data, sample_benefit_data):
        """Deleting a client should cascade-delete associated benefits."""
        create_resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client_id = json.loads(create_resp.data)['client']['id']
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')

        # Verify benefit exists
        assert len(json.loads(client.get('/api/benefits').data)['benefits']) == 1

        # Delete client
        client.delete(f'/api/clients/{client_id}')

        # Benefits should be gone
        assert json.loads(client.get('/api/benefits').data)['benefits'] == []

    def test_delete_client_cascades_commercial(self, client, sample_client_data, sample_commercial_data):
        """Deleting a client should cascade-delete associated commercial insurance."""
        create_resp = client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client_id = json.loads(create_resp.data)['client']['id']
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')

        assert len(json.loads(client.get('/api/commercial').data)['commercial']) == 1

        client.delete(f'/api/clients/{client_id}')

        assert json.loads(client.get('/api/commercial').data)['commercial'] == []

    def test_delete_individual_cascades_personal(self, client, sample_individual_data, sample_personal_data):
        """Deleting an individual should cascade-delete associated personal insurance."""
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind = json.loads(ind_resp.data)['individual']
        sample_personal_data['individual_id'] = ind['individual_id']
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        assert len(json.loads(client.get('/api/personal').data)['personal']) == 1

        client.delete(f'/api/individuals/{ind["id"]}')

        assert json.loads(client.get('/api/personal').data)['personal'] == []

    def test_import_clears_all_related_data(self, client, sample_client_data, sample_benefit_data,
                                             sample_commercial_data, sample_individual_data, sample_personal_data):
        """Import should delete ALL related data including child records."""
        # Create full data set
        client.post('/api/clients', data=json.dumps(sample_client_data), content_type='application/json')
        client.post('/api/benefits', data=json.dumps(sample_benefit_data), content_type='application/json')
        client.post('/api/commercial', data=json.dumps(sample_commercial_data), content_type='application/json')
        ind_resp = client.post('/api/individuals', data=json.dumps(sample_individual_data), content_type='application/json')
        ind_id = json.loads(ind_resp.data)['individual']['individual_id']
        sample_personal_data['individual_id'] = ind_id
        client.post('/api/personal', data=json.dumps(sample_personal_data), content_type='application/json')

        # Import empty file
        wb = Workbook()
        ws = wb.active
        ws.title = 'Clients'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        client.post('/api/import', data={'file': (output, 'empty.xlsx')}, content_type='multipart/form-data')

        # Verify everything is deleted
        assert json.loads(client.get('/api/clients').data)['clients'] == []
        assert json.loads(client.get('/api/individuals').data)['individuals'] == []
        assert json.loads(client.get('/api/benefits').data)['benefits'] == []
        assert json.loads(client.get('/api/commercial').data)['commercial'] == []
        assert json.loads(client.get('/api/personal').data)['personal'] == []
